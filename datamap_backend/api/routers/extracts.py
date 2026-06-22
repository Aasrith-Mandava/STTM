import asyncio
import logging
from typing import Any, Dict, List, Optional
import io
import json
import os
import re
from pydantic import BaseModel, Field
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from google.genai import types

from google.adk import Runner
from google.adk.apps import App
from utils.adk_runtime import VertexAiSessionService

# Attempting to import artifact service from adk if present
from google.adk.artifacts import GcsArtifactService

from config.settings import config
from agents.extract_agent.agents import parse_parallel_agent, ambiguity_detector_agent
from utils.extract_parser_utils import (
    parse_docx_brd,
    parse_xlsx_layout,
    extract_decisions_from_transcript,
)
from utils.llm_rate_utils import (
    record_llm_usage_and_get_wait,
    wait_for_llm_request_slot,
    is_resource_exhausted_error,
    is_transient_llm_transport_error,
    calculate_retry_delay,
    calculate_mapping_retry_delay,
)

logger = logging.getLogger(__name__)

router = APIRouter()


def _detect_mime_type(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return {
        "pdf": "application/pdf",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "txt": "text/plain",
        "md": "text/plain",
    }.get(ext, "application/octet-stream")


async def _update_step_status(
    session_service, app_name, user_id, session_id, step, status
):
    await session_service.update_session(
        app_name=app_name,
        user_id=user_id,
        session_id=session_id,
        state={"step_status": {step: status}},
    )


@router.post("/session")
async def create_extracts_session(
    session_id: str = Form(...),
    user_id: str = Form(...),
    app_name: str = Form(default=config.REASONING_ENGINE_RESOURCE),
    brd_file: UploadFile = File(...),
    layout_files: List[UploadFile] = File(...),
    transcript_files: Optional[List[UploadFile]] = File(default=[]),
    brd_section_reference: Optional[str] = Form(default=None),
    standards_version: Optional[str] = Form(default=None),
):
    try:
        session_service = VertexAiSessionService(
            project=config.GOOGLE_CLOUD_PROJECT, location=config.GOOGLE_CLOUD_LOCATION
        )
        # Using STAGING_BUCKET if GcsArtifactService requires it, but passing standard initialization
        artifact_service = GcsArtifactService(
            project=config.GOOGLE_CLOUD_PROJECT, location=config.GOOGLE_CLOUD_LOCATION
        )

        # 1. Read all file bytes
        brd_bytes = await brd_file.read()
        layout_bytes_list = [await f.read() for f in layout_files]
        transcript_bytes_list = (
            [await f.read() for f in transcript_files] if transcript_files else []
        )

        # 3. Call session_service.create_session
        initial_state = {
            "standards_version": standards_version,
            "brd_section_reference": brd_section_reference,
            "step_status": {},
            "artifact_keys": {},
        }
        await session_service.create_session(
            app_name=app_name,
            user_id=user_id,
            session_id=session_id,
            state=initial_state,
        )

        artifact_keys = {}

        # 4. Save artifacts
        # BRD
        brd_key = "brd_0" + (
            "" if not brd_file.filename else "." + brd_file.filename.rsplit(".", 1)[-1]
        )
        brd_mime = _detect_mime_type(brd_file.filename)
        await artifact_service.save_artifact(
            app_name=app_name,
            user_id=user_id,
            session_id=session_id,
            filename=brd_key,
            artifact=types.Part.from_bytes(data=brd_bytes, mime_type=brd_mime),
        )
        artifact_keys["brd"] = brd_key

        # Layouts
        artifact_keys["layouts"] = []
        for idx, (layout_file, l_bytes) in enumerate(
            zip(layout_files, layout_bytes_list)
        ):
            l_key = f"layout_{idx}" + (
                ""
                if not layout_file.filename
                else "." + layout_file.filename.rsplit(".", 1)[-1]
            )
            l_mime = _detect_mime_type(layout_file.filename)
            await artifact_service.save_artifact(
                app_name=app_name,
                user_id=user_id,
                session_id=session_id,
                filename=l_key,
                artifact=types.Part.from_bytes(data=l_bytes, mime_type=l_mime),
            )
            artifact_keys["layouts"].append(l_key)

        # Transcripts
        artifact_keys["transcripts"] = []
        for idx, (trans_file, t_bytes) in enumerate(
            zip(transcript_files, transcript_bytes_list)
        ):
            t_key = f"transcript_{idx}" + (
                ""
                if not trans_file.filename
                else "." + trans_file.filename.rsplit(".", 1)[-1]
            )
            t_mime = _detect_mime_type(trans_file.filename)
            await artifact_service.save_artifact(
                app_name=app_name,
                user_id=user_id,
                session_id=session_id,
                filename=t_key,
                artifact=types.Part.from_bytes(data=t_bytes, mime_type=t_mime),
            )
            artifact_keys["transcripts"].append(t_key)

        # 5. Call update_session to store artifact_keys
        await session_service.update_session(
            app_name=app_name,
            user_id=user_id,
            session_id=session_id,
            state={"artifact_keys": artifact_keys},
        )

        # 6. Return response
        return {
            "session_id": session_id,
            "artifact_keys": artifact_keys,
            "status": "created",
        }

    except Exception as e:
        logger.error(f"Error in create_extracts_session: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{session_id}/parse")
async def parse_extract_session(session_id: str, payload: dict):
    try:
        user_id = payload.get("user_id")
        app_name = payload.get("app_name", config.REASONING_ENGINE_RESOURCE)

        session_service = VertexAiSessionService(
            project=config.GOOGLE_CLOUD_PROJECT, location=config.GOOGLE_CLOUD_LOCATION
        )
        artifact_service = GcsArtifactService(
            project=config.GOOGLE_CLOUD_PROJECT, location=config.GOOGLE_CLOUD_LOCATION
        )

        # 1. Load session state
        session = await session_service.get_session(
            app_name=app_name, user_id=user_id, session_id=session_id
        )
        state = session.state
        artifact_keys = state.get("artifact_keys", {})
        brd_section_reference = state.get("brd_section_reference")

        parts = []

        # 2. Load BRD artifact
        if "brd" in artifact_keys:
            brd_key = artifact_keys["brd"]
            brd_artifact = await artifact_service.load_artifact(
                app_name=app_name,
                user_id=user_id,
                session_id=session_id,
                filename=brd_key,
            )
            if _detect_mime_type(brd_key) == "application/pdf":
                parts.append(brd_artifact)
            elif (
                _detect_mime_type(brd_key)
                == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            ):
                # docx
                extracted_text = parse_docx_brd(
                    brd_artifact.inline_data.data, brd_section_reference
                )
                parts.append(
                    types.Part.from_text(text=f"BRD CONTENT:\n{extracted_text}")
                )

        parts.append(
            types.Part.from_text(
                text="Please parse the BRD content following the BRD_PARSER_INSTRUCTION."
            )
        )

        # 3. Load layout artifacts
        for l_key in artifact_keys.get("layouts", []):
            l_artifact = await artifact_service.load_artifact(
                app_name=app_name,
                user_id=user_id,
                session_id=session_id,
                filename=l_key,
            )
            if (
                _detect_mime_type(l_key)
                == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ):
                # xlsx
                extracted_json = parse_xlsx_layout(l_artifact.inline_data.data)
                parts.append(
                    types.Part.from_text(
                        text=f"FILE LAYOUT ({l_key}):\n{extracted_json}"
                    )
                )
            elif _detect_mime_type(l_key) == "application/pdf":
                parts.append(l_artifact)

        # 4. Load transcript artifacts
        for t_key in artifact_keys.get("transcripts", []):
            t_artifact = await artifact_service.load_artifact(
                app_name=app_name,
                user_id=user_id,
                session_id=session_id,
                filename=t_key,
            )
            extracted_text = extract_decisions_from_transcript(
                t_artifact.inline_data.data.decode("utf-8")
            )
            parts.append(
                types.Part.from_text(text=f"TRANSCRIPT ({t_key}):\n{extracted_text}")
            )

        parts.append(
            types.Part.from_text(
                text="Please structure and analyze all the provided layout and transcript inputs according to your instructions."
            )
        )

        # 5. Assemble content
        message = types.Content(role="user", parts=parts)

        # 6. Run parallel_runner
        parallel_app = App(name=app_name, root_agent=parse_parallel_agent)
        parallel_runner = Runner(app=parallel_app, session_service=session_service)

        max_retries = 3
        for attempt in range(max_retries):
            try:
                await wait_for_llm_request_slot(f"parse_parallel:{session_id}")
                async for event in parallel_runner.run_async(
                    user_id=user_id, session_id=session_id, new_message=message
                ):
                    await record_llm_usage_and_get_wait(
                        event,
                        session_id=f"parse_parallel:{session_id}",
                        buffer_tokens=300,
                    )
                break
            except Exception as exc:
                if is_resource_exhausted_error(exc) and attempt < max_retries - 1:
                    delay = calculate_retry_delay(attempt)
                    logger.warning("parse_parallel agent 429 Resource Exhausted (attempt %d/%d). Retrying in %.2fs...", attempt+1, max_retries, delay)
                    await asyncio.sleep(delay)
                else: raise

        # 7. Reload session state to read parsed outputs
        session = await session_service.get_session(
            app_name=app_name, user_id=user_id, session_id=session_id
        )
        state = session.state

        parsed_brd = state.get("parsed_brd", {})
        parsed_layouts = state.get("parsed_layouts", [])
        parsed_transcript = state.get("parsed_transcript", {})

        # 8. Build summary types.Content
        summary_text = f"Parsed BRD: {json.dumps(parsed_brd)}\nParsed Layouts: {json.dumps(parsed_layouts)}\nParsed Transcript: {json.dumps(parsed_transcript)}"
        summary_message = types.Content(
            role="user", parts=[types.Part.from_text(text=summary_text)]
        )

        # 9. Run ambiguity_runner
        ambiguity_app = App(name=app_name, root_agent=ambiguity_detector_agent)
        ambiguity_runner = Runner(app=ambiguity_app, session_service=session_service)

        for attempt in range(max_retries):
            try:
                await wait_for_llm_request_slot(f"ambiguity:{session_id}")
                async for event in ambiguity_runner.run_async(
                    user_id=user_id, session_id=session_id, new_message=summary_message
                ):
                    await record_llm_usage_and_get_wait(
                        event,
                        session_id=f"ambiguity:{session_id}",
                        buffer_tokens=300,
                    )
                break
            except Exception as exc:
                if is_resource_exhausted_error(exc) and attempt < max_retries - 1:
                    delay = calculate_retry_delay(attempt)
                    logger.warning("ambiguity agent 429 Resource Exhausted (attempt %d/%d). Retrying in %.2fs...", attempt+1, max_retries, delay)
                    await asyncio.sleep(delay)
                else: raise

        # 10. Call _update_step_status
        await _update_step_status(
            session_service, app_name, user_id, session_id, "parse", "draft_ready"
        )

        # 11. Return
        return {"status": "draft_ready", "session_id": session_id}

    except Exception as e:
        logger.error(f"Error in parse_extract_session: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{session_id}/parse")
async def get_parse_status(
    session_id: str, user_id: str, app_name: str = config.REASONING_ENGINE_RESOURCE
):
    try:
        session_service = VertexAiSessionService(
            project=config.GOOGLE_CLOUD_PROJECT, location=config.GOOGLE_CLOUD_LOCATION
        )
        session = await session_service.get_session(
            app_name=app_name, user_id=user_id, session_id=session_id
        )
        state = session.state

        step_status = state.get("step_status", {}).get("parse", "unknown")
        ambiguity_report = state.get("ambiguity_report", {})
        can_proceed = ambiguity_report.get("can_proceed", False)

        return {
            "step": "parse",
            "status": step_status,
            "can_approve": can_proceed,
            "parsed_brd": state.get("parsed_brd", {}),
            "parsed_layouts": state.get("parsed_layouts", []),
            "parsed_transcript": state.get("parsed_transcript", {}),
            "domain_tagged_fields": state.get("domain_tagged_fields", {}),
            "ambiguity_report": ambiguity_report,
        }

    except Exception as e:
        logger.error(f"Error in get_parse_status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{session_id}/parse/approve")
async def approve_parse(session_id: str, payload: dict):
    try:
        user_id = payload.get("user_id")
        app_name = payload.get("app_name", config.REASONING_ENGINE_RESOURCE)

        session_service = VertexAiSessionService(
            project=config.GOOGLE_CLOUD_PROJECT, location=config.GOOGLE_CLOUD_LOCATION
        )
        session = await session_service.get_session(
            app_name=app_name, user_id=user_id, session_id=session_id
        )
        state = session.state

        if state.get("step_status", {}).get("parse") != "draft_ready":
            raise HTTPException(
                status_code=400, detail="Step is not in draft_ready state."
            )

        # Apply BSA overrides
        parsed_brd = state.get("parsed_brd", {})
        parsed_layouts = state.get("parsed_layouts", [])
        domain_tagged_fields = state.get("domain_tagged_fields", {})
        ambiguity_report = state.get("ambiguity_report", {})

        # Merge scope_additions, scope_removals, etc (naive merge for this example)
        if "in_scope_items" in parsed_brd:
            parsed_brd["in_scope_items"].extend(payload.get("scope_additions", []))
            for item in payload.get("scope_removals", []):
                if item in parsed_brd["in_scope_items"]:
                    parsed_brd["in_scope_items"].remove(item)

        # Snapshot
        from datetime import datetime

        approved_snapshot = {
            "parsed_brd": parsed_brd,
            "parsed_layouts": parsed_layouts,
            "domain_tagged_fields": domain_tagged_fields,
            "ambiguity_report": ambiguity_report,
            "approved_at": datetime.utcnow().isoformat(),
        }

        # Save to state
        await session_service.update_session(
            app_name=app_name,
            user_id=user_id,
            session_id=session_id,
            state={
                "parsed_brd": parsed_brd,
                "parsed_layouts": parsed_layouts,
                "domain_tagged_fields": domain_tagged_fields,
                "ambiguity_report": ambiguity_report,
                "approved_parsed_requirements": approved_snapshot,
                "step_status": {"parse": "bsa_approved"},
            },
        )

        return {
            "status": "bsa_approved",
            "message": "Step 2 (File Identification) is now available.",
        }

    except Exception as e:
        logger.error(f"Error in approve_parse: {e}")
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))


# ─── Doc Extraction (BRD / Transcript / File Layout / Validation) ───────────

from pathlib import Path as _Path
from typing import Optional as _Optional
import uuid as _uuid

from fastapi import File as _File, Form as _Form, UploadFile as _UploadFile
from fastapi.concurrency import run_in_threadpool as _run_in_threadpool

from api.models import (
    BrdExtractionResponse,
    BrdAcceptRequest,
    BrdRejectFreeformRequest,
    BrdCheckpointResponse,
    DocExtractionResponse,
    FileLayoutExtractionResponse,
    FileLayoutCheckpointRequest,
    FileLayoutCheckpointResponse,
    ValidationResultResponse,
    MappingApproverRequest,
    MappingApproverResponse,
    MappingFieldCheckpointRequest,
    MappingFieldCheckpointResponse,
)
from agents.extract_requirement_layer_agent.agents.brd_transcript_agent import (
    run_brd_extraction,
    run_validation,
    run_brd_checkpoint,
)
from agents.extract_requirement_layer_agent.agents.file_layout_agent import (
    run_file_layout_extraction,
)
from utils.gcs_artifact_utils import (
    upload_bytes as _upload_bytes,
    upload_json as _upload_json,
    upload_text as _upload_text,
    delete_folder as _delete_folder,
)

_TMP_DIR = _Path(config.DATA_DIR) / "tmp_doc_extraction"
_TMP_DIR.mkdir(parents=True, exist_ok=True)


def _read_upload(upload: _UploadFile) -> tuple[_Path, bytes]:
    raw = upload.file.read()
    safe_name = _Path(upload.filename).name.replace(" ", "_")
    dest = _TMP_DIR / f"{_uuid.uuid4().hex}_{safe_name}"
    with open(dest, "wb") as f:
        f.write(raw)
    return dest, raw


def _content_type(filename: str) -> str:
    return (
        "application/pdf"
        if filename.lower().endswith(".pdf")
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


@router.post("/upload-extract", response_model=DocExtractionResponse)
async def upload_extract(
    brd_file: _UploadFile = _File(..., description="BRD document (PDF or DOCX)"),
    user_id: str = _Form(..., description="user id"),
    interface_code: str = _Form(..., description="Interface code"),
    file_layout: _UploadFile = _File(
        ..., description="File layout document (PDF,DOCX or EXCEL)"
    ),
    session_id: str = _Form(..., description="Session ID"),
    transcript: _Optional[_UploadFile] = _File(
        None, description="Transcript document (PDF or DOCX), optional"
    ),
    bsa_notes: _Optional[str] = _Form(None, description="BSA text notes, optional"),
):
    tmp_paths: list[_Path] = []
    try:
        prefix = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/uploads"
        deleted_count = await _run_in_threadpool(_delete_folder, prefix=prefix)
        if deleted_count > 0:
            logger.info("Cleaned %d existing artifacts for session=%s", deleted_count, session_id)

        for upload in filter(None, [brd_file, transcript]):
            if _Path(upload.filename).suffix.lower() not in {".pdf", ".docx"}:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unsupported file type for '{upload.filename}': "
                    f"{_Path(upload.filename).suffix.lower()!r}",
                )
        for upload in filter(None, [file_layout]):
            if _Path(upload.filename).suffix.lower() not in {".pdf", ".docx", ".xlsx"}:
                raise HTTPException(
                    status_code=400,
                    detail=f"Unsupported file type for '{upload.filename}': "
                           f"{_Path(upload.filename).suffix.lower()!r}",
                )
        brd_tmp, brd_bytes = _read_upload(brd_file)
        tmp_paths.append(brd_tmp)
        layout_tmp, layout_bytes = _read_upload(file_layout)
        tmp_paths.append(layout_tmp)

        transcript_bytes: _Optional[bytes] = None
        if transcript is not None:
            transcript_tmp, transcript_bytes = _read_upload(transcript)
            tmp_paths.append(transcript_tmp)

        artifacts_uploaded: list[str] = []

        brd_object = f"{prefix}/brd_{_Path(brd_file.filename).name}"
        await _run_in_threadpool(
            _upload_bytes,
            object_name=brd_object,
            content=brd_bytes,
            content_type=_content_type(brd_file.filename),
        )
        artifacts_uploaded.append(brd_object)

        layout_object = f"{prefix}/file_layout_{_Path(file_layout.filename).name}"
        await _run_in_threadpool(
            _upload_bytes,
            object_name=layout_object,
            content=layout_bytes,
            content_type=_content_type(file_layout.filename),
        )
        artifacts_uploaded.append(layout_object)

        if transcript is not None and transcript_bytes is not None:
            transcript_object = f"{prefix}/transcript_{_Path(transcript.filename).name}"
            await _run_in_threadpool(
                _upload_bytes,
                object_name=transcript_object,
                content=transcript_bytes,
                content_type=_content_type(transcript.filename),
            )
            artifacts_uploaded.append(transcript_object)

        result_object = f"{prefix}/pipeline_result.json"
        await _run_in_threadpool(
            _upload_json,
            object_name=result_object,
            payload={
                "document_path": brd_file.filename,
                "status": "uploaded",
                "interface_code": interface_code,
            },
        )
        artifacts_uploaded.append(result_object)

        if bsa_notes:
            notes_object = f"{prefix}/bsa_notes.txt"
            await _run_in_threadpool(
                _upload_text, object_name=notes_object, content=bsa_notes
            )
            artifacts_uploaded.append(notes_object)

        if interface_code:
            interface_code_object = f"{prefix}/interface_code.txt"
            await _run_in_threadpool(
                _upload_text, object_name=interface_code_object, content=interface_code
            )
            artifacts_uploaded.append(interface_code_object)

        return DocExtractionResponse(
            success=True,
            session_id=session_id,
            message="BRD extraction complete. All artifacts uploaded to GCS.",
            gcs_prefix=f"gs://{config.MAPPING_ARTIFACT_BUCKET}/{prefix}",
            artifacts_uploaded=artifacts_uploaded,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("BRD extraction pipeline failed for session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        for p in tmp_paths:
            if p.exists():
                p.unlink()


@router.get(
    "/extract-brd-information/{session_id}", response_model=BrdExtractionResponse
)
async def extract_brd_information(session_id: str):
    try:
        extraction = await _run_in_threadpool(run_brd_extraction, session_id)
        return BrdExtractionResponse(
            success=True,
            session_id=session_id,
            message="BRD information extracted successfully.",
            artifacts_found=extraction.get("artifacts_found", []),
            brd_filename=extraction.get("brd_filename"),
            file_layout_filename=extraction.get("file_layout_filename"),
            transcript_filename=extraction.get("transcript_filename"),
            bsa_notes=extraction.get("bsa_notes"),
            markdown_uploads=extraction.get("markdown_uploads", []),
            requirement_layer=extraction.get("requirement_layer"),
            gcs_output_uri=extraction.get("gcs_output_uri"),
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    except Exception as exc:
        logger.exception("BRD information extraction failed for session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get(
    "/validate-requirement-layer/{session_id}", response_model=ValidationResultResponse
)
async def validate_requirement_layer(session_id: str):
    try:
        result = await _run_in_threadpool(run_validation, session_id)
        return ValidationResultResponse(
            success=True,
            session_id=session_id,
            validation_status="corrected"
            if result["corrections_made"]
            else "completed",
            corrections_made=result["corrections_made"],
            gcs_output_uri=result["gcs_output_uri"],
            validated_requirement_layer=result["validated_requirement_layer"],
            message=(
                f"Validation complete. {len(result['corrections'])} field(s) corrected."
                if result["corrections_made"]
                else "Validation complete. All fields verified — no corrections needed."
            ),
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        logger.exception("Validation failed for session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get(
    "/extract-file-layout/{session_id}", response_model=FileLayoutExtractionResponse
)
async def extract_file_layout(session_id: str):
    try:
        result = await _run_in_threadpool(run_file_layout_extraction, session_id)
        return FileLayoutExtractionResponse(
            success=True,
            session_id=session_id,
            message=f"File layout extracted successfully. {result['tables_extracted']} table(s) found.",
            file_layout_filename=result["file_layout_filename"],
            total_pages=result["total_pages"],
            tables_extracted=result["tables_extracted"],
            file_layout_tables=result["file_layout_tables"],
            gcs_output_uri=result["gcs_output_uri"],
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        logger.exception("File layout extraction failed for session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/brd-accept/{session_id}", response_model=BrdCheckpointResponse)
async def brd_accept_changes(session_id: str, body: BrdAcceptRequest):
    """
    Apply ONLY accepted edits as a strict patch.
    No LLM call. No re-derivation.
    """
    import json as _json
    from utils.gcs_artifact_utils import download_bytes as _dl
    from agents.extract_requirement_layer_agent.tools.brd_utils import (
        _safe_json_load as _sjl,
    )

    try:
        base_prefix = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}"
        validated_object = (
            f"{base_prefix}/extracted_data/validated_requirement_layer.json"
        )

        # Load current JSON
        current_bytes = _dl(object_name=validated_object)
        current: dict = _sjl(current_bytes.decode("utf-8"))

        accepted_fields_applied = []

        # 🔹 Strict deep patch
        def _deep_apply(base: dict, edits: dict, path=""):
            for k, v in edits.items():
                full_path = f"{path}.{k}".lstrip(".")
                if isinstance(v, dict) and isinstance(base.get(k), dict):
                    _deep_apply(base[k], v, full_path)
                else:
                    base[k] = v
                    accepted_fields_applied.append(full_path)

        if body.accepted_edits:
            _deep_apply(current, body.accepted_edits)

        # Save
        _upload_json(object_name=validated_object, payload=current)

        gcs_uri = f"gs://{config.MAPPING_ARTIFACT_BUCKET}/{validated_object}"

        return BrdCheckpointResponse(
            success=True,
            session_id=session_id,
            message="Accepted changes applied successfully.",
            validated_requirement_layer=current,
            gcs_output_uri=gcs_uri,
            accepted_fields_applied=accepted_fields_applied,
            rejected_field_results=[],
            unchanged_fields=[],
        )

    except Exception as e:
        logger.exception("BRD accept changes failed for session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/brd-reject/{session_id}", response_model=BrdCheckpointResponse)
async def brd_reject_changes(session_id: str, body: BrdRejectFreeformRequest):
    try:
        result = await _run_in_threadpool(
            run_brd_checkpoint,
            session_id,
            body.instruction,
        )

        return BrdCheckpointResponse(
            success=True,
            session_id=session_id,
            message="Requirement layer updated using instruction.",
            validated_requirement_layer=result["validated_requirement_layer"],
            gcs_output_uri=result["gcs_output_uri"],
            accepted_fields_applied=[],
            rejected_field_results=[],
            unchanged_fields=[],
        )

    except Exception as e:
        logger.exception("BRD freeform reject failed for session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(e))


@router.post(
    "/file-layout-checkpoint/{session_id}", response_model=FileLayoutCheckpointResponse
)
async def file_layout_checkpoint(session_id: str, body: FileLayoutCheckpointRequest):
    """
    Human checkpoint for file layout tables — direct edits only, no LLM re-run.

    - edited_tables: table keys → updated row arrays merged into the persisted
      file_layout_tables.json. Existing tables not mentioned are left untouched.

    Persists the merged tables back to file_layout_tables.json and updates
    final_with_layout.json in GCS.
    """
    import json as _json
    from utils.gcs_artifact_utils import download_bytes as _dl
    from agents.extract_requirement_layer_agent.tools.brd_utils import (
        _safe_json_load as _sjl,
    )

    from agents.extract_requirement_layer_agent.agents.brd_transcript_agent import run_cache_cleanup
    
    try:
        base_prefix = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}"
        layout_object = f"{base_prefix}/extracted_data/file_layout_tables.json"
        final_object = f"{base_prefix}/extracted_data/final_with_layout.json"

        try:
            existing_tables: dict = _sjl(_dl(object_name=layout_object).decode("utf-8"))
            # Strip legacy wrapper key if present
            existing_tables.pop("file_layout_tables", None)
            # Keep only list values (sheet name → rows)
            existing_tables = {k: v for k, v in existing_tables.items() if isinstance(v, list)}
        except Exception:
            existing_tables = {}

        # Unwrap if the client sent tables nested under "file_layout_tables"
        edits = body.edited_tables
        if "file_layout_tables" in edits and isinstance(edits["file_layout_tables"], dict):
            edits = edits["file_layout_tables"]

        # Merge — edited_tables keys overwrite, rest preserved
        existing_tables.update(edits)

        await _run_in_threadpool(
            _upload_json, object_name=layout_object, payload=existing_tables
        )

        # Update final_with_layout.json — always store tables under "file_layout_tables"
        # so the UI receives { "file_layout_tables": { "Sheet": [{...}] } }
        try:
            final: dict = _sjl(_dl(object_name=final_object).decode("utf-8"))
        except Exception:
            final = {}
        final["file_layout_tables"] = existing_tables
        await _run_in_threadpool(_upload_json, object_name=final_object, payload=final)

        gcs_uri = f"gs://{config.MAPPING_ARTIFACT_BUCKET}/{layout_object}"
        logger.info(
            "File layout checkpoint applied | session=%s tables_edited=%d",
            session_id,
            len(body.edited_tables),
        )

        run_cache_cleanup(session_id=session_id)
        
        return FileLayoutCheckpointResponse(
            success=True,
            session_id=session_id,
            message=f"File layout checkpoint applied. {len(body.edited_tables)} table(s) updated.",
            file_layout_tables=existing_tables,
            gcs_output_uri=gcs_uri,
        )
    except Exception as exc:
        logger.exception("File layout checkpoint failed for session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(exc))


# ─── Orchestrator Integration (Stages 2-5) ───────────────────────────────────

from agents.extract_agent.orchestrator import (
    ExtractPipelineOrchestrator,
    REQUIREMENT_JUDGE_STATE_KEY,
    DRIVER_JUDGE_STATE_KEY,
    METADATA_JUDGE_STATE_KEY,
)
from agents.extract_agent.pipeline_models import PipelineStage


@router.post("/{session_id}/driver")
async def run_driver_stage(session_id: str, payload: dict):
    user_id = payload.get("user_id")
    orchestrator = ExtractPipelineOrchestrator()
    return await orchestrator.run_stage(PipelineStage.DRIVER, user_id, session_id)


@router.post("/{session_id}/driver/approve")
async def approve_driver_stage(session_id: str, payload: dict):
    user_id = payload.get("user_id")
    overrides = payload.get("overrides", {})
    orchestrator = ExtractPipelineOrchestrator()
    return await orchestrator.approve_stage(
        PipelineStage.DRIVER, user_id, session_id, overrides
    )


@router.post("/{session_id}/discovery")
async def run_discovery_stage(session_id: str, payload: dict):
    user_id = payload.get("user_id")
    orchestrator = ExtractPipelineOrchestrator()
    return await orchestrator.run_stage(PipelineStage.DISCOVERY, user_id, session_id)


@router.post("/{session_id}/discovery/approve")
async def approve_discovery_stage(session_id: str, payload: dict):
    user_id = payload.get("user_id")
    overrides = payload.get("overrides", {})
    orchestrator = ExtractPipelineOrchestrator()
    return await orchestrator.approve_stage(
        PipelineStage.DISCOVERY, user_id, session_id, overrides
    )


@router.post("/{session_id}/metadata")
async def run_metadata_stage(session_id: str, payload: dict):
    user_id = payload.get("user_id")
    orchestrator = ExtractPipelineOrchestrator()
    # Metadata auto-completes, no approve endpoint needed
    return await orchestrator.run_stage(PipelineStage.METADATA, user_id, session_id)


# ---------------------------------------------------------------------------
# P0 Task 1 — Final metadata save: BSA approves metadata, persists to GCS
# ---------------------------------------------------------------------------

@router.post("/{session_id}/final_metadata_save")
async def final_metadata_save(session_id: str, payload: dict):
    """
    BSA-triggered save of the approved metadata output to GCS.
    Saves to: {prefix}/{session_id}/metadata/final_metadata_output.json
    Overwrites on each call (last approve wins).

    Payload shape (same as /extract-metadata response body):
      {
        "extracted_filespecs": { ... },
        "extracted_file1": {
            "entity_physical_name": ...,
            "attributes": [ ... ]
        }
      }
    """
    from fastapi.concurrency import run_in_threadpool

    object_name = (
        f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}"
        f"/metadata/final_metadata_output.json"
    )
    gcs_uri = f"gs://{config.MAPPING_ARTIFACT_BUCKET}/{object_name}"

    try:
        await run_in_threadpool(_upload_json, object_name=object_name, payload=payload)
        logger.info("final_metadata_save — session=%s saved to %s", session_id, gcs_uri)
        return {"session_id": session_id, "status": "saved", "gcs_output_uri": gcs_uri}
    except Exception as exc:
        logger.exception("final_metadata_save failed for session=%s: %s", session_id, exc)
        raise HTTPException(status_code=500, detail=str(exc))


# ---------------------------------------------------------------------------
# Mapping: common rules + rows from metadata + driver filter + mapping agent
# ---------------------------------------------------------------------------

@router.post("/{session_id}/mapping")
async def run_mapping_stage(session_id: str, payload: dict):
    from datetime import datetime
    import json as _json
    from agents.extract_agent.mapping_agent.tools.mapping_excel_tools import _build_common_rules_rows
    from agents.extract_agent.mapping_agent.agent import mapping_row_agent
    from google.adk import Runner
    from google.adk.apps.app import App
    from utils.adk_runtime import VertexAiSessionService
    from google.genai import types as _types
    from utils.gcs_artifact_utils import download_json_uri
    from fastapi.concurrency import run_in_threadpool

    # Vertex AI session credentials — required to run the mapping agent.
    # session_id in the URL is the user/GCS session; vertex_session_id is the ADK session.
    app_name         = payload.get("appName")
    vertex_session_id = payload.get("sessionId")
    user_id          = payload.get("user_id")

    # -------------------------------------------------------------------------
    # Common Rules tab + BRD rules — both from validated_requirement_layer.json
    # -------------------------------------------------------------------------
    common_rules: list = []
    common_rules_source: dict = {}
    brd_rules: dict = {"requirements_text": "", "default_values_note": "", "data_format_rules": ""}
    extract_context: dict = {}
    try:
        req_layer_uri = (
            f"gs://{config.MAPPING_ARTIFACT_BUCKET}/"
            f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/"
            f"extracted_data/validated_requirement_layer.json"
        )
        payload_json = await run_in_threadpool(download_json_uri, req_layer_uri)
        # Handle both flat and validated_requirement_layer-wrapped structures
        _rl = payload_json.get("validated_requirement_layer") or payload_json
        common_rules_source = dict(_rl.get("common_rules") or {})

        try:
            pipeline_meta = await run_in_threadpool(
                download_json_uri,
                f"gs://{config.MAPPING_ARTIFACT_BUCKET}/{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/uploads/pipeline_result.json",
            )
            if pipeline_meta.get("interface_code"):
                common_rules_source["interface_code"] = pipeline_meta["interface_code"]
        except Exception:
            pass

        common_rules = _build_common_rules_rows(common_rules_source)

        # Extract BRD rules for STEP 0 in the mapping agent — no extra GCS call needed
        _file_attrs = _rl.get("file_attributes_mapping") or {}
        brd_rules = {
            "requirements_text":  _rl.get("requirements", ""),
            "default_values_note": _file_attrs.get("default_values", ""),
            "data_format_rules":   _file_attrs.get("data_format_rules", ""),
        }

        # Build extract_context for L1 ranking — same fields used by fyi_lookup_tool
        # in the driver layer. Derived here from the already-downloaded BRD so no
        # extra GCS call is needed. Passed in every field_payload; agent uses it to
        # rank IndeMap candidates by subject-area alignment (member vs claim vs pharmacy).
        _file_specs   = _rl.get("file_specs") or {}
        _scope        = _rl.get("scope") or {}
        _filters      = _rl.get("filters_and_parameters") or {}
        _date_params  = _filters.get("date_parameters") or {}
        _in_scope_raw = _scope.get("in_scope") or _rl.get("in_scope") or ""
        _pop_type = str(_file_specs.get("file_population_type") or "").strip()
        if not _pop_type and _in_scope_raw:
            _pop_type = str(_in_scope_raw).strip()[:200]
        extract_context = {
            "file_population_type": _pop_type,
            "subject_areas":        _file_attrs.get("subject_areas", ""),
            "vendor_name":          _file_specs.get("vendor_name", ""),
            "file_name":            _file_specs.get("physical_file_name") or _file_attrs.get("file_naming_convention", ""),
            "interface_code":       common_rules_source.get("interface_code", ""),
            "effective_dates_from": common_rules_source.get("effective_dates_from", ""),
            "effective_dates_to":   common_rules_source.get("effective_dates_to", ""),
            "scope":                {k: v for k, v in _scope.items() if v},
            "filters":              {k: v for k, v in _filters.items() if k != "date_parameters" and v},
            "date_parameters":      {k: v for k, v in _date_params.items() if v},
        }
        logger.info(
            "brd_rules+extract_context extracted for session=%s req_len=%d "
            "population_type='%s' subject_areas='%s' interface='%s'",
            session_id, len(brd_rules["requirements_text"]),
            extract_context["file_population_type"],
            extract_context["subject_areas"],
            extract_context["interface_code"],
        )
    except Exception as exc:
        logger.warning("common_rules/brd_rules fetch failed for session=%s: %s", session_id, exc)

    # -------------------------------------------------------------------------
    # Transformation Rules — common_filter (merged cell) from driver layer
    # driver_data kept at outer scope for ibc_aha_context used by mapping agent
    # -------------------------------------------------------------------------
    common_filter: str | None = None
    driver_data: dict = {}
    try:
        driver_gcs_uri = (
            f"gs://{config.MAPPING_ARTIFACT_BUCKET}/"
            f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/"
            f"driver_data/approved_driver_layer_output.json"
        )
        driver_data = await run_in_threadpool(download_json_uri, driver_gcs_uri)
        common_filter = driver_data.get("sql_where_clause") or None
        if common_filter:
            extract_context["common_filter"] = common_filter
        logger.info("driver sql_where_clause loaded for session=%s", session_id)
    except Exception as exc:
        logger.warning(
            "driver_data fetch failed for session=%s (driver approve may not have been called yet): %s",
            session_id, exc,
        )

    # -------------------------------------------------------------------------
    # Transformation Rules — merged cells + rows from approved metadata
    # Loads: {session_id}/metadata/final_metadata_output.json
    # Fail-open: if metadata not yet saved, merged cells stay None, rows=[].
    # -------------------------------------------------------------------------
    target_entity: str | None = None
    driver_table_required = None
    history_data_pull = None
    rows: list = []

    try:
        metadata_gcs_uri = (
            f"gs://{config.MAPPING_ARTIFACT_BUCKET}/"
            f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/"
            f"metadata/final_metadata_output.json"
        )
        metadata = await run_in_threadpool(download_json_uri, metadata_gcs_uri)

        # Merged cell fields
        filespecs = metadata.get("extracted_filespecs") or {}
        target_entity = filespecs.get("Physical File Name") or None

        def _to_bool(val):
            if not val:
                return None
            return str(val).strip().lower() in ("true", "yes", "y", "1")

        driver_table_required = _to_bool(common_rules_source.get("driver_required"))
        history_data_pull = _to_bool(common_rules_source.get("history_required"))

        # Build one row per attribute — apply File Layout Status rules
        attributes = _collect_mapping_metadata_attributes(metadata)
        now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

        for attr in attributes:
            nullability = (attr.get("Nullability") or "").strip().lower()
            is_optional = nullability in ("optional", "conditional")

            key_parts = [
                k for k in [attr.get("Primary Key"), attr.get("Foreign Key"), attr.get("Alternate Key1")]
                if k and str(k).strip()
            ]

            row = {
                "target_attribute":       attr.get("Attribute Name"),
                "logical_attribute_name": attr.get("Logical Attribute Name"),
                "attribute_description":  attr.get("Attribute Description"),
                "data_type":              attr.get("Data Type"),
                "length":                 attr.get("Length"),
                "precision":              attr.get("Precision"),
                "format":                 attr.get("Format"),
                "nullable":               attr.get("Nullability"),
                "default_value":          attr.get("Default Value"),
                "order_no":               None,
                "cdc_indicator":          None,
                "key_columns":            ", ".join(key_parts) if key_parts else None,
                "rule_type":              "Default" if is_optional else None,
                "rule_name":              None,
                "source_entity":          None,
                "source_attribute":       None,
                "join":                   None,
                "filter":                 None,
                "transformation_rule":    "Populate Blank" if is_optional else None,
                "special_consideration":  None,
                "last_updated":           now_str,
                "match_level":            None,
                "match_score":            None,
                "open_item":              False,
                "open_item_reason":       None,
                # Internal flags — stripped before response is returned
                "_needs_agent":           not is_optional,
                "_file_name":             attr.get("File Name"),
            }
            rows.append(row)

        logger.info(
            "metadata loaded for session=%s target_entity=%s attributes=%d",
            session_id, target_entity, len(rows),
        )
    except Exception as exc:
        logger.warning(
            "metadata fetch failed for session=%s (final_metadata_save may not have been called yet): %s",
            session_id, exc,
        )

    # -------------------------------------------------------------------------
    # Mapping Agent — run L1→L2→L3 waterfall for each Required field
    # Requires appName + sessionId (Vertex AI) in the request payload.
    # Fail-open: if credentials missing or agent errors, Required rows stay null.
    # -------------------------------------------------------------------------
    ibc_aha_context = (driver_data.get("ibc_aha_context") if driver_data else None) or "both"

    required_rows = [r for r in rows if r.get("_needs_agent")]

    # TODO-PERF: remove this limit before production — for testing only
    _TEST_FIELD_LIMIT = int(payload.get("_test_field_limit", 0))
    if _TEST_FIELD_LIMIT > 0:
        required_rows = required_rows[:_TEST_FIELD_LIMIT]
        logger.info("TEST MODE: limiting agent to first %d required fields", _TEST_FIELD_LIMIT)

    if app_name and vertex_session_id and user_id and required_rows:
        try:
            session_service = VertexAiSessionService(
                project=config.GOOGLE_CLOUD_PROJECT,
                location=config.GOOGLE_CLOUD_LOCATION,
            )

            from pydantic import ValidationError as _PydanticValidationError
            try:
                mapping_app = App(name=app_name, root_agent=mapping_row_agent)
            except _PydanticValidationError as _ve:
                logger.warning("mapping App name validation failed (%s); using model_construct.", _ve)
                mapping_app = App.model_construct(name=app_name, root_agent=mapping_row_agent)
            runner = Runner(app=mapping_app, session_service=session_service)

            # TODO-DEV: remove event writing before production
            from pathlib import Path as _Path
            _EVENTS_DIR = _Path(__file__).parent.parent.parent / "ext_mapping_events"

            def _write_mapping_event(field_name: str, evt_idx: int, evt: Any) -> None:
                try:
                    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in field_name)[:40]
                    field_dir = _EVENTS_DIR / safe
                    field_dir.mkdir(parents=True, exist_ok=True)
                    (field_dir / f"event{evt_idx}.txt").write_text(
                        f"EVENT {evt_idx}:\n\n{evt}\n", encoding="utf-8"
                    )
                except Exception:
                    pass

            logger.info(
                "mapping agent starting — session=%s required_fields=%d",
                session_id, len(required_rows),
            )

            inter_field_delay_sec = float(os.getenv("MAPPING_INTER_FIELD_DELAY_SEC", "1.0"))
            agent_rows = []

            for row_index, row in enumerate(required_rows):
                field_payload = {
                    "target_attribute":       row["target_attribute"],
                    "logical_attribute_name": row["logical_attribute_name"],
                    "attribute_description":  row["attribute_description"],
                    "data_type":              row["data_type"],
                    "length":                 row["length"],
                    "precision":              row["precision"],
                    "format":                 row["format"],
                    "nullable":               row["nullable"],
                    "default_value":          row["default_value"],
                    "key_columns":            row["key_columns"],
                    "ibc_aha_context":        ibc_aha_context,
                    "brd_rules":              brd_rules,
                    "extract_context":        extract_context,
                    "file_name":              row.get("_file_name"),
                }
                msg = _types.Content(
                    role="user",
                    parts=[_types.Part(text=_json.dumps(field_payload, indent=2))],
                )
                
                max_retries = 3
                recommended_wait_sec = 0.0
                for attempt in range(max_retries):
                    try:
                        field_session = await session_service.create_session(
                            app_name=app_name,
                            user_id=user_id,
                            state={"mapping_rows": []},
                        )
                        field_session_id = str(field_session.id)
                        await wait_for_llm_request_slot(f"mapping:{session_id}")
                        event_count = 0
                        async for event in runner.run_async(
                            user_id=user_id,
                            session_id=field_session_id,
                            new_message=msg,
                        ):
                            recommended_wait_sec = max(
                                recommended_wait_sec,
                                await record_llm_usage_and_get_wait(
                                    event,
                                    session_id=f"mapping:{session_id}",
                                    buffer_tokens=300,
                                ),
                            )

                            # _write_mapping_event(row["target_attribute"], event_count, event)  # local dev debug events
                            event_count += 1
                            if event_count > 300:
                                logger.warning(
                                    "mapping agent safety limit reached for target=%s",
                                    row["target_attribute"],
                                )
                                break

                        field_session = await session_service.get_session(
                            app_name=app_name,
                            user_id=user_id,
                            session_id=field_session_id,
                        )
                        field_rows = (
                            (getattr(field_session, "state", None) or {}).get("mapping_rows")
                            or []
                        )
                        if field_rows:
                            agent_rows.extend(field_rows)
                        
                        # Success for this row, break retry loop
                        break

                    except Exception as field_exc:
                        if is_resource_exhausted_error(field_exc) and attempt < max_retries - 1:
                            delay = calculate_mapping_retry_delay(attempt)
                            logger.warning(
                                "mapping agent 429 Resource Exhausted for target=%s (attempt %d/%d). Retrying in %.2fs...",
                                row["target_attribute"], attempt + 1, max_retries, delay
                            )
                            await asyncio.sleep(delay)
                            continue
                        else:
                            logger.warning(
                                "mapping agent failed for target=%s: %s",
                                row["target_attribute"], field_exc,
                            )
                            # On final failure, don't crash the whole pipeline, just move to next row
                            break

                if row_index < len(required_rows) - 1:
                    delay_sec = max(inter_field_delay_sec, recommended_wait_sec)
                    if delay_sec > 0:
                        logger.info(
                            "mapping agent inter-field delay %.2fs after target=%s",
                            delay_sec,
                            row["target_attribute"],
                        )
                        await asyncio.sleep(delay_sec)

            # Merge agent rows into the ordered rows list by target_attribute
            agent_by_attr = {r.get("target_attribute"): r for r in agent_rows}
            for i, row in enumerate(rows):
                if not row.get("_needs_agent"):
                    continue
                target_attribute = row.get("target_attribute")
                if target_attribute in agent_by_attr:
                    rows[i] = agent_by_attr[target_attribute]
                    continue

                rows[i] = {
                    **row,
                    "rule_type": None,
                    "source_entity": None,
                    "source_attribute": None,
                    "join": None,
                    "filter": None,
                    "transformation_rule": None,
                    "match_level": None,
                    "match_score": None,
                    "open_item": True,
                    "open_item_reason": "Mapping agent did not return a result for this required field.",
                }

            logger.info(
                "mapping agent done — session=%s agent_rows=%d",
                session_id, len(agent_rows),
            )

        except Exception as agent_exc:
            logger.exception(
                "mapping agent block failed for session=%s: %s", session_id, agent_exc
            )
    else:
        if required_rows:
            logger.info(
                "mapping agent skipped — appName/sessionId not provided. "
                "Required rows returned as null placeholders. session=%s",
                session_id,
            )

    for row in rows:
        if row.get("_needs_agent") and not row.get("open_item") and not row.get("match_level"):
            row.update(
                {
                    "rule_type": None,
                    "source_entity": None,
                    "source_attribute": None,
                    "join": None,
                    "filter": None,
                    "transformation_rule": None,
                    "match_level": None,
                    "match_score": None,
                    "open_item": True,
                    "open_item_reason": "Mapping agent did not return a result for this required field.",
                }
            )

    # Strip internal flags before returning
    for row in rows:
        row.pop("_needs_agent", None)
        row.pop("_file_name", None)

    mapping_draft = {
        "session_id": session_id,
        "status": "draft_ready",
        "common_rules": common_rules,
        "transformation_rules": {
            "target_entity":         target_entity,
            "driver_table_required": driver_table_required,
            "history_data_pull":     history_data_pull,
            "common_filter":         common_filter,
            "rows":                  rows,
        },
    }

    # Persist draft so the approver endpoint can reference it
    try:
        from utils.gcs_artifact_utils import upload_json as _upload_json
        draft_object = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/mapping/mapping_result.json"
        await run_in_threadpool(_upload_json, object_name=draft_object, payload=mapping_draft)
    except Exception as exc:
        logger.warning("Failed to persist mapping draft for session=%s: %s", session_id, exc)

    return mapping_draft


@router.post(
    "/{session_id}/mapping/field/human-checkpoint",
    response_model=MappingFieldCheckpointResponse,
)
async def mapping_field_human_checkpoint(
    session_id: str,
    body: MappingFieldCheckpointRequest,
):
    import json as _json
    import uuid as _uuid
    from fastapi.concurrency import run_in_threadpool
    from pydantic import ValidationError as _PydanticValidationError
    from google.adk import Runner
    from google.adk.apps.app import App
    from google.adk.events import Event, EventActions
    from utils.adk_runtime import VertexAiSessionService
    from google.genai import types as _types
    from agents.extract_agent.mapping_agent.agent import mapping_field_checkpoint_agent
    from utils.gcs_artifact_utils import download_json_uri

    current_row = dict(body.current_row or {})
    current_target = current_row.get("target_attribute")

    if current_target and current_target != body.target_attribute:
        raise HTTPException(
            status_code=400,
            detail="current_row.target_attribute must match target_attribute.",
        )

    brd_rules = {
        "requirements_text": "",
        "default_values_note": "",
        "data_format_rules": "",
    }
    try:
        req_layer_uri = (
            f"gs://{config.MAPPING_ARTIFACT_BUCKET}/"
            f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/"
            f"extracted_data/validated_requirement_layer.json"
        )
        payload_json = await run_in_threadpool(download_json_uri, req_layer_uri)
        req_layer = payload_json.get("validated_requirement_layer") or payload_json
        file_attrs = req_layer.get("file_attributes_mapping") or {}
        brd_rules = {
            "requirements_text": req_layer.get("requirements", ""),
            "default_values_note": file_attrs.get("default_values", ""),
            "data_format_rules": file_attrs.get("data_format_rules", ""),
        }
    except Exception as exc:
        logger.warning(
            "checkpoint brd_rules fetch failed for session=%s: %s",
            session_id,
            exc,
        )

    ibc_aha_context = "both"
    try:
        driver_gcs_uri = (
            f"gs://{config.MAPPING_ARTIFACT_BUCKET}/"
            f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/"
            f"driver_data/approved_driver_layer_output.json"
        )
        driver_data = await run_in_threadpool(download_json_uri, driver_gcs_uri)
        ibc_aha_context = driver_data.get("ibc_aha_context") or "both"
    except Exception as exc:
        logger.warning(
            "checkpoint driver_data fetch failed for session=%s: %s",
            session_id,
            exc,
        )

    field_payload = {
        "current_row": current_row,
        "target_attribute": body.target_attribute,
        "logical_attribute_name": current_row.get("logical_attribute_name"),
        "attribute_description": current_row.get("attribute_description"),
        "data_type": current_row.get("data_type"),
        "length": current_row.get("length"),
        "precision": current_row.get("precision"),
        "format": current_row.get("format"),
        "nullable": current_row.get("nullable"),
        "default_value": current_row.get("default_value"),
        "key_columns": current_row.get("key_columns"),
        "ibc_aha_context": ibc_aha_context,
        "brd_rules": brd_rules,
        "bsa_instruction": body.bsa_instruction,
    }

    try:
        session_service = VertexAiSessionService(
            project=config.GOOGLE_CLOUD_PROJECT,
            location=config.GOOGLE_CLOUD_LOCATION,
        )
        session = await session_service.get_session(
            app_name=body.appName,
            user_id=body.user_id,
            session_id=body.sessionId,
        )
        clear_event = Event(
            author="system",
            invocation_id=f"sys-{_uuid.uuid4()}",
            actions=EventActions(state_delta={"mapping_rows": []}),
        )
        await session_service.append_event(session=session, event=clear_event)

        try:
            mapping_app = App(
                name=body.appName,
                root_agent=mapping_field_checkpoint_agent,
            )
        except _PydanticValidationError:
            mapping_app = App.model_construct(
                name=body.appName,
                root_agent=mapping_field_checkpoint_agent,
            )

        runner = Runner(app=mapping_app, session_service=session_service)
        msg = _types.Content(
            role="user",
            parts=[_types.Part(text=_json.dumps(field_payload, indent=2))],
        )

        max_retries = 3
        for attempt in range(max_retries):
            try:
                await wait_for_llm_request_slot(f"checkpoint:{session_id}")
                async for event in runner.run_async(
                    user_id=body.user_id,
                    session_id=body.sessionId,
                    new_message=msg,
                ):
                    await record_llm_usage_and_get_wait(
                        event,
                        session_id=f"checkpoint:{session_id}",
                        buffer_tokens=300,
                    )
                break
            except Exception as exc:
                if is_resource_exhausted_error(exc) and attempt < max_retries - 1:
                    delay = calculate_retry_delay(attempt)
                    logger.warning("mapping checkpoint 429 Resource Exhausted (attempt %d/%d). Retrying in %.2fs...", attempt+1, max_retries, delay)
                    await asyncio.sleep(delay)
                else: raise

        session = await session_service.get_session(
            app_name=body.appName,
            user_id=body.user_id,
            session_id=body.sessionId,
        )
        agent_rows = (getattr(session, "state", None) or {}).get("mapping_rows") or []
        row = next(
            (
                r
                for r in reversed(agent_rows)
                if r.get("target_attribute") == body.target_attribute
            ),
            None,
        )

        if not row:
            raise HTTPException(
                status_code=500,
                detail="Checkpoint agent did not return a mapping row.",
            )

        return MappingFieldCheckpointResponse(
            success=True,
            session_id=session_id,
            row=row,
        )

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(
            "mapping field checkpoint failed for session=%s target=%s",
            session_id,
            body.target_attribute,
        )
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/{session_id}/mapping/accept", response_model=MappingApproverResponse)
async def approve_mapping_with_edits(session_id: str, body: MappingApproverRequest):
    """
    Human-in-the-loop approver for the mapping stage.
    Fetches the initial mapping draft from GCS (if present), merges it with
    the UI-edited payload, and persists both as mapping_result.json.
    
    Expected body structure:
    {
      "common_rules": [{"Field": "...", "Value": "..."}],
      "transformation_rules": {
        "target_entity": "...",
        "driver_table_required": bool,
        "history_data_pull": bool,
        "common_filter": "...",
        "rows": [{...}]
      }
    }
    """
    from datetime import datetime
    from fastapi.concurrency import run_in_threadpool
    from utils.gcs_artifact_utils import upload_json as _upload_json, download_bytes as _dl

    base_prefix = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}"
    result_object = f"{base_prefix}/mapping/mapping_result.json"

    try:
        # Load initial mapping draft saved by run_mapping_stage (if any)
        initial_mapping: dict = {}
        try:
            raw = await run_in_threadpool(_dl, object_name=result_object)
            initial_mapping = json.loads(raw.decode("utf-8"))
        except FileNotFoundError:
            pass

        # Overwrite mapping_result.json with the approved data, preserving initial draft
        mapping_result = {
            **initial_mapping,
            "approved_at": datetime.utcnow().isoformat(),
            "status": "approved",
            "common_rules": body.common_rules,
            "transformation_rules": body.transformation_rules,
        }

        gcs_uri = await run_in_threadpool(
            _upload_json, object_name=result_object, payload=mapping_result
        )

        logger.info("Mapping result saved | session=%s uri=%s", session_id, gcs_uri)
        return MappingApproverResponse(
            success=True,
            session_id=session_id,
            message="Mapping approved and saved as mapping_result.json.",
            gcs_output_uri=gcs_uri,
            common_rules=body.common_rules,
            transformation_rules=body.transformation_rules,
        )
    except Exception as exc:
        logger.exception("Mapping approver failed for session=%s", session_id)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/{session_id}/state")
async def get_pipeline_state(session_id: str, user_id: str):
    orchestrator = ExtractPipelineOrchestrator()
    state = await orchestrator.get_pipeline_state(user_id, session_id)
    return state.model_dump(mode="json")


class ExtractsMetadataRequest(BaseModel):
    user_id: str
    session_id: str
    brd_gcs_uri: str
    layout_gcs_uri: str


def _is_blank_metadata_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"", "null", "none", "n/a", "na"}
    return False


def _metadata_key_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _parse_layout_data_type(raw_type: Any, field_name: str = "") -> tuple[Any, Any, Any]:
    text = str(raw_type or "").strip()
    if not text:
        name_lower = str(field_name or "").lower()
        if "date" in name_lower:
            return "DATE", None, None
        if "amount" in name_lower or "fee" in name_lower or "total" in name_lower:
            return "DECIMAL", None, None
        return "VARCHAR", None, None

    match = re.match(r"^\s*([a-zA-Z0-9_ ]+)\s*(?:\(([^)]*)\))?", text)
    if not match:
        return text.upper(), None, None

    data_type = match.group(1).strip().upper()
    length = None
    precision = None
    params = match.group(2)
    if params:
        parts = [part.strip() for part in params.split(",") if part.strip()]
        if parts:
            length = parts[0]
        if len(parts) > 1:
            precision = parts[1]
    return data_type, length, precision


def _extract_file_layout_tables(layout_content: Dict[str, Any]) -> Dict[str, list]:
    if not isinstance(layout_content, dict):
        return {}

    wrapped = layout_content.get("file_layout_tables")
    if isinstance(wrapped, dict):
        layout_content = wrapped

    return {
        str(table_name): rows
        for table_name, rows in layout_content.items()
        if isinstance(rows, list)
    }


def _build_metadata_file1_from_layout(layout_content: Dict[str, Any]) -> Dict[str, Any]:
    file_layout_tables = _extract_file_layout_tables(layout_content)
    if not file_layout_tables:
        return {}

    attributes: list[dict[str, Any]] = []
    for table_name, table_rows in file_layout_tables.items():
        if not isinstance(table_rows, list):
            continue
        column_name_key = None
        for row in table_rows:
            if isinstance(row, dict):
                column_name_key = next(
                    (
                        key
                        for key in row.keys()
                        if str(key).strip().lower().endswith("column name")
                    ),
                    None,
                )
                if column_name_key:
                    break
        if not column_name_key:
            continue

        for row in table_rows:
            if not isinstance(row, dict):
                continue
            attribute_name = str(row.get(column_name_key) or "").strip()
            if not attribute_name:
                continue

            raw_type = row.get("Column Data Type")
            data_type, length, precision = _parse_layout_data_type(
                raw_type, attribute_name
            )
            notes = str(row.get("Notes") or "").strip()
            sending = str(row.get("Sending? Y or N") or "").strip().lower()
            optional = "optional" in notes.lower() or sending == "n"
            description = str(row.get("Description") or "").strip()

            attributes.append(
                {
                    "Attribute Name": attribute_name,
                    "Logical Attribute Name": attribute_name,
                    "Attribute Description": description or "",
                    "Data Type": data_type,
                    "Length": length,
                    "Precision": precision,
                    "Format": "YYYY-MM-DD" if data_type == "DATE" else None,
                    "Nullability": "Optional" if optional else "NOT NULL",
                    "Default Value": None,
                    "Primary Key": "",
                    "Foreign Key": "",
                    "Alternate Key1": "",
                }
            )

    first_table_name = next(iter(file_layout_tables.keys()), None)
    return {
        "entity_type": "File",
        "file_type": "Outgoing",
        "entity_physical_name": None,
        "entity_business_name": first_table_name,
        "entity_description": (
            "Extracted from multi-file layout tables."
            if len(file_layout_tables) > 1
            else "Extracted from file layout table."
        ),
        "attributes": attributes,
    }


def _backfill_filespecs_from_brd(
    extracted_filespecs: Dict[str, Any], brd_content: Dict[str, Any]
) -> Dict[str, Any]:
    requirement_layer = (
        brd_content.get("corrected_requirement_layer")
        or brd_content.get("validated_requirement_layer")
        or brd_content.get("requirement_layer")
        or brd_content
    )
    brd_filespecs = requirement_layer.get("file_specs") or {}
    if not isinstance(brd_filespecs, dict):
        return extracted_filespecs

    brd_by_token = {
        _metadata_key_token(key): value for key, value in brd_filespecs.items()
    }
    backfilled = dict(extracted_filespecs or {})
    for key, value in list(backfilled.items()):
        token = _metadata_key_token(key)
        if _is_blank_metadata_value(value) and token in brd_by_token:
            backfilled[key] = brd_by_token[token]
    return backfilled


def _collect_mapping_metadata_attributes(metadata: Dict[str, Any]) -> list[dict]:
    """
    Collect mapping attributes from either a flattened extracted_file1 payload or
    multiple extracted_fileN sections. Adds file context when available so
    duplicate field names across files remain distinguishable to the mapping agent.
    """
    if not isinstance(metadata, dict):
        return []

    collected: list[dict] = []
    for key, value in metadata.items():
        if not str(key).startswith("extracted_file") or not isinstance(value, dict):
            continue

        file_name = (
            value.get("entity_business_name")
            or value.get("entity_physical_name")
            or key
        )

        for attr in value.get("attributes") or []:
            if not isinstance(attr, dict):
                continue
            attr_copy = dict(attr)
            # Store file name for display/output disambiguation — but do NOT prepend it
            # to Logical Attribute Name or Attribute Description. Those fields are used
            # directly as search signals (IndeMap vector search, L2 standards query) and
            # the file name prefix pollutes embeddings and search queries with noise.
            attr_copy.setdefault("File Name", file_name)
            collected.append(attr_copy)

    if collected:
        return collected

    return (metadata.get("extracted_file1") or {}).get("attributes") or []


class JudgeH1Request(BaseModel):
    user_id: str
    session_id: str
    brd_gcs_uri: str
    layout_gcs_uri: str
    transcript_gcs_uri: str | None = None
    brd_markdown_gcs_uri: str | None = None
    layout_markdown_gcs_uri: str | None = None
    judge_mode: str = "pre"
    bsa_rejection_feedback: str | None = None
    revision_number: int = 0


class JudgeH1Response(BaseModel):
    success: bool
    session_id: str
    judge_mode: str
    message: str
    source_artifacts: Dict[str, Any]
    normalized_requirement_model: Dict[str, Any]
    evaluation: Dict[str, Any]
    revision_directive: Optional[Dict[str, Any]] = None
    annotated_artifact: Dict[str, Any] = Field(default_factory=dict)
    bsa_review_summary: str


@router.post("/extract-metadata")
async def extract_metadata(request: ExtractsMetadataRequest):
    try:
        from utils.gcs_artifact_utils import download_json_uri
        import openpyxl
        from google.genai import types
        from google.adk.apps import App
        from utils.adk_runtime import VertexAiSessionService
        from google.adk import Runner

        from agents.extract_agent.metadata_agent.agent import metadata_extractor_agent

        # Load BRD and Layout JSONs
        brd_content = download_json_uri(request.brd_gcs_uri)
        layout_content = download_json_uri(request.layout_gcs_uri)

        # Load Excel template (server-relative)
        template_path = os.path.join(
            os.path.dirname(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            ),
            "templates",
            "extracts_metadata_template.xlsx",
        )

        try:
            wb = openpyxl.load_workbook(template_path)
            filespecs_sheet = wb["FileSpecs"]
            file1_sheet = wb["file1"]
        except Exception as e:
            logger.error(f"Failed to load template {template_path}: {e}")
            raise HTTPException(
                status_code=500, detail=f"Failed to load template: {str(e)}"
            )

        # --- Parse FileSpecs tab ---
        filespecs_expected_keys = []
        for row in filespecs_sheet.iter_rows(min_row=1, max_row=100):
            val = row[0].value
            if val and isinstance(val, str) and val.strip() != "Tab Name":
                filespecs_expected_keys.append(val.strip())

        # --- Parse file1 tab ---
        # Rows 1-5 are header fields (Entity Type, File Type, etc.)
        file1_header_fields = []
        for row_idx in range(1, 6):
            row_cells = list(
                file1_sheet.iter_rows(
                    min_row=row_idx, max_row=row_idx, values_only=False
                )
            )
            if row_cells:
                cells = row_cells[0]
                label = cells[0].value
                if label and isinstance(label, str):
                    file1_header_fields.append(label.strip())

        # Row 6 contains the attribute column headers
        file1_column_headers = []
        header_row = list(
            file1_sheet.iter_rows(min_row=6, max_row=6, values_only=False)
        )
        if header_row:
            for cell in header_row[0]:
                if cell.value and isinstance(cell.value, str):
                    file1_column_headers.append(cell.value.strip())

        # Format input for the agent
        prompt = (
            f"=== FileSpecs Expected Keys ===\n"
            f"{json.dumps(filespecs_expected_keys, indent=2)}\n\n"
            f"=== file1 Header Fields ===\n"
            f"{json.dumps(file1_header_fields, indent=2)}\n\n"
            f"=== file1 Expected Columns ===\n"
            f"{json.dumps(file1_column_headers, indent=2)}\n\n"
            f"=== BRD Content ===\n"
            f"{json.dumps(brd_content, indent=2)}\n\n"
            f"=== Layout Content ===\n"
            f"{json.dumps(layout_content, indent=2)}"
        )
        message = types.Content(role="user", parts=[types.Part.from_text(text=prompt)])

        # Run the agent
        session_service = VertexAiSessionService(
            project=config.GOOGLE_CLOUD_PROJECT, location=config.GOOGLE_CLOUD_LOCATION
        )
        app = App.model_construct(
            name=config.REASONING_ENGINE_RESOURCE, root_agent=metadata_extractor_agent
        )
        runner = Runner(app=app, session_service=session_service)

        max_retries = 3
        metadata_agent_session_id = None
        metadata_rate_key = "metadata:global"
        metadata_rpm_limit = int(
            os.getenv("METADATA_LLM_RPM_LIMIT", str(config.LLM_RPM_LIMIT))
        )
        metadata_tpm_limit = int(
            os.getenv("METADATA_LLM_TPM_LIMIT", str(config.LLM_TPM_LIMIT))
        )
        metadata_retry_base_delay = float(
            os.getenv("METADATA_LLM_RETRY_BASE_DELAY", "8.0")
        )
        metadata_retry_max_delay = float(
            os.getenv("METADATA_LLM_RETRY_MAX_DELAY", "30.0")
        )
        for attempt in range(max_retries):
            try:
                metadata_agent_session = await session_service.create_session(
                    app_name=config.REASONING_ENGINE_RESOURCE,
                    user_id=request.user_id,
                    state={},
                )
                metadata_agent_session_id = str(metadata_agent_session.id)
                logger.info(
                    "metadata agent temp session created original_session=%s temp_session=%s attempt=%d/%d",
                    request.session_id,
                    metadata_agent_session_id,
                    attempt + 1,
                    max_retries,
                )
                await wait_for_llm_request_slot(
                    metadata_rate_key,
                    rpm_limit=metadata_rpm_limit,
                )
                recommended_wait_sec = 0.0
                async for event in runner.run_async(
                    user_id=request.user_id,
                    session_id=metadata_agent_session_id,
                    new_message=message,
                ):
                    recommended_wait_sec = max(
                        recommended_wait_sec,
                        await record_llm_usage_and_get_wait(
                            event,
                            session_id=metadata_rate_key,
                            buffer_tokens=1000,
                            rpm_limit=metadata_rpm_limit,
                            tpm_limit=metadata_tpm_limit,
                            window_seconds=20,
                        ),
                    )
                if recommended_wait_sec > 0:
                    logger.info("metadata post-run delay %.2fs", recommended_wait_sec)
                    await asyncio.sleep(recommended_wait_sec)
                break
            except Exception as exc:
                if (
                    is_resource_exhausted_error(exc)
                    or is_transient_llm_transport_error(exc)
                ) and attempt < max_retries - 1:
                    delay = min(
                        metadata_retry_max_delay,
                        metadata_retry_base_delay * (2 ** max(0, attempt)),
                    )
                    logger.warning(
                        "metadata agent retryable LLM error (attempt %d/%d). Retrying in %.2fs... error=%s",
                        attempt + 1,
                        max_retries,
                        delay,
                        exc,
                    )
                    await asyncio.sleep(delay)
                else: raise

        if not metadata_agent_session_id:
            raise HTTPException(
                status_code=500,
                detail="Metadata agent did not create a temporary session.",
            )

        session = await session_service.get_session(
            app_name=config.REASONING_ENGINE_RESOURCE,
            user_id=request.user_id,
            session_id=metadata_agent_session_id,
        )

        # --- Parse extracted data from session state ---
        extracted_metadata = session.state.get("extracted_metadata", {})
        logger.info(
            f"extracted_metadata type={type(extracted_metadata).__name__}, value preview={str(extracted_metadata)[:500]}"
        )
        if isinstance(extracted_metadata, str):
            import re

            text_blob = extracted_metadata.strip()
            pattern = r"```(?:json)?\s*(.*?)\s*```"
            match = re.search(pattern, text_blob, re.DOTALL)
            if match:
                text_blob = match.group(1).strip()
            elif not (text_blob.startswith("{") or text_blob.startswith("[")):
                # Model wrapped the JSON in prose — slice out the object span.
                s, e = text_blob.find("{"), text_blob.rfind("}")
                if s != -1 and e > s:
                    text_blob = text_blob[s : e + 1]
            try:
                if text_blob.startswith("{") or text_blob.startswith("["):
                    extracted_metadata = json.loads(text_blob)
                else:
                    raise ValueError(
                        f"String does not start with JSON object: {text_blob[:50]}"
                    )
            except Exception as e:
                logger.warning(
                    f"Failed to parse extracted_metadata string as JSON: {e}"
                )
                extracted_metadata = {}

        # Extract the two sections — try session state sub-keys first,
        # then fall back to parsing from the combined extracted_metadata dict
        extracted_filespecs = session.state.get("extracted_filespecs", {})
        extracted_file1 = session.state.get("extracted_file1", {})

        # Handle string types for sub-keys
        if isinstance(extracted_filespecs, str):
            try:
                extracted_filespecs = json.loads(extracted_filespecs)
            except Exception:
                extracted_filespecs = {}
        if isinstance(extracted_file1, str):
            try:
                extracted_file1 = json.loads(extracted_file1)
            except Exception:
                extracted_file1 = {}

        # Fallback: if sub-keys are empty, parse from extracted_metadata
        if not extracted_filespecs and isinstance(extracted_metadata, dict):
            extracted_filespecs = extracted_metadata.get("filespecs", {})
            if (
                not extracted_filespecs
                and "filespecs" not in extracted_metadata
                and "file1" not in extracted_metadata
            ):
                # Old flat format — treat entire dict as filespecs
                extracted_filespecs = extracted_metadata
        if not extracted_file1 and isinstance(extracted_metadata, dict):
            extracted_file1 = extracted_metadata.get("file1", {})

        # if isinstance(extracted_filespecs, dict) and extracted_filespecs:
        #     extracted_filespecs = _backfill_filespecs_from_brd(
        #         extracted_filespecs, brd_content
        #     )

        # if (
        #     not isinstance(extracted_file1, dict)
        #     or not extracted_file1.get("attributes")
        # ):
        #     layout_tables = _extract_file_layout_tables(layout_content)
        #     logger.info(
        #         "metadata layout tables available for fallback: %s",
        #         list(layout_tables.keys()),
        #     )
        #     fallback_file1 = _build_metadata_file1_from_layout(layout_content)
        #     if fallback_file1.get("attributes"):
        #         logger.warning(
        #             "metadata agent returned no attributes; using deterministic layout fallback with %d rows",
        #             len(fallback_file1["attributes"]),
        #         )
        #         extracted_file1 = fallback_file1

        # --- BigQuery Persistence ---
        from agents.data_map_copilot_agent.sub_agents.metadata_fill_agent.tools.bq_tools import (
            create_metadata_and_filespecs_tables,
            append_filespecs_to_bq,
            append_chunk_to_bq,
        )

        unique_id = request.session_id.replace("-", "_")
        metadata_table_id, filespecs_table_id = create_metadata_and_filespecs_tables(
            unique_id, request.session_id
        )

        # Persist FileSpecs to BQ
        if isinstance(extracted_filespecs, dict) and extracted_filespecs:
            rows = [
                {"Field": str(k), "Value": str(v) if v is not None else ""}
                for k, v in extracted_filespecs.items()
            ]
            append_filespecs_to_bq(json.dumps(rows), filespecs_table_id)

        # Persist file1 attributes to metadata_template BQ table
        if isinstance(extracted_file1, dict) and extracted_file1.get("attributes"):
            append_chunk_to_bq(
                json.dumps(extracted_file1["attributes"]),
                metadata_table_id,
            )

        bq_reference = {
            "metadata_table": metadata_table_id,
            "filespecs_table": filespecs_table_id,
        }

        # Save to data.json
        data_json_path = os.path.join("data", "data.json")
        if os.path.exists(data_json_path):
            with open(data_json_path, "r") as f:
                data_json = json.load(f)

            if request.session_id not in data_json:
                data_json[request.session_id] = {}

            data_json[request.session_id]["metadata_bq_reference"] = bq_reference

            with open(data_json_path, "w") as f:
                json.dump(data_json, f, indent=2)

        return {
            "success": True,
            "session_id": request.session_id,
            "extracted_filespecs": extracted_filespecs,
            "extracted_file1": extracted_file1,
            "bq_reference": bq_reference,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error in extract_metadata: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _extract_requirement_layer_payload(brd_content: Dict[str, Any]) -> Dict[str, Any]:
    return (
        brd_content.get("corrected_requirement_layer")
        or brd_content.get("validated_requirement_layer")
        or brd_content.get("requirement_layer")
        or brd_content
    )


def _extract_layout_rows(layout_content: Dict[str, Any]) -> list[dict]:
    file_layout_tables = layout_content.get("file_layout_tables")
    if isinstance(file_layout_tables, dict):
        rows: list[dict] = []
        for table_name, table_rows in file_layout_tables.items():
            if not isinstance(table_rows, list):
                continue
            for index, row in enumerate(table_rows, start=1):
                row_copy = dict(row or {})
                row_copy.setdefault("__table_name__", table_name)
                row_copy.setdefault("__position__", index)
                rows.append(row_copy)
        return rows

    if isinstance(layout_content.get("layout_raw"), list):
        return [dict(row or {}) for row in layout_content["layout_raw"]]

    if isinstance(layout_content.get("layout_content"), list):
        return [dict(row or {}) for row in layout_content["layout_content"]]

    if isinstance(layout_content, dict) and all(
        isinstance(value, list) for value in layout_content.values()
    ):
        rows: list[dict] = []
        for table_name, table_rows in layout_content.items():
            for index, row in enumerate(table_rows or [], start=1):
                row_copy = dict(row or {})
                row_copy.setdefault("__table_name__", table_name)
                row_copy.setdefault("__position__", index)
                rows.append(row_copy)
        return rows

    return []


def _normalize_filter_values(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        return [value]
    return [value]


def _guess_domain(requirement_layer: Dict[str, Any]) -> tuple[str, str]:
    subject_areas = str(
        (requirement_layer.get("file_attributes_mapping") or {}).get("subject_areas")
        or ""
    ).lower()
    requirements_text = str(requirement_layer.get("requirements") or "").lower()
    combined = f"{subject_areas}\n{requirements_text}"
    if any(token in combined for token in ["member", "eligibility", "enrollment"]):
        return "Member", "Eligibility"
    if any(token in combined for token in ["claim", "medicaid reclamation"]):
        return "Claims", "Medical"
    if "provider" in combined:
        return "Provider", "Network"
    if "pharmacy" in combined:
        return "Pharmacy", "PBM"
    return "Other", "Other"


def _build_judge_requirement_model(
    requirement_layer: Dict[str, Any], layout_rows: list[dict]
):
    from judges.h1_requirement.schemas import RequirementModelInput

    filters_and_parameters = requirement_layer.get("filters_and_parameters") or {}
    common_rules = requirement_layer.get("common_rules") or {}
    file_specs = requirement_layer.get("file_specs") or {}
    scope_payload = requirement_layer.get("scope") or {}

    explicit_filters: list[dict[str, Any]] = []
    for key, value in filters_and_parameters.items():
        if value in (None, "", [], {}):
            continue
        operator = "exclude" if str(key).lower().startswith("excluded_") else "include"
        explicit_filters.append(
            {
                "field": key,
                "operator": operator,
                "values": _normalize_filter_values(value),
                "source": f"filters_and_parameters.{key}",
            }
        )

    output_fields: list[dict[str, Any]] = []
    for index, row in enumerate(layout_rows, start=1):
        output_fields.append(
            {
                "field_name": (
                    row.get("Field Name")
                    or row.get("field_name")
                    or row.get("Field")
                    or row.get("name")
                ),
                "position": row.get("__position__", index),
                "data_type": row.get("Data Type") or row.get("data_type"),
                "description": (
                    row.get("Field Description")
                    or row.get("Description")
                    or row.get("description")
                    or ""
                ),
                "required": (
                    row.get("Required / Optional / Conditional")
                    or row.get("Requirement")
                    or row.get("requirement")
                ),
                "source_table": row.get("__table_name__"),
            }
        )

    compliance_text = "\n".join(
        [
            str(requirement_layer.get("bsa_input") or ""),
            str(requirement_layer.get("requirements") or ""),
        ]
    ).lower()
    compliance_flags = [
        token
        for token in ["HIPAA", "GDPR", "SOX", "PHI", "PII", "COBRA"]
        if token.lower() in compliance_text
    ]

    primary_domain, sub_domain = _guess_domain(requirement_layer)
    date_parameters = filters_and_parameters.get("date_parameters") or {}
    date_range = date_parameters or {
        "effective_dates_from": common_rules.get("effective_dates_from"),
        "effective_dates_to": common_rules.get("effective_dates_to"),
        "posted_dates_from": common_rules.get("posted_dates_from"),
        "posted_dates_to": common_rules.get("posted_dates_to"),
    }
    agent_notes = [
        str(value)
        for value in [common_rules.get("comments"), file_specs.get("assumptions")]
        if value
    ]

    return RequirementModelInput(
        extract_purpose=str(
            requirement_layer.get("requirements")
            or requirement_layer.get("bsa_input")
            or ""
        ).strip(),
        scope={
            "company": filters_and_parameters.get("company")
            or scope_payload.get("in_scope")
            or "",
            "LOB": filters_and_parameters.get("line_of_business")
            or file_specs.get("file_population_type")
            or "",
            "funding": filters_and_parameters.get("financial_arrangement") or "",
            "date_range": date_range,
            "constraints": scope_payload.get("out_of_scope") or "",
        },
        explicit_filters=explicit_filters,
        compliance_flags=compliance_flags,
        stakeholder_references=[],
        output_fields=output_fields,
        total_field_count=len(output_fields),
        implicit_rules=[],
        conflicts_with_brd=[],
        ambiguities=[],
        blocking_count=0,
        primary_domain=primary_domain,
        sub_domain=sub_domain,
        domain_confidence=0.75,
        complexity_score=max(1, 1 + (len(output_fields) // 20) + (len(explicit_filters) // 5)),
        recommended_catalogs=[primary_domain] if primary_domain != "Other" else [],
        confidence_score=0.8,
        agent_notes=agent_notes,
    )


def _build_brd_text(requirement_layer: Dict[str, Any]) -> str:
    bsa_input = str(requirement_layer.get("bsa_input") or "").strip()
    if bsa_input:
        return bsa_input

    sections = [
        f"Scope In: {(requirement_layer.get('scope') or {}).get('in_scope', '')}",
        f"Scope Out: {(requirement_layer.get('scope') or {}).get('out_of_scope', '')}",
        f"Requirements: {requirement_layer.get('requirements', '')}",
        f"Filters: {json.dumps(requirement_layer.get('filters_and_parameters') or {}, indent=2)}",
        f"Common Rules: {json.dumps(requirement_layer.get('common_rules') or {}, indent=2)}",
        f"File Specs: {json.dumps(requirement_layer.get('file_specs') or {}, indent=2)}",
    ]
    return "\n\n".join(section for section in sections if section.strip())


async def _load_transcript_texts(transcript_gcs_uri: str | None) -> list[str]:
    if not transcript_gcs_uri:
        return []

    from fastapi.concurrency import run_in_threadpool
    from utils.gcs_artifact_utils import download_text, parse_gcs_uri

    _, object_name = parse_gcs_uri(transcript_gcs_uri)
    raw_text = await run_in_threadpool(download_text, object_name=object_name)
    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        return [raw_text]

    if isinstance(parsed, dict):
        if isinstance(parsed.get("transcript_texts"), list):
            return [str(item) for item in parsed["transcript_texts"]]
        parsed_transcript = parsed.get("parsed_transcript") or {}
        decisions = parsed_transcript.get("decisions") or []
        if decisions:
            return [
                "\n".join(
                    str(item.get("decision_text") or "") for item in decisions if item
                )
            ]
    if isinstance(parsed, list):
        return [str(item) for item in parsed]
    return [raw_text]


@router.post("/judge-h1", response_model=JudgeH1Response)
async def run_requirement_judge_h1(request: JudgeH1Request):
    try:
        from fastapi.concurrency import run_in_threadpool
        from judges.h1_requirement.post_judge import PostJudgeH1
        from judges.h1_requirement.pre_judge import PreJudgeH1
        from judges.h1_requirement.schemas import JudgeInputH1
        from utils.gcs_artifact_utils import download_json_uri, download_text, parse_gcs_uri

        judge_mode = request.judge_mode.strip().lower()
        if judge_mode not in {"pre", "post"}:
            raise HTTPException(
                status_code=422, detail="judge_mode must be 'pre' or 'post'."
            )
        if judge_mode == "post" and not (request.bsa_rejection_feedback or "").strip():
            raise HTTPException(
                status_code=422,
                detail="bsa_rejection_feedback is required when judge_mode='post'.",
            )

        brd_content = await run_in_threadpool(download_json_uri, request.brd_gcs_uri)
        layout_content = await run_in_threadpool(
            download_json_uri, request.layout_gcs_uri
        )
        transcript_texts = await _load_transcript_texts(request.transcript_gcs_uri)

        requirement_layer = _extract_requirement_layer_payload(brd_content)
        layout_rows = _extract_layout_rows(layout_content)
        requirement_model = _build_judge_requirement_model(
            requirement_layer, layout_rows
        )
        if request.brd_markdown_gcs_uri:
            _, object_name = parse_gcs_uri(request.brd_markdown_gcs_uri)
            if object_name and object_name.strip():
                brd_text = await run_in_threadpool(download_text, object_name=object_name)
            else:
                logger.warning("brd_markdown_gcs_uri has no object path, falling back to brd_text from requirement_layer")
                brd_text = _build_brd_text(requirement_layer)
        else:
            brd_text = _build_brd_text(requirement_layer)

        layout_text = None
        if request.layout_markdown_gcs_uri:
            _, object_name = parse_gcs_uri(request.layout_markdown_gcs_uri)
            if object_name and object_name.strip():
                layout_text = await run_in_threadpool(download_text, object_name=object_name)
            else:
                logger.warning("layout_markdown_gcs_uri has no object path, skipping layout_text download")

        judge_input = JudgeInputH1(
            session_id=request.session_id,
            requirement_model=requirement_model,
            brd_text=brd_text,
            layout_text=layout_text,
            layout_raw=layout_rows,
            transcript_texts=transcript_texts,
            bsa_rejection_feedback=request.bsa_rejection_feedback,
            revision_number=request.revision_number,
        )

        judge_output = (
            await PreJudgeH1().evaluate(judge_input)
            if judge_mode == "pre"
            else await PostJudgeH1().evaluate(judge_input)
        )

        orchestrator = ExtractPipelineOrchestrator()
        session_state = await orchestrator._load_session_state(
            request.user_id, request.session_id
        )
        judge_state = session_state.get(REQUIREMENT_JUDGE_STATE_KEY, {}) or {}
        judge_state.update(
            {
                "source_artifacts": {
                    "brd_gcs_uri": request.brd_gcs_uri,
                    "layout_gcs_uri": request.layout_gcs_uri,
                    "transcript_gcs_uri": request.transcript_gcs_uri,
                },
                "requirement_model": requirement_model.model_dump(mode="json"),
                "annotated_artifact": judge_output.annotated_artifact,
                "bsa_review_summary": judge_output.bsa_review_summary,
            }
        )
        if judge_mode == "pre":
            judge_state["pre_judge_evaluation"] = judge_output.evaluation.model_dump(
                mode="json"
            )
        else:
            judge_state["post_judge_evaluation"] = judge_output.evaluation.model_dump(
                mode="json"
            )
            if judge_output.revision_directive is not None:
                judge_state["revision_directive"] = (
                    judge_output.revision_directive.model_dump(mode="json")
                )
        await orchestrator._update_session_state(
            request.user_id,
            request.session_id,
            {REQUIREMENT_JUDGE_STATE_KEY: judge_state},
        )

        return JudgeH1Response(
            success=True,
            session_id=request.session_id,
            judge_mode=judge_mode,
            message=(
                "H1 pre-judge completed successfully."
                if judge_mode == "pre"
                else "H1 post-judge completed successfully."
            ),
            source_artifacts={
                "brd_gcs_uri": request.brd_gcs_uri,
                "layout_gcs_uri": request.layout_gcs_uri,
                "transcript_gcs_uri": request.transcript_gcs_uri,
            },
            normalized_requirement_model=requirement_model.model_dump(mode="json"),
            evaluation=judge_output.evaluation.model_dump(mode="json"),
            revision_directive=(
                judge_output.revision_directive.model_dump(mode="json")
                if judge_output.revision_directive is not None
                else None
            ),
            annotated_artifact=judge_output.annotated_artifact,
            bsa_review_summary=judge_output.bsa_review_summary,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(
            "run_requirement_judge_h1 failed for session=%s", request.session_id
        )
        raise HTTPException(status_code=500, detail=str(exc))


# ─── H2 Driver Judge ─────────────────────────────────────────────────────────


class JudgeH2Request(BaseModel):
    user_id: str
    session_id: str
    brd_gcs_uri: str
    layout_gcs_uri: str
    driver_criteria_gcs_uri: Optional[str] = None
    driver_criteria: Optional[Dict[str, Any]] = None
    h1_requirement_model: Optional[Dict[str, Any]] = None
    standards_dictionary: Optional[Dict[str, str]] = None
    brd_markdown_gcs_uri: Optional[str] = None
    judge_mode: str = "pre"
    bsa_rejection_feedback: Optional[str] = None
    revision_number: int = 0


class JudgeH2Response(BaseModel):
    success: bool
    session_id: str
    judge_mode: str
    message: str
    source_artifacts: Dict[str, Any]
    normalized_driver_criteria: Dict[str, Any]
    evaluation: Dict[str, Any]
    revision_directive: Optional[Dict[str, Any]] = None
    annotated_driver: Dict[str, Any] = Field(default_factory=dict)
    sql_analysis_report: Dict[str, Any] = Field(default_factory=dict)
    bsa_review_summary: str


def _default_standards_dictionary() -> Dict[str, str]:
    """Fallback standards mapping used when caller does not supply one.
    In production this should be sourced from the ADW Standards catalog."""
    return {
        "company": "CO_CD_ROLLUP_ID",
        "line of business": "MED_LOB_ROLLUP_ID",
        "lob": "MED_LOB_ROLLUP_ID",
        "ibc focus": "IBC_FOC_LVL_CD",
        "funding": "FUNDING_TYPE_CD",
        "funding type": "FUNDING_TYPE_CD",
        "coverage type": "CVG_CTG_CD",
        "coverage category": "CVG_CTG_CD",
        "enrollment status": "ENRL_STAT_CD",
        "status": "ENRL_STAT_CD",
        "effective date": "ENRL_EFF_DT",
        "termination date": "ENRL_TERM_DT",
        "group": "GRP_LGL_ENTITY_CD",
        "member id": "MBR_ID",
        "exclude fep": "CO_CD_ROLLUP_ID",
    }


def _build_h1_requirement_model_from_layer(requirement_layer: Dict[str, Any]) -> Dict[str, Any]:
    """Convert the H1 requirement_layer payload into the H1ApprovedRequirementModel shape."""
    filters_and_parameters = requirement_layer.get("filters_and_parameters") or {}
    scope_payload = requirement_layer.get("scope") or {}

    explicit_filters: list[dict] = []
    for key, value in filters_and_parameters.items():
        if value in (None, "", [], {}):
            continue
        operator = "exclude" if str(key).lower().startswith("excluded_") else "include"
        explicit_filters.append(
            {
                "field": key,
                "operator": operator,
                "values": value if isinstance(value, list) else [value],
                "source_text": str(value),
                "source": f"filters_and_parameters.{key}",
            }
        )

    domain, _ = _guess_domain(requirement_layer)
    return {
        "extract_purpose": str(
            requirement_layer.get("requirements") or requirement_layer.get("bsa_input") or ""
        ).strip(),
        "scope": {
            "company": filters_and_parameters.get("company") or scope_payload.get("in_scope") or "",
            "LOB": filters_and_parameters.get("line_of_business") or "",
            "funding": filters_and_parameters.get("financial_arrangement") or "",
            "constraints": scope_payload.get("out_of_scope") or "",
        },
        "explicit_filters": explicit_filters,
        "implicit_rules": [],
        "ambiguities": [],
        "primary_domain": domain,
        "complexity_score": max(1, len(explicit_filters)),
        "bsa_h1_resolutions": requirement_layer.get("bsa_h1_resolutions") or {},
    }


def _normalize_driver_criteria_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Coerce arbitrary driver outputs into the DriverCriteriaInput schema shape."""
    if not isinstance(payload, dict):
        return {"where_clause": "", "predicates": []}

    where_clause = (
        payload.get("where_clause")
        or payload.get("sql_where_clause")
        or ""
    )
    predicates_raw = (
        payload.get("predicates")
        or payload.get("global_filters")
        or payload.get("filters")
        or []
    )

    predicates: list[dict] = []
    if isinstance(predicates_raw, list):
        for entry in predicates_raw:
            if not isinstance(entry, dict):
                continue
            predicates.append(
                {
                    "id": entry.get("id") or entry.get("filter_id"),
                    "business_field": entry.get("business_field") or entry.get("business_term"),
                    "standard_field": entry.get("standard_field") or entry.get("field"),
                    "operator": entry.get("operator"),
                    "values": entry.get("values"),
                    "direction": entry.get("direction"),
                    "brd_source_text": entry.get("brd_source_text") or entry.get("source_text"),
                    "brd_section": entry.get("brd_section") or entry.get("section"),
                    "fyi_used": entry.get("fyi_used", False),
                    "parameterization_applied": entry.get("parameterization_applied", False),
                    "raw": entry.get("raw") or entry.get("predicate") or entry.get("sql"),
                }
            )

    return {
        "where_clause": where_clause,
        "predicates": predicates,
        "normalized_filters": payload.get("normalized_filters") or [],
        "incomplete_filters": payload.get("incomplete_filters") or [],
        "unmapped_fields": payload.get("unmapped_fields") or [],
        "fyi_lookups": payload.get("fyi_lookups") or [],
        "logic_connectors": payload.get("logic_connectors") or {},
        "estimated_row_impact": payload.get("estimated_row_impact"),
        "activated_rules": payload.get("activated_rules") or [],
        "bypassed_rules": payload.get("bypassed_rules") or [],
        "validation_passed": bool(payload.get("validation_passed", True)),
        "validation_notes": payload.get("validation_notes") or [],
        "confidence_score": float(payload.get("confidence_score") or 0.0),
        "agent_notes": payload.get("agent_notes") or [],
    }


@router.post("/judge-h2", response_model=JudgeH2Response)
async def run_driver_judge_h2(request: JudgeH2Request):
    try:
        from fastapi.concurrency import run_in_threadpool
        from judges.h2_driver.post_judge import PostJudgeH2
        from judges.h2_driver.pre_judge import PreJudgeH2
        from judges.h2_driver.schemas import (
            DriverCriteriaInput,
            H1ApprovedRequirementModel,
            JudgeInputH2,
        )
        from utils.gcs_artifact_utils import (
            download_json_uri,
            download_text,
            parse_gcs_uri,
        )

        judge_mode = request.judge_mode.strip().lower()
        if judge_mode not in {"pre", "post"}:
            raise HTTPException(
                status_code=422, detail="judge_mode must be 'pre' or 'post'."
            )
        if judge_mode == "post" and not (request.bsa_rejection_feedback or "").strip():
            raise HTTPException(
                status_code=422,
                detail="bsa_rejection_feedback is required when judge_mode='post'.",
            )

        # Resolve driver_criteria
        if request.driver_criteria is not None:
            driver_payload = request.driver_criteria
        elif request.driver_criteria_gcs_uri:
            driver_payload = await run_in_threadpool(
                download_json_uri, request.driver_criteria_gcs_uri
            )
        else:
            raise HTTPException(
                status_code=422,
                detail="Either driver_criteria or driver_criteria_gcs_uri must be supplied.",
            )
        normalized_driver = _normalize_driver_criteria_payload(driver_payload)

        # Resolve H1 requirement model — explicit or derived from BRD JSON
        brd_content = await run_in_threadpool(download_json_uri, request.brd_gcs_uri)
        requirement_layer = _extract_requirement_layer_payload(brd_content)
        if request.h1_requirement_model is not None:
            h1_payload = request.h1_requirement_model
        else:
            h1_payload = _build_h1_requirement_model_from_layer(requirement_layer)

        # BRD text
        if request.brd_markdown_gcs_uri:
            _, object_name = parse_gcs_uri(request.brd_markdown_gcs_uri)
            brd_text = await run_in_threadpool(download_text, object_name=object_name)
        else:
            brd_text = _build_brd_text(requirement_layer)

        # Standards dictionary
        standards_dictionary = (
            request.standards_dictionary or _default_standards_dictionary()
        )

        judge_input = JudgeInputH2(
            session_id=request.session_id,
            driver_criteria=DriverCriteriaInput(**normalized_driver),
            h1_requirement_model=H1ApprovedRequirementModel(**h1_payload),
            brd_text=brd_text,
            standards_dictionary=standards_dictionary,
            bsa_rejection_feedback=request.bsa_rejection_feedback,
            revision_number=request.revision_number,
        )

        judge_output = (
            await PreJudgeH2().evaluate(judge_input)
            if judge_mode == "pre"
            else await PostJudgeH2().evaluate(judge_input)
        )

        orchestrator = ExtractPipelineOrchestrator()
        session_state = await orchestrator._load_session_state(
            request.user_id, request.session_id
        )
        judge_state = session_state.get(DRIVER_JUDGE_STATE_KEY, {}) or {}
        judge_state.update(
            {
                "source_artifacts": {
                    "brd_gcs_uri": request.brd_gcs_uri,
                    "layout_gcs_uri": request.layout_gcs_uri,
                    "driver_criteria_gcs_uri": request.driver_criteria_gcs_uri,
                },
                "driver_criteria": normalized_driver,
                "h1_requirement_model": h1_payload,
                "standards_dictionary_used": standards_dictionary,
                "annotated_driver": judge_output.annotated_driver,
                "sql_analysis_report": judge_output.sql_analysis_report,
                "bsa_review_summary": judge_output.bsa_review_summary,
            }
        )
        if judge_mode == "pre":
            judge_state["pre_judge_evaluation"] = judge_output.evaluation.model_dump(
                mode="json"
            )
        else:
            judge_state["post_judge_evaluation"] = judge_output.evaluation.model_dump(
                mode="json"
            )
            if judge_output.revision_directive is not None:
                judge_state["revision_directive"] = (
                    judge_output.revision_directive.model_dump(mode="json")
                )
        await orchestrator._update_session_state(
            request.user_id,
            request.session_id,
            {DRIVER_JUDGE_STATE_KEY: judge_state},
        )

        return JudgeH2Response(
            success=True,
            session_id=request.session_id,
            judge_mode=judge_mode,
            message=(
                "H2 pre-judge completed successfully."
                if judge_mode == "pre"
                else "H2 post-judge completed successfully."
            ),
            source_artifacts={
                "brd_gcs_uri": request.brd_gcs_uri,
                "layout_gcs_uri": request.layout_gcs_uri,
                "driver_criteria_gcs_uri": request.driver_criteria_gcs_uri,
            },
            normalized_driver_criteria=normalized_driver,
            evaluation=judge_output.evaluation.model_dump(mode="json"),
            revision_directive=(
                judge_output.revision_directive.model_dump(mode="json")
                if judge_output.revision_directive is not None
                else None
            ),
            annotated_driver=judge_output.annotated_driver,
            sql_analysis_report=judge_output.sql_analysis_report,
            bsa_review_summary=judge_output.bsa_review_summary,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(
            "run_driver_judge_h2 failed for session=%s", request.session_id
        )
        raise HTTPException(status_code=500, detail=str(exc))


# ─── Driver Pipeline Judge ───────────────────────────────────────────────────

DRIVER_PIPELINE_JUDGE_STATE_KEY = "driver_pipeline_judge_state"


class JudgeDriverRequest(BaseModel):
    """
    Payload for the driver pipeline judge.

    Pass the 3 step outputs directly — the judge does not re-run the pipeline
    or read from session state.  brd_uri is required so the judge can download
    the BRD requirement-layer JSON and cross-check what the BRD demanded against
    what the agents produced.
    """
    userId: str
    sessionId: str                              # used only for logging / audit
    brd_uri: str                                # GCS URI of validated_requirement_layer JSON
    driver_mapping: Dict[str, Any]              # output of business_mapping_agent (Step 1)
    driver_logic: Dict[str, Any]                # output of logic_builder_agent (Step 2)
    driver_validation: Dict[str, Any]           # output of driver_validator_agent (Step 3)
    revision_number: int = 0


class StepJudgmentResponse(BaseModel):
    step: str
    verdict: str
    score: float
    summary: str
    findings: List[str] = Field(default_factory=list)
    recommendations: List[str] = Field(default_factory=list)
    details: Dict[str, Any] = Field(default_factory=dict)


class JudgeDriverResponse(BaseModel):
    success: bool
    session_id: str
    overall_verdict: str
    overall_score: float
    overall_summary: str
    can_proceed: bool
    step_judgments: List[StepJudgmentResponse]
    recommendations: List[str] = Field(default_factory=list)
    bsa_review_summary: str
    judged_at: str
    steps_available: Dict[str, bool] = Field(default_factory=dict)
    quality_scorecard: Dict[str, Any] = Field(default_factory=dict)
    rule_scores: List[Dict[str, Any]] = Field(default_factory=list)


def _build_brd_context_from_layer(requirement_layer: Dict[str, Any]):
    """Build BrdContext from a normalized requirement layer dict."""
    from judges.h2_driver.schemas import BrdContext

    fp = requirement_layer.get("filters_and_parameters") or {}
    scope = requirement_layer.get("scope") or {}
    in_scope = (
        fp.get("in_scope")
        or scope.get("in_scope")
        or requirement_layer.get("in_scope")
        or ""
    )
    out_of_scope = (
        fp.get("out_of_scope")
        or scope.get("out_of_scope")
        or requirement_layer.get("out_of_scope")
        or ""
    )

    # Collect non-empty filter keys (excludes date_parameters sub-object itself)
    active_keys: List[str] = []
    for k, v in fp.items():
        if k == "date_parameters":
            # Expand date sub-keys
            if isinstance(v, dict):
                active_keys.extend(sk for sk, sv in v.items() if sv)
            continue
        if v not in (None, "", [], {}):
            active_keys.append(k)

    return BrdContext(
        in_scope=str(in_scope),
        out_of_scope=str(out_of_scope),
        requirements=str(requirement_layer.get("requirements") or ""),
        filters_and_parameters=fp,
        active_filter_keys=active_keys,
    )


@router.post("/judge-driver", response_model=JudgeDriverResponse)
async def run_driver_pipeline_judge(request: JudgeDriverRequest):
    """
    Unified LLM judge for the 3-step driver generation pipeline.

    Payload:
      userId, sessionId, brd_uri,
      driver_mapping    (Step 1 output),
      driver_logic      (Step 2 output),
      driver_validation (Step 3 output)

    The caller passes the step outputs directly — no session state is read
    and the pipeline is not re-run.  brd_uri is downloaded once to supply
    the BRD requirement-layer context used for cross-checking.

    Returns per-step judgments (PASS/WARN/BLOCK) and an overall verdict.
    """
    try:
        from fastapi.concurrency import run_in_threadpool
        from judges.h2_driver.pipeline_judge import DriverPipelineJudge
        from judges.h2_driver.schemas import BrdContext, DriverPipelineJudgeInput

        # 1. Download BRD JSON → build cross-check context
        try:
            from utils.gcs_artifact_utils import download_json_uri
            brd_content = await run_in_threadpool(download_json_uri, request.brd_uri)
            requirement_layer = _extract_requirement_layer_payload(brd_content)
            brd_context = _build_brd_context_from_layer(requirement_layer)
        except Exception as brd_exc:
            logger.warning(
                "judge-driver: BRD load failed — judging without BRD cross-check. Error: %s",
                brd_exc,
            )
            brd_context = BrdContext()

        # 2. Run judge directly on the provided step outputs
        judge_input = DriverPipelineJudgeInput(
            session_id=request.sessionId,
            driver_mapping=request.driver_mapping,
            driver_logic=request.driver_logic,
            driver_validation=request.driver_validation,
            brd_context=brd_context,
            revision_number=request.revision_number,
        )
        judge_output = await DriverPipelineJudge().evaluate(judge_input)

        return JudgeDriverResponse(
            success=True,
            session_id=request.sessionId,
            overall_verdict=judge_output.overall_verdict,
            overall_score=judge_output.overall_score,
            overall_summary=judge_output.overall_summary,
            can_proceed=judge_output.can_proceed,
            step_judgments=[
                StepJudgmentResponse(**sj.model_dump(mode="json"))
                for sj in judge_output.step_judgments
            ],
            recommendations=judge_output.recommendations,
            bsa_review_summary=judge_output.bsa_review_summary,
            judged_at=judge_output.judged_at,
            steps_available={
                "driver_mapping":    bool(request.driver_mapping),
                "driver_logic":      bool(request.driver_logic),
                "driver_validation": bool(request.driver_validation),
            },
            quality_scorecard=judge_output.quality_scorecard,
            rule_scores=judge_output.rule_scores,
        )

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("run_driver_pipeline_judge failed for session=%s", request.sessionId)
        raise HTTPException(status_code=500, detail=str(exc))


# ─── H5 Metadata Judge ───────────────────────────────────────────────────────


class JudgeH5Request(BaseModel):
    user_id: str
    session_id: str
    brd_gcs_uri: str
    layout_gcs_uri: str
    metadata_output_gcs_uri: Optional[str] = None
    metadata_output: Optional[Dict[str, Any]] = None
    h4_mapping_spec: Optional[Dict[str, Any]] = None
    h4_mapping_gcs_uri: Optional[str] = None
    original_layout_fields: Optional[List[Dict[str, Any]]] = None
    judge_mode: str = "pre"
    bsa_rejection_feedback: Optional[str] = None
    revision_number: int = 0


class JudgeH5Response(BaseModel):
    success: bool
    session_id: str
    judge_mode: str
    message: str
    source_artifacts: Dict[str, Any]
    normalized_metadata_output: Dict[str, Any]
    evaluation: Dict[str, Any]
    revision_directive: Optional[Dict[str, Any]] = None
    annotated_metadata: Dict[str, Any] = Field(default_factory=dict)
    quality_scorecard: Dict[str, Any] = Field(default_factory=dict)
    auto_corrected_output: Dict[str, Any] = Field(default_factory=dict)
    bsa_review_summary: str


def _layout_rows_to_layout_fields(layout_rows: list[dict]) -> list[Dict[str, Any]]:
    fields: list[Dict[str, Any]] = []
    for index, row in enumerate(layout_rows or [], start=1):
        fields.append(
            {
                "field_name": (
                    row.get("Field Name")
                    or row.get("field_name")
                    or row.get("Field")
                    or row.get("name")
                ),
                "position": row.get("__position__", index),
                "data_type": row.get("Data Type") or row.get("data_type"),
                "description": (
                    row.get("Field Description")
                    or row.get("Description")
                    or row.get("description")
                    or ""
                ),
                "required": (
                    row.get("Required / Optional / Conditional")
                    or row.get("Requirement")
                    or row.get("requirement")
                ),
                "source_table": row.get("__table_name__"),
            }
        )
    return fields


def _normalize_metadata_output_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Coerce arbitrary metadata-builder outputs into MetadataBuildOutput shape."""
    if not isinstance(payload, dict):
        return {
            "file_metadata": {},
            "attributes": [],
            "indimap_template_json": "",
        }

    file_metadata = payload.get("file_metadata") or payload.get("filespecs") or {}
    attributes_raw = (
        payload.get("attributes")
        or payload.get("file1", {}).get("attributes")
        or payload.get("metadata_attributes")
        or []
    )

    template = payload.get("indimap_template_json")
    if isinstance(template, dict):
        template = json.dumps(template)
    elif template is None:
        template = ""

    return {
        "file_metadata": file_metadata,
        "attributes": attributes_raw,
        "naming_conformance_score": float(payload.get("naming_conformance_score") or 0.0),
        "type_conformance_score": float(payload.get("type_conformance_score") or 0.0),
        "completeness_score": float(payload.get("completeness_score") or 0.0),
        "indimap_template_json": template,
        "naming_auto_corrections": payload.get("naming_auto_corrections") or [],
        "naming_manual_flags": payload.get("naming_manual_flags") or [],
        "type_casts_applied": payload.get("type_casts_applied") or [],
        "type_cast_warnings": payload.get("type_cast_warnings") or [],
        "confidence_score": float(payload.get("confidence_score") or 0.0),
        "agent_notes": payload.get("agent_notes") or [],
    }


def _normalize_h4_mapping_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Coerce arbitrary H4 outputs into H4ApprovedMappingSpec shape."""
    if not isinstance(payload, dict):
        return {
            "session_id": "",
            "fields": [],
            "total_field_count": 0,
            "no_match_fields": [],
            "bsa_h4_overrides": {},
        }
    fields = payload.get("fields") or payload.get("mappings") or []
    return {
        "session_id": str(payload.get("session_id") or ""),
        "fields": fields if isinstance(fields, list) else [],
        "total_field_count": int(payload.get("total_field_count") or len(fields or [])),
        "no_match_fields": payload.get("no_match_fields") or [],
        "bsa_h4_overrides": payload.get("bsa_h4_overrides") or {},
    }


@router.post("/judge-h5", response_model=JudgeH5Response)
async def run_metadata_judge_h5(request: JudgeH5Request):
    try:
        from fastapi.concurrency import run_in_threadpool
        from judges.h5_metadata.post_judge import PostJudgeH5
        from judges.h5_metadata.pre_judge import PreJudgeH5
        from judges.h5_metadata.schemas import (
            H4ApprovedMappingSpec,
            JudgeInputH5,
            MetadataBuildOutput,
        )
        from utils.gcs_artifact_utils import download_json_uri

        judge_mode = request.judge_mode.strip().lower()
        if judge_mode not in {"pre", "post"}:
            raise HTTPException(
                status_code=422, detail="judge_mode must be 'pre' or 'post'."
            )
        if judge_mode == "post" and not (request.bsa_rejection_feedback or "").strip():
            raise HTTPException(
                status_code=422,
                detail="bsa_rejection_feedback is required when judge_mode='post'.",
            )

        # Resolve metadata output
        if request.metadata_output is not None:
            metadata_payload = request.metadata_output
        elif request.metadata_output_gcs_uri:
            metadata_payload = await run_in_threadpool(
                download_json_uri, request.metadata_output_gcs_uri
            )
        else:
            raise HTTPException(
                status_code=422,
                detail="Either metadata_output or metadata_output_gcs_uri must be supplied.",
            )
        normalized_metadata = _normalize_metadata_output_payload(metadata_payload)

        # Resolve H4 mapping spec
        if request.h4_mapping_spec is not None:
            h4_payload = request.h4_mapping_spec
        elif request.h4_mapping_gcs_uri:
            h4_payload = await run_in_threadpool(
                download_json_uri, request.h4_mapping_gcs_uri
            )
        else:
            h4_payload = {"session_id": request.session_id, "fields": []}
        normalized_h4 = _normalize_h4_mapping_payload(h4_payload)

        # Resolve original layout fields
        if request.original_layout_fields is not None:
            layout_fields = request.original_layout_fields
        else:
            layout_content = await run_in_threadpool(
                download_json_uri, request.layout_gcs_uri
            )
            layout_rows = _extract_layout_rows(layout_content)
            layout_fields = _layout_rows_to_layout_fields(layout_rows)

        judge_input = JudgeInputH5(
            session_id=request.session_id,
            metadata_output=MetadataBuildOutput(**normalized_metadata),
            h4_mapping_spec=H4ApprovedMappingSpec(**normalized_h4),
            original_layout_fields=layout_fields,
            bsa_rejection_feedback=request.bsa_rejection_feedback,
            revision_number=request.revision_number,
        )

        judge_output = (
            await PreJudgeH5().evaluate(judge_input)
            if judge_mode == "pre"
            else await PostJudgeH5().evaluate(judge_input)
        )

        orchestrator = ExtractPipelineOrchestrator()
        session_state = await orchestrator._load_session_state(
            request.user_id, request.session_id
        )
        judge_state = session_state.get(METADATA_JUDGE_STATE_KEY, {}) or {}
        judge_state.update(
            {
                "source_artifacts": {
                    "brd_gcs_uri": request.brd_gcs_uri,
                    "layout_gcs_uri": request.layout_gcs_uri,
                    "metadata_output_gcs_uri": request.metadata_output_gcs_uri,
                    "h4_mapping_gcs_uri": request.h4_mapping_gcs_uri,
                },
                "metadata_output": normalized_metadata,
                "h4_mapping_spec": normalized_h4,
                "annotated_metadata": judge_output.annotated_metadata,
                "quality_scorecard": judge_output.quality_scorecard,
                "auto_corrected_output": judge_output.auto_corrected_output,
                "bsa_review_summary": judge_output.bsa_review_summary,
            }
        )
        if judge_mode == "pre":
            judge_state["pre_judge_evaluation"] = judge_output.evaluation.model_dump(
                mode="json"
            )
        else:
            judge_state["post_judge_evaluation"] = judge_output.evaluation.model_dump(
                mode="json"
            )
            if judge_output.revision_directive is not None:
                judge_state["revision_directive"] = (
                    judge_output.revision_directive.model_dump(mode="json")
                )
        await orchestrator._update_session_state(
            request.user_id,
            request.session_id,
            {METADATA_JUDGE_STATE_KEY: judge_state},
        )

        return JudgeH5Response(
            success=True,
            session_id=request.session_id,
            judge_mode=judge_mode,
            message=(
                "H5 pre-judge completed successfully."
                if judge_mode == "pre"
                else "H5 post-judge completed successfully."
            ),
            source_artifacts={
                "brd_gcs_uri": request.brd_gcs_uri,
                "layout_gcs_uri": request.layout_gcs_uri,
                "metadata_output_gcs_uri": request.metadata_output_gcs_uri,
                "h4_mapping_gcs_uri": request.h4_mapping_gcs_uri,
            },
            normalized_metadata_output=normalized_metadata,
            evaluation=judge_output.evaluation.model_dump(mode="json"),
            revision_directive=(
                judge_output.revision_directive.model_dump(mode="json")
                if judge_output.revision_directive is not None
                else None
            ),
            annotated_metadata=judge_output.annotated_metadata,
            quality_scorecard=judge_output.quality_scorecard,
            auto_corrected_output=judge_output.auto_corrected_output,
            bsa_review_summary=judge_output.bsa_review_summary,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(
            "run_metadata_judge_h5 failed for session=%s", request.session_id
        )
        raise HTTPException(status_code=500, detail=str(exc))


# ─── Metadata Pipeline Judge ─────────────────────────────────────────────────


class JudgeMetadataRequest(BaseModel):
    """
    Payload for the metadata extraction judge.

    Pass the extractor output directly — the judge does not re-run any agents.
    brd_uri and layout_uri are downloaded once to supply cross-check context.
    """

    userId: str
    sessionId: str                        # used only for logging / audit
    brd_uri: str                          # GCS URI of validated_requirement_layer JSON
    layout_uri: str                       # GCS URI of layout JSON
    extracted_metadata: Dict[str, Any]    # output of metadata_extractor_agent
    revision_number: int = 0


class MetadataStepJudgmentResponse(BaseModel):
    step: str
    verdict: str
    score: float
    summary: str
    findings: List[str] = Field(default_factory=list)
    recommendations: List[str] = Field(default_factory=list)
    details: Dict[str, Any] = Field(default_factory=dict)


class JudgeMetadataResponse(BaseModel):
    success: bool
    session_id: str
    overall_verdict: str
    overall_score: float
    overall_summary: str
    can_proceed: bool
    step_judgments: List[MetadataStepJudgmentResponse]
    recommendations: List[str] = Field(default_factory=list)
    bsa_review_summary: str
    judged_at: str
    steps_available: Dict[str, bool] = Field(default_factory=dict)
    quality_scorecard: Dict[str, Any] = Field(default_factory=dict)
    rule_scores: List[Dict[str, Any]] = Field(default_factory=list)


def _parse_layout_columns(layout_content: Any) -> List[Dict[str, Any]]:
    """Extract a flat list of column dicts from a layout GCS artifact."""
    if isinstance(layout_content, list):
        return layout_content
    if isinstance(layout_content, dict):
        for key in ("columns", "fields", "attributes", "layout", "rows", "data"):
            val = layout_content.get(key)
            if isinstance(val, list):
                return val
        # file1.attributes structure
        file1 = layout_content.get("file1") or {}
        if isinstance(file1, dict) and isinstance(file1.get("attributes"), list):
            return file1["attributes"]
    return []


@router.post("/judge-metadata", response_model=JudgeMetadataResponse)
async def run_metadata_pipeline_judge(request: JudgeMetadataRequest):
    """
    Unified LLM judge for the 2-step metadata agent pipeline.

    Payload:
      userId, sessionId, brd_uri, layout_uri,
      normalized_metadata  (Step 1 — metadata_normalizer_agent output),
      extracted_metadata   (Step 2 — metadata_extractor_agent output)

    The caller passes the step outputs directly — no session state is read
    and no agents are re-run.  brd_uri and layout_uri are downloaded once
    to supply BRD requirement context and the authoritative column list for
    cross-checking extraction completeness.

    Returns per-step judgments (PASS/WARN/BLOCK) and an overall verdict.
    """
    try:
        from fastapi.concurrency import run_in_threadpool
        from judges.h5_metadata.pipeline_judge import MetadataPipelineJudge
        from judges.h5_metadata.schemas import MetadataBrdContext, MetadataPipelineJudgeInput
        from utils.gcs_artifact_utils import download_json_uri

        # 1. Download BRD → build requirement context
        try:
            brd_content = await run_in_threadpool(download_json_uri, request.brd_uri)
            requirement_layer = _extract_requirement_layer_payload(brd_content)
            fp = requirement_layer.get("filters_and_parameters") or {}
            scope = requirement_layer.get("scope") or {}
            brd_context = MetadataBrdContext(
                in_scope=str(
                    fp.get("in_scope")
                    or scope.get("in_scope")
                    or requirement_layer.get("in_scope")
                    or ""
                ),
                out_of_scope=str(
                    fp.get("out_of_scope")
                    or scope.get("out_of_scope")
                    or requirement_layer.get("out_of_scope")
                    or ""
                ),
                requirements=str(requirement_layer.get("requirements") or ""),
                filters_and_parameters=fp,
            )
        except Exception as brd_exc:
            logger.warning(
                "judge-metadata: BRD load failed — judging without BRD context. Error: %s",
                brd_exc,
            )
            brd_context = MetadataBrdContext()

        # 2. Download layout → extract column list
        layout_columns: List[Dict[str, Any]] = []
        try:
            layout_content = await run_in_threadpool(download_json_uri, request.layout_uri)
            layout_columns = _parse_layout_columns(layout_content)
        except Exception as layout_exc:
            logger.warning(
                "judge-metadata: layout load failed — coverage check skipped. Error: %s",
                layout_exc,
            )

        # 3. Build judge input from provided extractor output
        judge_input = MetadataPipelineJudgeInput(
            session_id=request.sessionId,
            extracted_metadata=request.extracted_metadata,
            brd_context=brd_context,
            layout_columns=layout_columns,
            revision_number=request.revision_number,
        )

        # 4. Run judge
        judge_output = await MetadataPipelineJudge().evaluate(judge_input)

        return JudgeMetadataResponse(
            success=True,
            session_id=request.sessionId,
            overall_verdict=judge_output.overall_verdict,
            overall_score=judge_output.overall_score,
            overall_summary=judge_output.overall_summary,
            can_proceed=judge_output.can_proceed,
            step_judgments=[
                MetadataStepJudgmentResponse(**sj.model_dump(mode="json"))
                for sj in judge_output.step_judgments
            ],
            recommendations=judge_output.recommendations,
            bsa_review_summary=judge_output.bsa_review_summary,
            judged_at=judge_output.judged_at,
            steps_available={
                "extracted_metadata": bool(request.extracted_metadata),
            },
            quality_scorecard=judge_output.quality_scorecard,
            rule_scores=judge_output.rule_scores,
        )

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(
            "run_metadata_pipeline_judge failed for session=%s", request.sessionId
        )
        raise HTTPException(status_code=500, detail=str(exc))


# ─── Mapping Pipeline Judge ──────────────────────────────────────────────────


class JudgeMappingRequest(BaseModel):
    """
    Payload for the mapping pipeline judge.

    All cross-check artifacts are GCS URIs already saved under the user's session;
    the mapping_result itself can be passed inline OR via mapping_uri.
    """

    userId: str
    sessionId: str                                    # used only for logging / audit
    brd_uri: str                                      # validated_requirement_layer.json
    driver_uri: str                                   # approved_driver_layer_output.json
    metadata_uri: str                                 # final_metadata_output.json
    mapping_result: Optional[Dict[str, Any]] = None   # inline mapping_result.json
    mapping_uri: Optional[str] = None                 # OR a GCS URI to the same artifact
    revision_number: int = 0


class MappingStepJudgmentResponse(BaseModel):
    step: str
    verdict: str
    score: float
    summary: str
    findings: List[str] = Field(default_factory=list)
    recommendations: List[str] = Field(default_factory=list)
    details: Dict[str, Any] = Field(default_factory=dict)


class JudgeMappingResponse(BaseModel):
    success: bool
    session_id: str
    overall_verdict: str
    overall_score: float
    overall_summary: str
    can_proceed: bool
    step_judgments: List[MappingStepJudgmentResponse]
    recommendations: List[str] = Field(default_factory=list)
    bsa_review_summary: str
    judged_at: str
    artifacts_loaded: Dict[str, bool] = Field(default_factory=dict)
    quality_scorecard: Dict[str, Any] = Field(default_factory=dict)
    rule_scores: List[Dict[str, Any]] = Field(default_factory=list)


def _extract_layout_attributes_from_metadata(metadata_content: Any) -> List[Dict[str, Any]]:
    """Pull the attribute list out of the metadata extraction artifact."""
    if not isinstance(metadata_content, dict):
        return []
    file1 = (
        metadata_content.get("extracted_file1")
        or metadata_content.get("file1")
        or {}
    )
    if isinstance(file1, dict) and isinstance(file1.get("attributes"), list):
        return file1["attributes"]
    if isinstance(metadata_content.get("attributes"), list):
        return metadata_content["attributes"]
    return []


@router.post("/judge-mapping", response_model=JudgeMappingResponse)
async def run_mapping_pipeline_judge(request: JudgeMappingRequest):
    """
    Unified LLM judge for the mapping generation agent.

    Payload:
      userId, sessionId, brd_uri, driver_uri, metadata_uri,
      mapping_result (inline) OR mapping_uri (GCS).

    Downloads BRD, driver, and metadata artifacts from GCS for cross-check,
    then judges the mapping_result against R1-R7 (field coverage, match accuracy,
    transformation correctness, join minimisation, NO MATCH handling, IndiMap
    reuse declaration, transformation/driver separation).

    Returns per-rule scores, KPI scorecard, and overall verdict.
    """
    try:
        from fastapi.concurrency import run_in_threadpool
        from judges.h4_mapping.pipeline_judge import MappingPipelineJudge
        from judges.h4_mapping.schemas import (
            MappingBrdContext,
            MappingPipelineJudgeInput,
        )
        from utils.gcs_artifact_utils import download_json_uri

        artifacts_loaded: Dict[str, bool] = {
            "brd": False,
            "driver": False,
            "metadata": False,
            "mapping_result": False,
        }

        # 1. BRD requirement layer → context
        brd_context = MappingBrdContext()
        try:
            brd_content = await run_in_threadpool(download_json_uri, request.brd_uri)
            requirement_layer = _extract_requirement_layer_payload(brd_content)
            scope = requirement_layer.get("scope") or {}
            brd_context = MappingBrdContext(
                in_scope=str(scope.get("in_scope") or requirement_layer.get("in_scope") or ""),
                out_of_scope=str(
                    scope.get("out_of_scope") or requirement_layer.get("out_of_scope") or ""
                ),
                requirements=str(requirement_layer.get("requirements") or ""),
                common_rules=requirement_layer.get("common_rules") or {},
                file_attributes_mapping=requirement_layer.get("file_attributes_mapping") or {},
            )
            artifacts_loaded["brd"] = True
        except Exception as brd_exc:
            logger.warning("judge-mapping: BRD load failed: %s", brd_exc)

        # 2. Driver layer → common_filter + predicates
        common_filter = ""
        driver_predicates: List[Dict[str, Any]] = []
        try:
            driver_content = await run_in_threadpool(download_json_uri, request.driver_uri)
            common_filter = str(driver_content.get("sql_where_clause") or "")
            driver_predicates = (
                driver_content.get("common_filters")
                or driver_content.get("predicates")
                or []
            )
            artifacts_loaded["driver"] = True
        except Exception as driver_exc:
            logger.warning("judge-mapping: driver load failed: %s", driver_exc)

        # 3. Metadata → layout attributes (source of truth for R1)
        metadata_attributes: List[Dict[str, Any]] = []
        try:
            metadata_content = await run_in_threadpool(
                download_json_uri, request.metadata_uri
            )
            metadata_attributes = _extract_layout_attributes_from_metadata(metadata_content)
            artifacts_loaded["metadata"] = bool(metadata_attributes)
        except Exception as meta_exc:
            logger.warning("judge-mapping: metadata load failed: %s", meta_exc)

        # 4. Mapping result — inline or via URI
        mapping_result: Dict[str, Any] = request.mapping_result or {}
        if not mapping_result and request.mapping_uri:
            try:
                mapping_result = await run_in_threadpool(
                    download_json_uri, request.mapping_uri
                )
            except Exception as map_exc:
                logger.warning("judge-mapping: mapping_uri load failed: %s", map_exc)
        artifacts_loaded["mapping_result"] = bool(mapping_result)

        if not mapping_result:
            raise HTTPException(
                status_code=422,
                detail="Either mapping_result or mapping_uri must resolve to a mapping payload.",
            )

        # 5. Build judge input + run
        judge_input = MappingPipelineJudgeInput(
            session_id=request.sessionId,
            mapping_result=mapping_result,
            brd_context=brd_context,
            common_filter=common_filter,
            driver_predicates=driver_predicates,
            layout_columns=metadata_attributes,
            metadata_attributes=metadata_attributes,
            revision_number=request.revision_number,
        )
        judge_output = await MappingPipelineJudge().evaluate(judge_input)

        return JudgeMappingResponse(
            success=True,
            session_id=request.sessionId,
            overall_verdict=judge_output.overall_verdict,
            overall_score=judge_output.overall_score,
            overall_summary=judge_output.overall_summary,
            can_proceed=judge_output.can_proceed,
            step_judgments=[
                MappingStepJudgmentResponse(**sj.model_dump(mode="json"))
                for sj in judge_output.step_judgments
            ],
            recommendations=judge_output.recommendations,
            bsa_review_summary=judge_output.bsa_review_summary,
            judged_at=judge_output.judged_at,
            artifacts_loaded=artifacts_loaded,
            quality_scorecard=judge_output.quality_scorecard,
            rule_scores=judge_output.rule_scores,
        )

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(
            "run_mapping_pipeline_judge failed for session=%s", request.sessionId
        )
        raise HTTPException(status_code=500, detail=str(exc))


class ReviewMetadataRequest(BaseModel):
    user_id: str
    session_id: str
    status: str
    feedback: Optional[str] = None
    bq_reference: Optional[dict] = None


class ManualUpdateMetadataRequest(BaseModel):
    user_id: str
    session_id: str
    updated_filespecs: Optional[dict] = None
    updated_file1: Optional[dict] = None
    bq_reference: dict


@router.post("/extract-metadata/manual-update")
async def manual_update_metadata(request: ManualUpdateMetadataRequest):
    try:
        from agents.data_map_copilot_agent.sub_agents.metadata_fill_agent.tools.bq_tools import (
            overwrite_filespecs_in_bq,
            append_chunk_to_bq,
        )

        # Overwrite FileSpecs BQ table
        filespecs_table_id = request.bq_reference.get("filespecs_table")
        if filespecs_table_id and request.updated_filespecs:
            rows = [
                {"Field": str(k), "Value": str(v) if v is not None else ""}
                for k, v in request.updated_filespecs.items()
            ]
            overwrite_filespecs_in_bq(json.dumps(rows), filespecs_table_id)

        # Overwrite metadata_template (file1) BQ table
        metadata_table_id = request.bq_reference.get("metadata_table")
        if metadata_table_id and request.updated_file1:
            attributes = request.updated_file1.get("attributes", [])
            if attributes:
                from utils import local_warehouse as _bq
                from utils.bg_query_utils import get_bigquery_client

                client = get_bigquery_client()
                job_config = _bq.LoadJobConfig(
                    write_disposition=_bq.WriteDisposition.WRITE_TRUNCATE,
                    source_format=_bq.SourceFormat.NEWLINE_DELIMITED_JSON,
                )
                load_job = client.load_table_from_json(
                    attributes, metadata_table_id, job_config=job_config
                )
                load_job.result()

        return {
            "success": True,
            "session_id": request.session_id,
            "message": "Manual update successfully saved to BigQuery.",
        }
    except Exception as e:
        logger.exception(f"Error in manual_update_metadata: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/extract-metadata/review")
async def review_extract_metadata(request: ReviewMetadataRequest):
    try:
        from google.genai import types
        from google.adk.apps import App
        from utils.adk_runtime import VertexAiSessionService
        from google.adk import Runner

        from agents.extract_agent.metadata_agent.agent import metadata_extractor_agent

        session_service = VertexAiSessionService(
            project=config.GOOGLE_CLOUD_PROJECT, location=config.GOOGLE_CLOUD_LOCATION
        )

        if request.status.lower() == "approved":
            await _update_step_status(
                session_service=session_service,
                app_name=config.REASONING_ENGINE_RESOURCE,
                user_id=request.user_id,
                session_id=request.session_id,
                step_name="extract_metadata",
                status="approved",
            )
            session = await session_service.get_session(
                app_name=config.REASONING_ENGINE_RESOURCE,
                user_id=request.user_id,
                session_id=request.session_id,
            )
            extracted_filespecs = session.state.get("extracted_filespecs", {})
            extracted_file1 = session.state.get("extracted_file1", {})
            return {
                "success": True,
                "status": "approved",
                "session_id": request.session_id,
                "extracted_filespecs": extracted_filespecs,
                "extracted_file1": extracted_file1,
            }

        elif request.status.lower() == "rejected":
            feedback_msg = f"""
                Your previous extraction was rejected by the reviewer. Please re-evaluate the contents and try again, ensuring all rules are followed.
                {request.feedback.strip() if request.feedback else ""} """

            # Fetch previous data from both BQ tables to provide full context
            if request.bq_reference:
                from agents.data_map_copilot_agent.sub_agents.metadata_fill_agent.tools.bq_tools import (
                    get_bq_table_rows_range,
                )

                prev_context_parts = []
                if "filespecs_table" in request.bq_reference:
                    filespecs_data = get_bq_table_rows_range(
                        request.bq_reference["filespecs_table"], 0, 1000
                    )
                    if "rows" in filespecs_data:
                        prev_context_parts.append(
                            f"Previous FileSpecs data:\n{filespecs_data['rows']}"
                        )

                if "metadata_table" in request.bq_reference:
                    metadata_data = get_bq_table_rows_range(
                        request.bq_reference["metadata_table"], 0, 1000
                    )
                    if "rows" in metadata_data:
                        prev_context_parts.append(
                            f"Previous file1 attribute data:\n{metadata_data['rows']}"
                        )

                if prev_context_parts:
                    prev_context = "\n\n".join(prev_context_parts)
                    feedback_msg = (
                        f"Here is the previously generated metadata:\n{prev_context}\n\n"
                        f"The human reviewer rejected it with the following feedback: {feedback_msg}"
                    )

            message = types.Content(
                role="user", parts=[types.Part.from_text(text=feedback_msg)]
            )

            app = App.model_construct(
                name=config.REASONING_ENGINE_RESOURCE,
                root_agent=metadata_extractor_agent,
            )
            runner = Runner(app=app, session_service=session_service)

            max_retries = 3
            for attempt in range(max_retries):
                try:
                    await wait_for_llm_request_slot(f"metadata_review:{request.session_id}")
                    async for event in runner.run_async(
                        user_id=request.user_id,
                        session_id=request.session_id,
                        new_message=message,
                    ):
                        await record_llm_usage_and_get_wait(
                            event,
                            session_id=f"metadata_review:{request.session_id}",
                            buffer_tokens=300,
                        )
                    break
                except Exception as exc:
                    if is_resource_exhausted_error(exc) and attempt < max_retries - 1:
                        delay = calculate_retry_delay(attempt)
                        logger.warning("metadata_review agent 429 Resource Exhausted (attempt %d/%d). Retrying in %.2fs...", attempt+1, max_retries, delay)
                        await asyncio.sleep(delay)
                    else: raise

            session = await session_service.get_session(
                app_name=config.REASONING_ENGINE_RESOURCE,
                user_id=request.user_id,
                session_id=request.session_id,
            )

            # --- Parse extracted data from session state ---
            extracted_metadata = session.state.get("extracted_metadata", {})
            if isinstance(extracted_metadata, str):
                import re

                text_blob = extracted_metadata.strip()
                pattern = r"```(?:json)?\s*(.*?)\s*```"
                match = re.search(pattern, text_blob, re.DOTALL)
                if match:
                    text_blob = match.group(1).strip()
                try:
                    if text_blob.startswith("{") or text_blob.startswith("["):
                        extracted_metadata = json.loads(text_blob)
                    else:
                        raise ValueError(
                            f"String does not start with JSON object: {text_blob[:50]}"
                        )
                except Exception as e:
                    logger.warning(
                        f"Failed to parse extracted_metadata string as JSON: {e}"
                    )
                    extracted_metadata = {}

            extracted_filespecs = session.state.get("extracted_filespecs", {})
            extracted_file1 = session.state.get("extracted_file1", {})

            if isinstance(extracted_filespecs, str):
                try:
                    extracted_filespecs = json.loads(extracted_filespecs)
                except Exception:
                    extracted_filespecs = {}
            if isinstance(extracted_file1, str):
                try:
                    extracted_file1 = json.loads(extracted_file1)
                except Exception:
                    extracted_file1 = {}

            if not extracted_filespecs and isinstance(extracted_metadata, dict):
                extracted_filespecs = extracted_metadata.get("filespecs", {})
                if (
                    not extracted_filespecs
                    and "filespecs" not in extracted_metadata
                    and "file1" not in extracted_metadata
                ):
                    extracted_filespecs = extracted_metadata
            if not extracted_file1 and isinstance(extracted_metadata, dict):
                extracted_file1 = extracted_metadata.get("file1", {})

            # --- BigQuery Persistence ---
            if request.bq_reference:
                from agents.data_map_copilot_agent.sub_agents.metadata_fill_agent.tools.bq_tools import (
                    overwrite_filespecs_in_bq,
                )

                filespecs_table_id = request.bq_reference.get("filespecs_table")
                if (
                    filespecs_table_id
                    and isinstance(extracted_filespecs, dict)
                    and extracted_filespecs
                ):
                    rows = [
                        {"Field": str(k), "Value": str(v) if v is not None else ""}
                        for k, v in extracted_filespecs.items()
                    ]
                    overwrite_filespecs_in_bq(json.dumps(rows), filespecs_table_id)

                metadata_table_id = request.bq_reference.get("metadata_table")
                if (
                    metadata_table_id
                    and isinstance(extracted_file1, dict)
                    and extracted_file1.get("attributes")
                ):
                    from utils import local_warehouse as _bq
                    from utils.bg_query_utils import get_bigquery_client

                    client = get_bigquery_client()
                    job_config = _bq.LoadJobConfig(
                        write_disposition=_bq.WriteDisposition.WRITE_TRUNCATE,
                        source_format=_bq.SourceFormat.NEWLINE_DELIMITED_JSON,
                    )
                    load_job = client.load_table_from_json(
                        extracted_file1["attributes"],
                        metadata_table_id,
                        job_config=job_config,
                    )
                    load_job.result()

            return {
                "success": True,
                "status": "needs_review",
                "session_id": request.session_id,
                "extracted_filespecs": extracted_filespecs,
                "extracted_file1": extracted_file1,
            }

        else:
            raise HTTPException(
                status_code=400,
                detail="Invalid status. Must be 'approved' or 'rejected'.",
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error in review_extract_metadata: {e}")
        raise HTTPException(status_code=500, detail=str(e))
