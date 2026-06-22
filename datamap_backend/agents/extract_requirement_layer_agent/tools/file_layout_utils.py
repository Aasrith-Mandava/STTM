from __future__ import annotations

import io
import json
import logging
import os
import sys
import time
from pathlib import Path
from typing import Any

from pypdf import PdfReader, PdfWriter

from config.settings import config
from utils.doc_chunker.splitter import convert_docx_to_pdf
from utils.gcs_artifact_utils import (
    download_bytes,
    list_blobs,
    upload_json,
    upload_text,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Excel → PDF
# ---------------------------------------------------------------------------

def _excel_bytes_to_pdf(raw: bytes) -> bytes:
    """
    Convert an Excel workbook (any number of sheets) to PDF bytes.
    Each sheet becomes a section with the tab name as a bold heading.
    Multi-sheet workbooks have all sheets concatenated in order.
    """
    try:
        import openpyxl
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Excel file layout conversion requires 'openpyxl' and 'reportlab'. "
            "Install the missing dependency or upload PDF/DOCX file layouts instead."
        ) from exc

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    styles = getSampleStyleSheet()
    heading_style = styles["Heading1"]
    normal_style = styles["Normal"]

    story = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        story.append(Paragraph(sheet_name, heading_style))
        story.append(Spacer(1, 0.3 * cm))

        # Build table data — replace None with ""
        table_data = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in rows
        ]

        col_count = max(len(r) for r in table_data)
        col_width = (landscape(A4)[0] - 2 * cm) / max(col_count, 1)

        tbl = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 0.6 * cm))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=cm, rightMargin=cm,
                            topMargin=cm, bottomMargin=cm)
    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Ensure PDF bytes (DOCX / XLSX / PDF)
# ---------------------------------------------------------------------------

def _to_pdf_bytes(raw: bytes, filename: str) -> bytes:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return raw
    if ext in {".xlsx", ".xls", ".xlsm"}:
        return _excel_bytes_to_pdf(raw)
    if ext == ".docx":
        tmp_dir = Path(config.DATA_DIR) / "tmp_layout_conversion"
        tmp_dir.mkdir(parents=True, exist_ok=True)

        docx_path = tmp_dir / f"_layout_{os.urandom(4).hex()}.docx"
        pdf_path = docx_path.with_name(docx_path.stem + "_converted.pdf")

        try:
            docx_path.write_bytes(raw)

            if sys.platform == "win32":
                import pythoncom
                pythoncom.CoInitialize()

            try:
                convert_docx_to_pdf(str(docx_path))
            finally:
                if sys.platform == "win32":
                    pythoncom.CoUninitialize()

            for _ in range(3):
                if pdf_path.exists():
                    break
                time.sleep(0.5)

            if not pdf_path.exists():
                raise FileNotFoundError(f"Converted PDF not found: {pdf_path}")

            logger.info("DOCX → PDF SUCCESS | file=%s", filename)

            return pdf_path.read_bytes()

        except Exception as e:
            logger.error("DOCX → PDF FAILED | file=%s error=%s", filename, e)
            raise

        finally:
            # safer cleanup
            time.sleep(0.2)
            for p in (docx_path, pdf_path):
                try:
                    if p.exists():
                        p.unlink()
                except Exception:
                    logger.warning("Cleanup failed for %s", p)
        raise ValueError(f"Unsupported file layout format: {ext!r}")


# ---------------------------------------------------------------------------
# PDF chunking
# ---------------------------------------------------------------------------

def _chunk_pdf(pdf_bytes: bytes, chunk_size: int) -> tuple[list[bytes], int]:
    reader = PdfReader(io.BytesIO(pdf_bytes))
    total = len(reader.pages)
    chunks: list[bytes] = []
    for start in range(0, total, chunk_size):
        writer = PdfWriter()
        for i in range(start, min(start + chunk_size, total)):
            writer.add_page(reader.pages[i])
        buf = io.BytesIO()
        writer.write(buf)
        chunks.append(buf.getvalue())
    return chunks, total

# ---------------------------------------------------------------------------
# Merge chunk results
# ---------------------------------------------------------------------------

def _merge_layout_chunks(chunks_results: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Merge per-chunk extraction dicts.
    Tables with the same header key have their rows concatenated.
    """
    merged: dict[str, Any] = {}
    for chunk in chunks_results:
        if not isinstance(chunk, dict):
            logger.warning("Skipping non-dict chunk result: %s", type(chunk))
            continue
        for table_key, rows in chunk.items():
            if table_key in merged:
                if isinstance(merged[table_key], list) and isinstance(rows, list):
                    merged[table_key].extend(rows)
            else:
                merged[table_key] = rows
    return merged
