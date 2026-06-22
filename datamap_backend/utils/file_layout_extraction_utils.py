"""
File layout extraction utilities.

Pipeline for extract-file-layout/{session_id}:
  1. Fetch file_layout_* artifact from GCS for the session
  2. Convert to PDF:
       - DOCX  → PDF via convert_docx_to_pdf
       - XLSX  → PDF with each sheet rendered as a section headed by the tab name
       - PDF   → used as-is
  3. Chunk the PDF and send each chunk to the LLM to extract structured JSON
     - Each table in the PDF is keyed by its header text for 1-to-1 accuracy
  4. Merge all chunk results into a single file_layout JSON
  5. Load validated_requirement_layer.json from GCS, attach file_layout under
     "file_layout_tables", persist back as final_with_layout.json
  6. Return the merged result
"""
from __future__ import annotations

import io
import json
import logging
import os
import sys
import time
from pathlib import Path
from typing import Any

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from google.genai.types import Part
from pypdf import PdfReader, PdfWriter

from config.settings import config
from utils.brd_extraction_utils import _get_client, _create_cache, _delete_cache
from utils.doc_chunker.splitter import convert_docx_to_pdf
from utils.gcs_artifact_utils import (
    download_bytes,
    list_blobs,
    upload_json,
    upload_text,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Excel → PDF
# ---------------------------------------------------------------------------

def _excel_bytes_to_pdf(raw: bytes) -> bytes:
    """
    Convert an Excel workbook (any number of sheets) to PDF bytes.
    Each sheet becomes a section with the tab name as a bold heading.
    Multi-sheet workbooks have all sheets concatenated in order.
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    styles = getSampleStyleSheet()
    heading_style = styles["Heading1"]
    normal_style = styles["Normal"]

    story = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        story.append(Paragraph(sheet_name, heading_style))
        story.append(Spacer(1, 0.3 * cm))

        # Build table data — replace None with ""
        table_data = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in rows
        ]

        col_count = max(len(r) for r in table_data)
        col_width = (landscape(A4)[0] - 2 * cm) / max(col_count, 1)

        tbl = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 0.6 * cm))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=cm, rightMargin=cm,
                            topMargin=cm, bottomMargin=cm)
    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Ensure PDF bytes (DOCX / XLSX / PDF)
# ---------------------------------------------------------------------------

def _to_pdf_bytes(raw: bytes, filename: str) -> bytes:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return raw
    if ext in {".xlsx", ".xls", ".xlsm"}:
        return _excel_bytes_to_pdf(raw)
    if ext == ".docx":
        tmp_dir = Path(config.DATA_DIR) / "tmp_layout_conversion"
        tmp_dir.mkdir(parents=True, exist_ok=True)

        docx_path = tmp_dir / f"_layout_{os.urandom(4).hex()}.docx"
        pdf_path = docx_path.with_name(docx_path.stem + "_converted.pdf")

        try:
            docx_path.write_bytes(raw)

            if sys.platform == "win32":
                import pythoncom
                pythoncom.CoInitialize()

            try:
                convert_docx_to_pdf(str(docx_path))
            finally:
                if sys.platform == "win32":
                    pythoncom.CoUninitialize()

            for _ in range(3):
                if pdf_path.exists():
                    break
                time.sleep(0.5)

            if not pdf_path.exists():
                raise FileNotFoundError(f"Converted PDF not found: {pdf_path}")

            logger.info("DOCX → PDF SUCCESS | file=%s", filename)

            return pdf_path.read_bytes()

        except Exception as e:
            logger.error("DOCX → PDF FAILED | file=%s error=%s", filename, e)
            raise

        finally:
            # safer cleanup
            time.sleep(0.2)
            for p in (docx_path, pdf_path):
                try:
                    if p.exists():
                        p.unlink()
                except Exception:
                    logger.warning("Cleanup failed for %s", p)
        raise ValueError(f"Unsupported file layout format: {ext!r}")


# ---------------------------------------------------------------------------
# PDF chunking
# ---------------------------------------------------------------------------

def _chunk_pdf(pdf_bytes: bytes, chunk_size: int) -> tuple[list[bytes], int]:
    reader = PdfReader(io.BytesIO(pdf_bytes))
    total = len(reader.pages)
    chunks: list[bytes] = []
    for start in range(0, total, chunk_size):
        writer = PdfWriter()
        for i in range(start, min(start + chunk_size, total)):
            writer.add_page(reader.pages[i])
        buf = io.BytesIO()
        writer.write(buf)
        chunks.append(buf.getvalue())
    return chunks, total


# ---------------------------------------------------------------------------
# LLM extraction prompt
# ---------------------------------------------------------------------------

_LAYOUT_CHUNK_PROMPT = """\
You are a precise document extraction assistant specialising in file layout specifications.

Extract ALL tables from this PDF chunk (pages {page_range}) as structured JSON.

OUTPUT FORMAT — return a single JSON object where:
- Each key is the EXACT table header / section heading as it appears in the document
- Each value is an array of row objects, where each row object maps column headers to cell values
- Preserve ALL rows — do not skip or summarise any row
- Preserve exact column header text
- If a cell spans multiple columns, repeat the value under each column header
- If a section has no tabular data but has key-value pairs, represent as [{{"key": "...", "value": "..."}}]
- If a table continues from the previous chunk, use the same key and continue appending rows

Previous chunk ended with:
{handoff}

Return STRICT JSON only. No markdown fences, no commentary.
"""


# ---------------------------------------------------------------------------
# Merge chunk results
# ---------------------------------------------------------------------------

def _merge_layout_chunks(chunks_results: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Merge per-chunk extraction dicts.
    Tables with the same header key have their rows concatenated.
    """
    merged: dict[str, Any] = {}
    for chunk in chunks_results:
        for table_key, rows in chunk.items():
            if table_key in merged:
                if isinstance(merged[table_key], list) and isinstance(rows, list):
                    merged[table_key].extend(rows)
            else:
                merged[table_key] = rows
    return merged


# ---------------------------------------------------------------------------
# Main extraction function
# ---------------------------------------------------------------------------

def run_file_layout_extraction(session_id: str) -> dict[str, Any]:
    """
    1. Fetch file_layout_* artifact from GCS
    2. Convert to PDF (DOCX / Excel / PDF)
    3. Chunk PDF and extract structured JSON via LLM (table header → rows)
    4. Load validated_requirement_layer.json, attach file_layout_tables
    5. Persist final_with_layout.json to GCS
    6. Return result dict
    """
    prefix = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/uploads/"
    blobs = list_blobs(prefix=prefix)

    layout_blob = next(
        (b for b in blobs if Path(b.name).name.startswith("file_layout_")),
        None,
    )
    if layout_blob is None:
        raise FileNotFoundError(
            f"No file_layout_* artifact found for session_id={session_id!r}"
        )

    filename = Path(layout_blob.name).name
    raw = download_bytes(object_name=layout_blob.name)
    logger.info("Downloaded file layout | name=%s size=%d", filename, len(raw))

    pdf_bytes = _to_pdf_bytes(raw, filename)
    logger.info("Converted file layout to PDF | session=%s bytes=%d", session_id, len(pdf_bytes))

    chunk_size = getattr(config, "MARKDOWN_CHUNK_PAGES", 20)
    chunks, total_pages = _chunk_pdf(pdf_bytes, chunk_size)
    logger.info("Chunked file layout PDF | session=%s pages=%d chunks=%d", session_id, total_pages, len(chunks))

    client = _get_client()
    chunk_results: list[dict[str, Any]] = []
    handoff = "(start of document)"

    # Cache the static extraction prompt system text if large enough
    system_hint = (
        "You are a precise document extraction assistant for file layout specifications. "
        "Extract every table exactly as it appears, keyed by its header."
    )
    cache_name = _create_cache(client, content=system_hint, display_name=f"layout-prompt-{session_id}")

    try:
        for idx, chunk_bytes in enumerate(chunks):
            start_page = idx * chunk_size + 1
            end_page = min((idx + 1) * chunk_size, total_pages)
            page_range = f"{start_page}-{end_page}"

            prompt = _LAYOUT_CHUNK_PROMPT.format(page_range=page_range, handoff=handoff)
            pdf_part = Part.from_bytes(data=chunk_bytes, mime_type="application/pdf")

            call_config: dict[str, Any] = {
                "temperature": 0.0,
                "response_mime_type": "application/json",
                "max_output_tokens": 8192,
            }
            if cache_name:
                call_config["cached_content"] = cache_name

            last_exc: Exception | None = None
            for attempt in range(3):
                try:
                    resp = client.models.generate_content(
                        model=config.AGENT_MODEL,
                        contents=[pdf_part, prompt],
                        config=call_config,
                    )
                    text = (resp.text or "{}").strip()
                    chunk_json = json.loads(text)
                    chunk_results.append(chunk_json)
                    # Rolling handoff — last table key + row count for continuity
                    last_key = list(chunk_json.keys())[-1] if chunk_json else ""
                    handoff = f"Last table: {last_key!r}" if last_key else "(empty chunk)"
                    logger.info("Layout extraction chunk %d/%d pages=%s tables=%d",
                                idx + 1, len(chunks), page_range, len(chunk_json))
                    break
                except Exception as exc:
                    last_exc = exc
                    logger.warning("Layout chunk %d attempt %d failed: %s", idx, attempt + 1, exc)
                    time.sleep(1.5 * (attempt + 1))
            else:
                logger.error("Layout chunk %d permanently failed: %s", idx, last_exc)
                chunk_results.append({})
    finally:
        if cache_name:
            _delete_cache(client, cache_name)

    file_layout_tables = _merge_layout_chunks(chunk_results)
    logger.info("Merged file layout tables | session=%s table_count=%d", session_id, len(file_layout_tables))

    # Persist raw layout JSON to GCS
    layout_object = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/extracted_data/file_layout_tables.json"
    upload_json(object_name=layout_object, payload=file_layout_tables)

    # Load validated requirement layer and attach file_layout_tables
    validated_object = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/extracted_data/validated_requirement_layer.json"
    try:
        validated_bytes = download_bytes(object_name=validated_object)
        validated: dict[str, Any] = json.loads(validated_bytes.decode("utf-8"))
    except FileNotFoundError:
        raise FileNotFoundError(
            f"validated_requirement_layer.json not found for session_id={session_id!r}. "
            "Run GET /validate-requirement-layer first."
        )

    validated["file_layout_tables"] = file_layout_tables

    final_object = f"{config.BSA_EXTRACT_ARTIFACT_PREFIX}/{session_id}/extracted_data/final_with_layout.json"
    upload_json(object_name=final_object, payload=validated)
    gcs_uri = f"gs://{config.MAPPING_ARTIFACT_BUCKET}/{final_object}"
    logger.info("Persisted final_with_layout | uri=%s", gcs_uri)

    return {
        "session_id": session_id,
        "file_layout_filename": filename,
        "total_pages": total_pages,
        "tables_extracted": len(file_layout_tables),
        "file_layout_tables": file_layout_tables,
        "final_requirement_layer": validated,
        "gcs_output_uri": gcs_uri,
    }
