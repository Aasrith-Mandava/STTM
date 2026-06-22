# server/utils/metadata_fill_functions.py

import pandas as pd
import json
import io
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Any, List, Optional
import logging
import os

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def metadata_fill_tool(tool_input: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generates a comprehensive metadata Excel file from data dictionary and profiling outputs.

    This tool creates a SINGLE Excel workbook with MULTIPLE SHEETS - one sheet per file/table
    found in the data dictionary output. Each sheet contains detailed metadata including
    attribute names, data types, primary/foreign keys, and profiling statistics.

    The tool should be called after data_dictionary_tool and intelligent_profiling_tool have
    completed their analysis.

    Args:
        tool_input: Dictionary containing:
            - datadict_output (required): JSON string or list from data_dictionary_tool containing
              fields: File Name, Field Name, Data Type, Length, Primary Key, Foreign Key
            - profiling_output (optional): JSON from intelligent_profiling_tool for additional
              statistics like sample values, distinct counts, null counts
            - template_path (optional): Path to template Excel file, defaults to
              "template/metadata_template.xlsx"

    Returns:
        Dictionary with:
            - status: "success" or "error" indicating the operation result
            - message: Human-readable description of the result
            - excel_bytes: Base64 encoded Excel file (only on success)
            - file_count: Number of files processed (only on success)
            - sheets_created: List of sheet names created (only on success)
            - error_details: Detailed error information (only on error)
    """

    try:
        datadict_output = tool_input.get("datadict_output")
        profiling_output = tool_input.get("profiling_output", [])
        template_path = tool_input.get("template_path", "template/metadata_template.xlsx")

        # Parse datadict_output if it's a JSON string
        if isinstance(datadict_output, str):
            datadict_data = json.loads(datadict_output)
        else:
            datadict_data = datadict_output

        if not isinstance(datadict_data, list):
            return {
                "status": "error",
                "message": "Invalid datadict_output format",
                "error_details": "Expected 'datadict_output' to be a list of dictionaries."
            }

        if not datadict_data:
            return {
                "status": "error",
                "message": "Empty datadict_output provided",
                "error_details": "The 'datadict_output' list is empty. No data to process."
            }

        # Parse profiling output if needed
        if isinstance(profiling_output, str):
            profiling_data = json.loads(profiling_output)
        else:
            profiling_data = profiling_output

        # Create profiling lookup for additional stats
        profiling_lookup = _build_profiling_lookup(profiling_data)

        # Group data by file name
        files_data = {}
        for row in datadict_data:
            file_name = row.get("File Name")
            if file_name:
                if file_name not in files_data:
                    files_data[file_name] = []
                files_data[file_name].append(row)

        if not files_data:
            return {
                "status": "error",
                "message": "No files found in datadict_output",
                "error_details": "Could not extract any file names from the provided data dictionary output."
            }

        # Check if template exists
        full_template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), template_path)

        if os.path.exists(full_template_path):
            # Load existing template
            wb = load_workbook(full_template_path)
            # Remove default sheets if they exist and aren't needed
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        else:
            # Create new workbook
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        sheets_created = []

        # Create a sheet for each file
        for file_name, rows in files_data.items():
            # Create sheet with file name (Excel sheet names have max 31 chars)
            sheet_name = file_name[:31] if len(file_name) > 31 else file_name

            # Create new sheet
            ws = wb.create_sheet(title=sheet_name)
            sheets_created.append(sheet_name)

            # Apply template styling and populate data
            _populate_sheet(ws, rows, profiling_lookup, file_name)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Convert to base64 for transmission
        import base64
        excel_bytes = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')

        return {
            "status": "success",
            "message": f"Successfully generated metadata Excel file with {len(files_data)} file(s) processed into {len(sheets_created)} sheet(s)",
            "excel_bytes": excel_bytes,
            "file_count": len(files_data),
            "sheets_created": sheets_created
        }

    except json.JSONDecodeError as e:
        logger.error(f"JSON parsing error: {e}")
        return {
            "status": "error",
            "message": "Failed to parse JSON input",
            "error_details": f"JSON parsing error: {str(e)}"
        }
    except Exception as e:
        logger.error(f"Error in metadata_fill_tool: {e}", exc_info=True)
        return {
            "status": "error",
            "message": "Unexpected error while generating metadata file",
            "error_details": str(e)
        }


def _build_profiling_lookup(profiling_data: List[Dict]) -> Dict[str, Dict]:
    """
    Build a lookup dictionary from profiling data for quick access.
    Returns: {table_name: {column_name: {stats}}}
    """
    lookup = {}
    if not isinstance(profiling_data, list):
        return lookup

    for table_data in profiling_data:
        table_ref = table_data.get("table_reference", "")
        table_name = table_ref.split(".")[-1] if table_ref else ""
        column_analysis = table_data.get("column_analysis", {})

        if table_name and column_analysis:
            lookup[table_name] = column_analysis

    return lookup


def _populate_sheet(ws, rows: List[Dict], profiling_lookup: Dict, file_name: str):
    """
    Populate a worksheet with data from data dictionary rows.
    Applies formatting and styling similar to a typical metadata template.
    """

    # Define headers based on typical metadata requirements
    headers = [
        "Attribute Name",
        "Description",
        "Data Type",
        "Length/Precision",
        "Scale",
        "Nullable",
        "Primary Key",
        "Foreign Key",
        "Default Value",
        "Constraints",
        "Sample Values",
        "Distinct Count",
        "Null Count"
    ]

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    column_widths = [25, 40, 15, 15, 10, 10, 12, 12, 15, 20, 25, 12, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Get profiling data for this file
    file_profiling = profiling_lookup.get(file_name, {})

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        field_name = row_data.get("Field Name", "")
        data_type = row_data.get("Data Type", "")
        length = row_data.get("Length", "")
        primary_key = row_data.get("Primary Key", "No")
        foreign_key = row_data.get("Foreign Key", "No")

        # Get profiling stats for this column
        col_stats = file_profiling.get(field_name, {})
        sample_values = col_stats.get("sample_values", [])
        distinct_count = col_stats.get("distinct_count", "")
        null_count = col_stats.get("null_count", "")

        # Prepare row values
        row_values = [
            field_name,  # Attribute Name
            "",  # Description (to be filled by business analyst or AI)
            data_type,  # Data Type
            length if length else "",  # Length/Precision
            "",  # Scale
            "Yes" if null_count else "No",  # Nullable (inferred)
            primary_key,  # Primary Key
            foreign_key,  # Foreign Key
            "",  # Default Value
            "",  # Constraints
            ", ".join(str(v) for v in sample_values[:3]) if sample_values else "",  # Sample Values
            distinct_count,  # Distinct Count
            null_count  # Null Count
        ]

        # Write row values
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Highlight primary/foreign keys
            if col_idx == 7 and primary_key == "Yes":  # Primary Key column
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif col_idx == 8 and foreign_key != "No":  # Foreign Key column
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
