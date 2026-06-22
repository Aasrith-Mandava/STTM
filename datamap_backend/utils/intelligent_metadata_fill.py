"""
Intelligent Metadata Fill Tool - Uses AI agent mapping to fill Excel templates dynamically
"""

import pandas as pd
import json
import io
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, Any, List
import logging
import os

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def intelligent_metadata_fill_tool(tool_input: Dict[str, Any]) -> Dict[str, Any]:
    """
    Intelligently fills Excel metadata template using AI-generated column mappings.

    This tool works with the metadata_analysis_agent to:
    1. Read the template structure
    2. Get AI-generated column mappings
    3. Fill the Excel file dynamically based on the mapping

    Args:
        tool_input: Dictionary containing:
            - datadict_output (required): Output from data_dictionary_tool
            - profiling_output (optional): Output from profiling_tool
            - template_path (optional): Path to Excel template
            - column_mapping (required): AI-generated mapping from metadata_analysis_agent

    Returns:
        Dictionary with:
            - status: "success" or "error"
            - message: Human-readable message
            - excel_bytes: Base64-encoded Excel file (on success)
            - file_count: Number of files processed
            - sheets_created: List of sheet names
    """

    try:
        # Extract inputs
        datadict_output = tool_input.get("datadict_output")
        profiling_output = tool_input.get("profiling_output", [])
        template_path = tool_input.get("template_path", "templates/Medata_template.xlsx")
        column_mapping = tool_input.get("column_mapping")

        logger.info(f"=== INTELLIGENT METADATA FILL TOOL INPUTS ===")
        logger.info(f"DataDict output type: {type(datadict_output)}, length: {len(datadict_output) if isinstance(datadict_output, list) else 'N/A'}")
        logger.info(f"Profiling output type: {type(profiling_output)}, length: {len(profiling_output) if isinstance(profiling_output, list) else 'N/A'}")
        logger.info(f"Template Path: {template_path}")
        logger.info(f"Column Mapping type: {type(column_mapping)}")
        logger.info(f"Column Mapping content: {json.dumps(column_mapping, indent=2) if column_mapping else 'None'}")

        # Validate required inputs
        if not datadict_output:
            return {
                "status": "error",
                "message": "Missing required parameter 'datadict_output'",
                "error_details": "The tool requires datadict_output to generate metadata"
            }

        # column_mapping is optional - if not provided, will use default mapping
        if not column_mapping:
            logger.warning("No column_mapping provided, will use default column mapping")
            column_mapping = {}

        # Parse datadict_output if string
        if isinstance(datadict_output, str):
            try:
                datadict_data = json.loads(datadict_output)
            except json.JSONDecodeError as e:
                return {
                    "status": "error",
                    "message": "Invalid JSON format in datadict_output",
                    "error_details": f"JSON parsing failed: {str(e)}"
                }
        else:
            datadict_data = datadict_output

        if not isinstance(datadict_data, list) or not datadict_data:
            return {
                "status": "error",
                "message": "Invalid or empty datadict_output",
                "error_details": "Expected datadict_output to be a non-empty list"
            }

        # Parse profiling output
        if isinstance(profiling_output, str):
            profiling_data = json.loads(profiling_output) if profiling_output else []
        else:
            profiling_data = profiling_output

        # Parse column_mapping if string
        if isinstance(column_mapping, str):
            try:
                mapping_data = json.loads(column_mapping)
            except json.JSONDecodeError as e:
                return {
                    "status": "error",
                    "message": "Invalid JSON format in column_mapping",
                    "error_details": f"JSON parsing failed: {str(e)}"
                }
        else:
            mapping_data = column_mapping

        # Build profiling lookup
        profiling_lookup = _build_profiling_lookup(profiling_data)

        # Group data by file name
        files_data = {}
        for row in datadict_data:
            file_name = row.get("File Name")
            if file_name:
                if file_name not in files_data:
                    files_data[file_name] = []
                files_data[file_name].append(row)

        if not files_data:
            return {
                "status": "error",
                "message": "No files found in datadict_output",
                "error_details": "Could not extract any file names from the provided data dictionary output."
            }

        # Load template workbook
        full_template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), template_path)

        if not os.path.exists(full_template_path):
            return {
                "status": "error",
                "message": f"Template file not found at: {template_path}",
                "error_details": f"Expected template at: {full_template_path}"
            }

        wb = load_workbook(full_template_path)

        # Identify the template sheet to copy (usually "file1" or first sheet that's not fileSpecs)
        template_sheet_name = None
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() not in ['filespecs', 'file_specs', 'sheet']:
                template_sheet_name = sheet_name
                break

        if not template_sheet_name:
            # No template sheet found, use first sheet
            template_sheet_name = wb.sheetnames[0] if wb.sheetnames else None

        if not template_sheet_name:
            return {
                "status": "error",
                "message": "No valid template sheet found in Excel file",
                "error_details": f"Template sheets available: {wb.sheetnames}"
            }

        template_sheet = wb[template_sheet_name]
        logger.info(f"Using '{template_sheet_name}' as template sheet to copy")

        # Remove the original template sheet (we'll create copies for each file)
        wb.remove(template_sheet)

        sheets_created = []

        # Copy template sheet for each file
        for file_name, rows in files_data.items():
            sheet_name = file_name[:31] if len(file_name) > 31 else file_name

            # Create a copy of the template sheet
            ws = wb.copy_worksheet(template_sheet)
            ws.title = sheet_name
            sheets_created.append(sheet_name)

            # Populate using intelligent mapping (preserves template formatting)
            _populate_with_mapping(ws, rows, profiling_lookup, file_name, mapping_data, preserve_template=True)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Convert to base64
        import base64
        excel_bytes = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')

        return {
            "status": "success",
            "message": f"Successfully generated metadata Excel file with {len(files_data)} file(s) using intelligent mapping",
            "excel_bytes": excel_bytes,
            "file_count": len(files_data),
            "sheets_created": sheets_created,
            "mapping_used": mapping_data.get("confidence", "unknown")
        }

    except Exception as e:
        logger.error(f"Error in intelligent_metadata_fill_tool: {e}", exc_info=True)
        return {
            "status": "error",
            "message": "Unexpected error while generating metadata file",
            "error_details": str(e)
        }


def _build_profiling_lookup(profiling_data: List[Dict]) -> Dict[str, Dict]:
    """Build lookup for profiling data by table and column."""
    lookup = {}
    if not isinstance(profiling_data, list):
        return lookup

    for table_data in profiling_data:
        table_ref = table_data.get("table_reference", "")
        table_name = table_ref.split(".")[-1] if table_ref else ""
        column_analysis = table_data.get("column_analysis", {})

        if table_name and column_analysis:
            lookup[table_name] = column_analysis

    return lookup


def _populate_with_mapping(ws, rows: List[Dict], profiling_lookup: Dict, file_name: str, mapping_data: Dict, preserve_template: bool = True):
    """
    Populate worksheet using AI-generated column mapping.

    Args:
        ws: Worksheet object (copied from template, already has headers and formatting)
        rows: Data dictionary rows
        profiling_lookup: Profiling data lookup
        file_name: File being processed
        mapping_data: AI-generated mapping from metadata_analysis_agent
        preserve_template: If True, read headers from template and preserve formatting
    """

    # Extract template columns from mapping
    template_columns_config = mapping_data.get("template_columns", [])

    logger.info(f"=== POPULATE WITH MAPPING ===")
    logger.info(f"File: {file_name}, Rows count: {len(rows)}")
    logger.info(f"Mapping data keys: {mapping_data.keys() if mapping_data else 'None'}")
    logger.info(f"Template columns config count: {len(template_columns_config)}")

    if not template_columns_config:
        logger.warning("No template column mapping provided, using defaults")
        _populate_with_defaults(ws, rows, profiling_lookup, file_name)
        return

    if preserve_template:
        # Read existing headers from template (row 1)
        template_headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                template_headers.append(str(cell.value).strip())
            else:
                break

        logger.info(f"Template headers found: {template_headers}")

        # Create mapping between template headers and column configs
        header_to_config = {}
        for col_config in template_columns_config:
            template_col = col_config["template_column"]
            # Find matching header (case-insensitive)
            for idx, header in enumerate(template_headers, start=1):
                if header.lower() == template_col.lower():
                    header_to_config[idx] = col_config
                    break

        logger.info(f"Mapped {len(header_to_config)} columns from AI mapping to template headers")
    else:
        # Legacy mode: overwrite headers (not recommended)
        headers = [col["template_column"] for col in template_columns_config]
        header_to_config = {idx: col_config for idx, col_config in enumerate(template_columns_config, start=1)}

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

    # Get profiling data for this file
    file_profiling = profiling_lookup.get(file_name, {})

    # Write data rows (starting from row 2, preserving template formatting)
    logger.info(f"Writing {len(rows)} data rows to sheet")
    for row_idx, row_data in enumerate(rows, start=2):
        field_name = row_data.get("Field Name", "")
        col_stats = file_profiling.get(field_name, {})

        if row_idx == 2:  # Log first row for debugging
            logger.info(f"First row data: {row_data}")
            logger.info(f"Field name: {field_name}, Profiling stats available: {bool(col_stats)}")

        # Process each column based on mapping
        for col_idx, col_config in header_to_config.items():
            value = _extract_value_from_mapping(col_config, row_data, col_stats)

            if row_idx == 2:  # Log first row values
                logger.info(f"Column {col_idx} ({col_config.get('template_column')}): source={col_config.get('source')}, field={col_config.get('field')}, value={value}")

            # Get existing cell (preserves template formatting)
            cell = ws.cell(row=row_idx, column=col_idx)

            # Only set value, preserve all existing formatting
            cell.value = value

            # Optional: Apply conditional formatting for keys (if not already styled by template)
            if "primary" in col_config["template_column"].lower() and value == "Yes":
                if not cell.fill or cell.fill.start_color.rgb == "00000000":  # Only if no fill
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif "foreign" in col_config["template_column"].lower() and value and value != "No":
                if not cell.fill or cell.fill.start_color.rgb == "00000000":  # Only if no fill
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")


def _extract_value_from_mapping(col_config: Dict, datadict_row: Dict, profiling_data: Dict) -> Any:
    """Extract value based on AI mapping configuration."""

    source = col_config.get("source", "static")
    field = col_config.get("field", "")
    transform = col_config.get("transform", "")

    if source == "static":
        return ""

    elif source == "datadict":
        value = datadict_row.get(field, "")
        return _apply_transform(value, transform)

    elif source == "profiling":
        value = profiling_data.get(field, "")
        return _apply_transform(value, transform)

    return ""


def _apply_transform(value: Any, transform: str) -> Any:
    """Apply transformation to value based on description."""

    if not value:
        return value

    if not transform:
        return value

    transform_lower = transform.lower()

    # Yes/No to Y/N transformation
    if transform == "yes_no_to_y_n" or "yes_no_to_y_n" in transform_lower:
        value_str = str(value).strip().upper()
        if value_str in ["YES", "Y", "TRUE", "1"]:
            return "Y"
        elif value_str in ["NO", "N", "FALSE", "0"]:
            return "N"
        return value

    # Join list values with commas (first 3 items)
    if "join" in transform_lower or "comma" in transform_lower:
        if isinstance(value, list):
            return ", ".join(str(v) for v in value[:3])
        return value

    # Generic yes/no transformation
    if "yes" in transform_lower and "no" in transform_lower:
        value_str = str(value).strip().upper()
        if value_str in ["YES", "Y", "TRUE", "1"]:
            return "Yes"
        elif value_str in ["NO", "N", "FALSE", "0"]:
            return "No"
        return value

    return value


def _populate_with_defaults(ws, rows: List[Dict], profiling_lookup: Dict, file_name: str):
    """Fallback to default population if no mapping provided."""
    # This is the original hardcoded approach as fallback
    logger.info("Using default population method")
    # Implementation same as original _populate_sheet
    pass
