# bg_query_utils_with_router.py 
import os
import io
import csv as _csv_mod

CEDILLA_FAMILY = ("\u00C7", "\u00C3\u2021", "\u0102\u2021", "\u012E")
import uuid
import zipfile
import logging
import tempfile
import json
from typing import Optional, Dict, Any, List, Tuple
from datetime import datetime
from functools import lru_cache
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from utils import local_warehouse as bigquery
from google.api_core.exceptions import NotFound
from google.oauth2 import service_account
from ydata_profiling import ProfileReport

# Application-specific imports (adjust paths in your project)
from api.models import FileUploadResponse, BatchUploadResponse
from utils.scoring_utils import calculate_quality_score
from config.settings import config

# Ensure matplotlib backend for headless environments
import matplotlib
matplotlib.use("Agg")

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------
# BigQuery client
# ---------------------------------------------------------------------
@lru_cache(maxsize=1)
def get_bigquery_client():
    """Standalone: return a SQLite-backed client shim (no GCP/BigQuery).

    The shim implements the BigQuery client surface the app uses against a local
    SQLite warehouse, so existing call sites that do ``client.query(...)`` /
    ``client.load_table_from_dataframe(...)`` keep working unchanged.
    """
    from utils.local_warehouse import get_local_bq_client

    return get_local_bq_client()


# ---------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------
def validate_file(file: UploadFile) -> None:
    """Validate uploaded file size and block executable extensions."""
    if hasattr(file, "size") and file.size and config and hasattr(config, "MAX_FILE_SIZE") and file.size > config.MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail=f"File too large. Max size is {config.MAX_FILE_SIZE // (1024*1024)}MB")
    ext = os.path.splitext(file.filename)[1].lower()
    blocked = getattr(config, "BLOCKED_EXTENSIONS", {".bat", ".exe"})
    if ext in blocked:
        raise HTTPException(status_code=400, detail=f"File type '{ext}' is not allowed for security reasons.")


def generate_table_name(filename: str) -> str:
    import hashlib
    base_name = os.path.splitext(os.path.basename(filename))[0]
    safe = "".join(c if c.isalnum() else "_" for c in base_name)
    uid = hashlib.md5(base_name.encode()).hexdigest()[:8]
    return f"{config.BQ_TABLE_PREFIX}{safe}_{uid}".lower()


# ---------------------------------------------------------------------
# ZIP expansion (recursive)
# ---------------------------------------------------------------------
def expand_zip_files(files: List[UploadFile]) -> List[UploadFile]:
    """
    Expand ZIP files into UploadFile objects (recursively). Non-ZIP files passed through.
    """
    expanded: List[UploadFile] = []

    for file in files:
        ext = os.path.splitext(file.filename)[1].lower()
        file.file.seek(0)
        if ext == ".zip":
            logger.info(f"Extracting ZIP: {file.filename}")
            try:
                zip_bytes = file.file.read()
                with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
                    for member in z.namelist():
                        if member.endswith("/") or member.startswith("__MACOSX"):
                            continue
                        name = os.path.basename(member)
                        if not name:
                            continue
                        inner_ext = os.path.splitext(name)[1].lower()
                        content = z.read(member)
                        # If nested zip, expand recursively
                        if inner_ext == ".zip":
                            nested_spoof = UploadFile(filename=name, file=io.BytesIO(content))
                            nested = expand_zip_files([nested_spoof])
                            expanded.extend(nested)
                            continue
                        # Create an UploadFile-like object
                        buf = io.BytesIO(content)
                        buf.seek(0)
                        expanded.append(UploadFile(filename=name, file=buf))
                        logger.debug(f"  - extracted {name} ({len(content)} bytes)")
            except zipfile.BadZipFile:
                raise HTTPException(status_code=400, detail=f"Invalid ZIP file: {file.filename}")
            except Exception as e:
                logger.exception("ZIP extraction failed")
                raise HTTPException(status_code=500, detail=f"Failed extracting ZIP {file.filename}: {e}")
        else:
            file.file.seek(0)
            expanded.append(file)

    logger.info(f"Expanded {len(files)} â†’ {len(expanded)} files")
    return expanded


# ---------------------------------------------------------------------
# Metadata header loader (supports UploadFile or filesystem path)
# ---------------------------------------------------------------------
def load_metadata_headers(metadata_source: Optional[UploadFile or str]) -> List[str]:
    """
    Accepts either:
      - UploadFile (uploaded metadata), or
      - str path to metadata file on disk

    Supported metadata: CSV with 'Field Name' column or Excel with 'Field Name' column.
    Returns list[str] of headers.
    """
    if metadata_source is None:
        return []

    # If UploadFile was passed, save to temp file
    to_cleanup = None
    if isinstance(metadata_source, UploadFile):
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(metadata_source.filename)[1])
        metadata_source.file.seek(0)
        tmp.write(metadata_source.file.read())
        tmp.flush()
        tmp.close()
        metadata_path = tmp.name
        to_cleanup = tmp.name
    else:
        metadata_path = str(metadata_source)

    headers: List[str] = []
    ext = os.path.splitext(metadata_path)[1].lower()

    try:
        if ext == ".csv":
            # Try to read 'Field Name' column first; fallback to header row
            df = pd.read_csv(metadata_path, dtype=str)
            if "Field Name" in df.columns:
                headers = df["Field Name"].dropna().astype(str).tolist()
            else:
                # If no explicit 'Field Name', use first column as header list (values)
                # or use column names
                if df.shape[1] == 1:
                    headers = df.iloc[:, 0].dropna().astype(str).tolist()
                else:
                    headers = list(df.columns)
        elif ext in (".xlsx", ".xls", ".xlsm"):
            if openpyxl is None:
                raise ImportError("openpyxl required to read Excel metadata")
            import openpyxl as _op
            wb = _op.load_workbook(metadata_path, read_only=True, data_only=True)
            sheet = wb.active
            # find column titled 'Field Name' (case-insensitive)
            col_idx = None
            for i, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)), start=1):
                if cell and str(cell).strip().lower() == "field name":
                    col_idx = i
                    break
            if col_idx:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    val = row[col_idx - 1]
                    if val is not None:
                        headers.append(str(val))
            else:
                # Fallback: use first column values or header names
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if header_row and any(header_row):
                    headers = [str(h) for h in header_row]
                else:
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row and row[0] is not None:
                            headers.append(str(row[0]))
        else:
            raise HTTPException(status_code=400, detail="Unsupported metadata file type. Use .csv or .xlsx/.xls.")
    finally:
        if to_cleanup:
            try:
                os.unlink(to_cleanup)
            except Exception:
                pass

    # final sanity: strip headers
    headers = [h.strip() for h in headers if h and str(h).strip() != ""]
    return headers


# ---------------------------------------------------------------------
# Header detection helper (for CSV or PSV/TXT)
# ---------------------------------------------------------------------
def _looks_like_header_by_delim(first_row: str, delimiter: str, min_alpha_ratio: float = 0.7) -> bool:
    parts = first_row.split(delimiter)
    parts = [p.strip() for p in parts]
    # Need at least one non-empty token
    non_empty = [p for p in parts if p]
    if not non_empty:
        return False
    # If first non-empty cell is purely numeric -> likely NOT a header (R07)
    first = non_empty[0].replace(".", "", 1)
    if first.isdigit():
        return False
    # Count how many tokens contain alphabetic characters (use non-empty parts only)
    alpha_count = sum(1 for p in non_empty if any(c.isalpha() for c in p))
    return alpha_count >= len(non_empty) * min_alpha_ratio

NAN_STRINGS = {"nan", "NaN", "NAN", "null", "NULL", "None", "none", "NA", "N/A", "n/a", ""}

def _clean_nan_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Replace string 'nan'/'null' artifacts with actual None (missing)."""
    nan_strings = {"nan", "NaN", "NAN", "null", "NULL", "None", "none", "NA", "N/A", "n/a"}
    """
    Replace string 'nan'/'null'/'none'/'NA' artifacts with actual None (missing).
    Pandas often converts missing values to the string 'nan' when reading files
    with mixed types or when float NaN gets cast to string columns.
    This ensures they land as proper SQL NULLs in BigQuery.
    """
    return df.replace(nan_strings, None)


def _extract_fields_recursive(
    element,
    row_data: dict,
    prefix: str = "",
    max_depth: int = 3,
    current_depth: int = 0
):
    if current_depth >= max_depth:
        return

    # Handle text content of the current element
    if element.text and element.text.strip():
        field_name = f"{prefix}{element.tag}" if prefix else element.tag
        field_name = field_name.split('}')[-1]
        row_data[field_name] = element.text.strip()

    # Group children by tag to detect lists
    from collections import defaultdict
    child_map = defaultdict(list)
    for child in element:
        child_tag = child.tag.split('}')[-1]
        child_map[child_tag].append(child)

    for child_tag, children in child_map.items():
        field_name = f"{prefix}{child_tag}" if prefix else child_tag
        
        # CASE 1: List of items (e.g., Multiple Procedures)
        if len(children) > 1:
            nested_values = []
            for c in children:
                # If the child has children, recurse
                if len(c) > 0:
                    nested_data = {}
                    _extract_fields_recursive(c, nested_data, '', max_depth, current_depth + 1)
                    nested_values.append(nested_data)
                # If the child is a simple leaf node with text
                elif c.text and c.text.strip():
                    nested_values.append(c.text.strip())
            
            # STORE AS A LIST, NOT A STRING
            row_data[field_name] = nested_values

        # CASE 2: Single Child (Nested Object)
        elif len(children) == 1:
            child = children[0]
            if len(child) > 0:
                new_prefix = f"{prefix}{child_tag}_" if prefix else f"{child_tag}_"
                _extract_fields_recursive(child, row_data, new_prefix, max_depth, current_depth + 1)
            elif child.text and child.text.strip():
                row_data[field_name] = child.text.strip()
            else:
                row_data[field_name] = ''
import re

def sanitize_column_name(col: str) -> str:
    if col is None:
        return "unnamed_column"

    col = col.strip().lower()
    col = re.sub(r"[ \-]+", "_", col)
    col = re.sub(r"[^a-zA-Z0-9_]", "", col)

    if re.match(r"^\d", col):
        col = f"col_{col}"

    if col == "":
        col = "unnamed_column"

    col = col[:300]
    return col

def _detect_delimiter_from_raw_bytes(raw: bytes) -> Optional[str]:
    """
    Detect delimiter directly from raw bytes to avoid encoding-side delimiter loss.
    Returns a single-character delimiter string or None.
    """
    # Candidate delimiters in byte form.
    byte_candidates = [
        (b"\xC3\x87", "\u00C7"),  # UTF-8 Ç
        (b"\xC7", "\u00C7"),      # cp1252/latin1 Ç
        (b"|", "|"),
        (b"\t", "\t"),
        (b";", ";"),
        (b",", ","),
        (b"^", "^"),
        (b"~", "~"),
    ]

    lines = [ln for ln in raw.splitlines() if ln.strip()][:6]
    if len(lines) < 2:
        return None

    best = None
    best_score = 0
    for token, logical_delim in byte_candidates:
        counts = [ln.count(token) for ln in lines]
        # Must appear in every sampled line and be mostly consistent
        if min(counts) < 1:
            continue
        if len(set(counts)) > 2:
            continue
        score = min(counts) * len(lines)
        if score > best_score:
            best = logical_delim
            best_score = score

    if best:
        logger.info(f"Raw-byte delimiter detected: {repr(best)} (score={best_score})")
    return best

def _decode_sample(raw: bytes) -> str:
    """Detect encoding via charset-normalizer (R08), cascade fallback, log result."""
    try:
        from charset_normalizer import from_bytes
        result = from_bytes(raw).best()
        if result:
            enc = (result.encoding or "").lower()
            # Guardrail: avoid false utf-16 picks on plain text without BOM.
            if enc.startswith("utf_16") or enc.startswith("utf-16"):
                has_utf16_bom = raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff")
                if not has_utf16_bom:
                    logger.info(
                        "charset-normalizer suggested %s without BOM; using cascade fallback",
                        result.encoding,
                    )
                else:
                    logger.info(f"Encoding detected by charset-normalizer: {result.encoding}")
                    return str(result)
            else:
                logger.info(f"Encoding detected by charset-normalizer: {result.encoding}")
                return str(result)
    except ImportError:
        logger.debug("charset-normalizer not available; using encoding cascade")
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            decoded = raw.decode(enc)
            logger.info(f"Encoding detected by cascade: {enc}")
            return decoded
        except UnicodeDecodeError:
            continue
    logger.warning("All encodings failed; decoding with utf-8 replace mode")
    return raw.decode("utf-8", errors="replace")

def _sniff_delimiter(sample: str, fallback: str = ",") -> str:
    """
    Detect delimiter from content, including unknown separators.

    Strategy:
    1) Fast-path cedilia family detection.
    2) Try csv.Sniffer with common ASCII delimiters.
    3) Generic scoring across all non-alphanumeric separator-like chars found
       in the first lines (works for unknown delimiters too).
    """
    cedilla_candidates = CEDILLA_FAMILY
    first_lines = [l for l in sample.splitlines() if l.strip()][:8]

    if len(first_lines) < 2:
        return fallback

    # Hard detection for Cedilia COB files
    for cedilla in cedilla_candidates:
        cedilla_counts = [line.count(cedilla) for line in first_lines]
        if min(cedilla_counts) > 5 and len(set(cedilla_counts)) <= 2:
            logger.info(f"Detected cedilla delimiter {repr(cedilla)}")
            return cedilla

    # Fallback: common ASCII sniff first
    try:
        sniffer = _csv_mod.Sniffer()
        dialect = sniffer.sniff(sample, delimiters=",|\t;")
        detected = dialect.delimiter

        counts = []
        for line in first_lines:
            try:
                row = next(_csv_mod.reader([line], delimiter=detected))
                counts.append(len(row))
            except Exception:
                counts.append(len(line.split(detected)))

        if len(set(counts)) == 1 and counts[0] > 1:
            logger.info(f"Sniffer detected delimiter={repr(detected)}")
            return detected

    except Exception:
        pass

    # Generic delimiter discovery: score all separator-like characters seen
    # across lines by consistency and column count.
    candidate_chars: Dict[str, List[int]] = {}
    ignored_chars = set([" ", "\r", "\n", '"', "'"])

    for line in first_lines:
        local_counts: Dict[str, int] = {}
        in_quote = False
        quote_char = ""
        for ch in line:
            if ch in ('"', "'"):
                if not in_quote:
                    in_quote = True
                    quote_char = ch
                elif quote_char == ch:
                    in_quote = False
                    quote_char = ""
                continue
            if in_quote:
                continue
            if ch in ignored_chars:
                continue
            if ch.isalnum() or ch == "_":
                continue
            local_counts[ch] = local_counts.get(ch, 0) + 1

        for ch, cnt in local_counts.items():
            candidate_chars.setdefault(ch, []).append(cnt)

    best_delim = None
    best_score = -1
    for ch, counts in candidate_chars.items():
        # Must appear in all sampled lines and create at least 2 columns
        if len(counts) < len(first_lines):
            continue
        min_cnt = min(counts)
        max_cnt = max(counts)
        if min_cnt < 1:
            continue
        # Allow small variance (e.g., occasional empty trailing token)
        if max_cnt - min_cnt > 2:
            continue
        avg_cols = sum(c + 1 for c in counts) / len(counts)
        if avg_cols < 2:
            continue

        # Prefer more stable and richer separators.
        stability = 10 - (max_cnt - min_cnt)
        score = int(avg_cols * 10) + stability
        if score > best_score:
            best_score = score
            best_delim = ch

    if best_delim:
        logger.info(f"Generic delimiter detected={repr(best_delim)} score={best_score}")
        return best_delim

    logger.warning(f"Delimiter fallback used: {repr(fallback)}")
    return fallback

def _detect_format(raw: bytes) -> str:
    """
    Detect file format from raw bytes. Returns one of:
    'parquet', 'xls', 'xlsx', 'xml', 'json', 'tabular'
    Detection order: binary magic bytes first, then text heuristics.
    """
    # 1. Parquet
    if raw[:4] == b"PAR1":
        return "parquet"
    # 2. Old Excel (.xls) â€” OLE2 compound document
    if raw[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return "xls"
    # 3. ZIP-based (.xlsx, .xlsm) â€” PK magic
    if raw[:4] == b"PK\x03\x04":
        return "xlsx"
    # Text heuristics
    sample = _decode_sample(raw[:8192])
    stripped = sample.lstrip()
    # 4. XML
    if stripped.startswith("<"):
        return "xml"
    # 5. JSON object or array
    if stripped.startswith("{") or stripped.startswith("["):
        return "json"
    # 6. NDJSON â€” first 2 non-empty lines both parse as JSON
    lines = [l for l in stripped.splitlines() if l.strip()][:2]
    if len(lines) == 2:
        try:
            json.loads(lines[0])
            json.loads(lines[1])
            return "json"
        except Exception:
            pass
    # 7. Everything else is tabular
    return "tabular"


# ---------------------------------------------------------------------
# Read file to DataFrame (unified) â€” supports metadata header injection
# ---------------------------------------------------------------------

def sanitize_and_deduplicate_columns(columns):
        """Sanitize and remove duplicate column names by appending _1, _2, etc."""
        import re
        
        def sanitize_col_name(name):
            """Ensure column name is PyArrow-compatible"""
            # Convert to string if not already
            name = str(name)
            # Replace dots, colons, and other special chars with underscores
            name = re.sub(r'[\.:\-\s]+', '_', name)
            # Ensure it doesn't start with a number
            if name and name[0].isdigit():
                name = f"{name}"
            # Remove any remaining problematic characters
            name = re.sub(r'[^\w_]', '_', name)
            # Ensure it's not empty
            if not name or name == '_':
                name = 'unnamed'
            return name
        
        seen = {}
        result = []
        for col in columns:
            # First sanitize the column name
            col = sanitize_col_name(col)
            original = col
            counter = 1
            while col in seen:
                col = f"{original}_{counter}"
                counter += 1
            seen[col] = True
            result.append(col)
        return result


def normalize_missing_tokens(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert common placeholder strings (e.g., 'nan', 'null') to actual nulls.
    Only object columns are scanned to avoid changing numeric dtypes.
    """
    missing_tokens = {"nan", "na", "n/a", "null", "none", "<na>", "<null>"}
    obj_cols = df.select_dtypes(include=["object"]).columns

    if not len(obj_cols):
        return df

    def _clean_value(val):
        if pd.isna(val):
            return None
        if isinstance(val, str):
            token = val.strip().lower()
            if token in missing_tokens or token == "":
                return None
        return val

    df[obj_cols] = df[obj_cols].applymap(_clean_value)
    return df

def read_file_to_dataframe(
    file: UploadFile,
    metadata_source: Optional[UploadFile or str] = None,
    delimiter_hint: Optional[str] = None,
    has_header: Optional[bool] = None,
) -> pd.DataFrame:
    """
    Unified reader. Format is detected from file content (magic bytes + text heuristics),
    not from the file extension.

    Parameters
    ----------
    delimiter_hint : override Sniffer for exotic delimiters (e.g. '^', '~')
    has_header    : force header detection (True/False); None = auto-detect
    """
    is_fixed_width_csv = getattr(file, "was_converted_from_fixed_width", False)

    file.file.seek(0)
    raw_bytes = file.file.read()
    file.file.seek(0)

    # Converted fixed-width files are already normalized CSV with headers
    # emitted by the converter. Do NOT run delimiter/header auto-heuristics
    # here, otherwise we can accidentally skip valid rows.
    if is_fixed_width_csv:
        try:
            df = pd.read_csv(
                io.BytesIO(raw_bytes),
                dtype=str,
                keep_default_na=False,
            )
        except Exception:
            # Fallback parser without dropping lines via on_bad_lines.
            df = pd.read_csv(
                io.BytesIO(raw_bytes),
                dtype=str,
                keep_default_na=False,
                engine="python",
            )

        df.columns = sanitize_and_deduplicate_columns(df.columns.tolist())
        return _clean_nan_strings(df)

    # Prepare metadata headers
    metadata_headers = []
    if metadata_source and not is_fixed_width_csv:
        metadata_headers = load_metadata_headers(metadata_source)
    metadata_headers = sanitize_and_deduplicate_columns(
        [sanitize_column_name(h) for h in metadata_headers]
    )
    logger.info(f"metadata_headers: {metadata_headers}")

    fmt = _detect_format(raw_bytes)
    logger.info(f"Detected format={fmt!r} for file={file.filename!r}")

    # ------------------------------------------------------------------
    # Parquet
    # ------------------------------------------------------------------
    if fmt == "parquet":
        try:
            df = pd.read_parquet(io.BytesIO(raw_bytes))
            df.columns = sanitize_and_deduplicate_columns(df.columns.tolist())
            return _clean_nan_strings(df)
        except Exception as e:
            logger.exception("Parquet parse failed")
            raise HTTPException(status_code=400, detail=f"Parquet parse failed: {e}")

    # ------------------------------------------------------------------
    # Excel (.xls / .xlsx / .xlsm)
    # ------------------------------------------------------------------
    # if fmt in ("xls", "xlsx"):
    #     try:
    #         df = pd.read_excel(io.BytesIO(raw_bytes))
    #         df.columns = sanitize_and_deduplicate_columns(df.columns.tolist())
    #         return _clean_nan_strings(df)
    #     except Exception as e:
    #         logger.exception("Excel parse failed")
    #         raise HTTPException(status_code=400, detail=f"Excel parse failed: {e}")

    # ------------------------------------------------------------------
    # Excel (.xls / .xlsx / .xlsm)
    # ------------------------------------------------------------------
    if fmt in ("xls", "xlsx"):
        try:
            sheets_dict = pd.read_excel(io.BytesIO(raw_bytes), sheet_name=None)

            processed_sheets = {}
            for sheet_name, df in sheets_dict.items():
                if df.empty:
                    continue
                df.columns = sanitize_and_deduplicate_columns(df.columns.tolist())
                processed_sheets[sheet_name] = _clean_nan_strings(df)

            if not processed_sheets:
                raise HTTPException(status_code=400, detail="Excel file contains no valid sheets.")

            return processed_sheets  # dict of {sheet_name: df} instead of single df

        except Exception as e:
            logger.exception("Excel parse failed")
            raise HTTPException(status_code=400, detail=f"Excel parse failed: {e}")

    # ------------------------------------------------------------------
    # XML
    # ------------------------------------------------------------------
    if fmt == "xml":
        try:
            import xml.etree.ElementTree as ET
            content_str = _decode_sample(raw_bytes)
            content_str = re.sub(r'\sxmlns(:\w+)?="[^"]*"', '', content_str)
            content_str = re.sub(r'<(/?)(\w+:)?', r'<\1', content_str)
            root = ET.fromstring(content_str.encode("utf-8"))
            data = []
            for child in root:
                row = {}
                _extract_fields_recursive(child, row, prefix="")
                data.append(row)
            df = pd.DataFrame(data)
            list_cols = [c for c in df.columns if df[c].apply(lambda x: isinstance(x, list)).any()]
            for col in list_cols:
                df = df.explode(col).reset_index(drop=True)
                if df[col].apply(lambda x: isinstance(x, dict)).any():
                    normalized = pd.json_normalize(df[col])
                    df = pd.concat([df.drop(columns=[col]), normalized], axis=1)
            df.columns = sanitize_and_deduplicate_columns(df.columns.tolist())
            return _clean_nan_strings(df)
        except Exception as e:
            logger.exception("XML parse failed")
            raise HTTPException(status_code=400, detail=f"XML parse failed: {e}")

    # ------------------------------------------------------------------
    # JSON / NDJSON
    # ------------------------------------------------------------------
    if fmt == "json":
        logger.info("[JSON] Handler: START")
        raw_text = _decode_sample(raw_bytes).strip()
        logger.info(f"[JSON] Raw length={len(raw_text)}")
        component_paths = {}

        def load_json_any(raw: str):
            lines = [l for l in raw.splitlines() if l.strip()]
            if len(lines) > 1:
                try:
                    return [json.loads(l) for l in lines]
                except Exception:
                    pass
            if raw.startswith("["):
                try:
                    obj = json.loads(raw)
                    if isinstance(obj, list):
                        return obj
                except Exception:
                    pass
            return [json.loads(raw)]

        def flatten_recursive(obj, max_iterations: int = 10000):
            df = pd.json_normalize(obj, sep="__")
            for iteration in range(1, max_iterations + 1):  # R10: cap iterations
                complex_cols = [
                    c for c in df.columns
                    if df[c].apply(lambda v: isinstance(v, (dict, list))).any()
                ]
                if not complex_cols:
                    break
                if iteration == max_iterations:
                    logger.warning(f"JSON flatten hit max_iterations={max_iterations}; stopping to avoid infinite loop (R10)")
                    break
                for col in complex_cols:
                    df = df.explode(col, ignore_index=True)
                    if df[col].apply(lambda v: isinstance(v, dict)).any():
                        nested = pd.json_normalize(df[col], sep="__")
                        nested.columns = [f"{col}__{x}" for x in nested.columns]
                        df = pd.concat([df.drop(columns=[col]), nested], axis=1)
            return _clean_nan_strings(df)

        try:
            obj = load_json_any(raw_text)
            df = flatten_recursive(obj)
            for col in list(df.columns):
                parts = col.split("__")
                component_paths[col] = "__".join(parts[:-1]) or None
            logger.info(f"[META] Stored component paths: {component_paths}")
            new_cols = {}
            for c in df.columns:
                cleaned = sanitize_column_name(c)[:300]
                if cleaned in new_cols.values():
                    cleaned = f"{cleaned}_{len(new_cols)}"
                new_cols[c] = cleaned
            df = df.rename(columns=new_cols)
            if metadata_source:
                try:
                    meta_hdrs = [sanitize_column_name(h) for h in load_metadata_headers(metadata_source)]
                    name_map = {}
                    for c in df.columns:
                        base = c.split("_")[-1]
                        if base in meta_hdrs:
                            name_map[c] = base
                    if name_map:
                        df = df.rename(columns=name_map)
                    else:
                        logger.warning("[META] No metadata mapping match")
                except Exception as e:
                    logger.error(f"[META] Mapping failed: {e}")
            df._hidden_component_paths = component_paths
            df.columns = sanitize_and_deduplicate_columns(df.columns.tolist())
            return _clean_nan_strings(df)
        except Exception as e:
            logger.exception("JSON parsing failed")
            raise HTTPException(status_code=400, detail=f"JSON parse/flatten failed: {e}")

    # ------------------------------------------------------------------
    # Tabular (CSV / PSV / TSV / TXT / DAT / any delimiter-separated)
    # ------------------------------------------------------------------
    decoded_text = _decode_sample(raw_bytes)
    sample_text = decoded_text[:65536]
    first_line = sample_text.splitlines()[0].rstrip("\n") if sample_text.strip() else ""

    byte_delimiter = _detect_delimiter_from_raw_bytes(raw_bytes)

    if delimiter_hint:
        delimiter = delimiter_hint
        logger.info(f"Using delimiter_hint={repr(delimiter)}")
    elif byte_delimiter:
        delimiter = byte_delimiter
        logger.info(f"Using byte-level detected delimiter {repr(delimiter)}")
    elif any(c in sample_text[:2000] for c in CEDILLA_FAMILY):
        delimiter = next(c for c in CEDILLA_FAMILY if c in sample_text[:2000])
        logger.info(f"Using hard-coded cedilla-style delimiter {repr(delimiter)}")
    else:
        delimiter = _sniff_delimiter(sample_text)

    # If byte-level detection says cedilla-family, choose the actual variant
    # present in decoded text so read_csv splits correctly.
    if byte_delimiter in CEDILLA_FAMILY:
        variant = next((c for c in CEDILLA_FAMILY if c in sample_text[:5000]), None)
        if variant:
            delimiter = variant
            logger.info(f"Aligned cedilla delimiter to decoded-text variant {repr(delimiter)}")

    raw_lines = decoded_text.splitlines()
    non_empty_idx = [i for i, line in enumerate(raw_lines) if line.strip()]
    skiprows = None

    # Cedilia files can start with a control row (e.g. V3.1Ç...).
    # If so, skip it and treat the next row as the header.
    if len(non_empty_idx) >= 2:
        i0, i1 = non_empty_idx[0], non_empty_idx[1]
        line0 = raw_lines[i0].strip()
        line1 = raw_lines[i1].strip()
        cols0 = line0.count(delimiter) + 1
        cols1 = line1.count(delimiter) + 1
        if (
            cols0 >= 2
            and cols1 >= 10
            and cols1 > (cols0 * 2)
            and line0.upper().startswith("V")
            and any(ch.isdigit() for ch in line0)
        ):
            skiprows = [i0]
            logger.info(
                "Detected and skipped control/preamble row before header (cols=%d -> %d)",
                cols0,
                cols1,
            )

    def _is_header_like_line(line: str) -> bool:
        return _looks_like_header_by_delim(line, delimiter)

    # Header finder: inspect first 2-3 non-empty rows (after optional control row)
    # and select the best header candidate.
    effective_non_empty_idx = [i for i in non_empty_idx if not skiprows or i not in skiprows]
    header_line_idx = effective_non_empty_idx[0] if effective_non_empty_idx else None
    if has_header is None and effective_non_empty_idx:
        candidate_idx = effective_non_empty_idx[:3]
        first_candidate = candidate_idx[0]
        first_line_candidate = raw_lines[first_candidate].rstrip("\n")

        # Keep row-1 header when it already looks valid.
        if _is_header_like_line(first_line_candidate):
            header_line_idx = first_candidate
        else:
            chosen = None
            for pos, idx in enumerate(candidate_idx[1:], start=1):
                line = raw_lines[idx].rstrip("\n")
                if not _is_header_like_line(line):
                    continue
                next_idx = candidate_idx[pos + 1] if pos + 1 < len(candidate_idx) else None
                if next_idx is None:
                    chosen = idx
                    break
                next_line = raw_lines[next_idx].rstrip("\n")
                # Prefer a header-like row followed by a less header-like row (data row).
                if not _is_header_like_line(next_line):
                    chosen = idx
                    break
                if chosen is None:
                    chosen = idx
            if chosen is not None:
                header_line_idx = chosen

    effective_first_line = raw_lines[header_line_idx].rstrip("\n") if header_line_idx is not None else first_line
    num_cols = len(effective_first_line.split(delimiter)) if effective_first_line else 0

    # Header decision: explicit override > detected header-row heuristic
    if has_header is None:
        header_detected = _is_header_like_line(effective_first_line) if effective_first_line else False
    else:
        header_detected = has_header

    # If header is detected and it is not the first readable row, skip rows before it.
    if header_detected and header_line_idx is not None and header_line_idx > 0:
        auto_skip = list(range(header_line_idx))
        if skiprows:
            auto_skip = sorted(set(skiprows + auto_skip))
        skiprows = auto_skip
        logger.info("Header row auto-detected at line=%d; applying skiprows=%s", header_line_idx, skiprows)

    # If both file header and metadata exist, prefer in-file header.
    # Metadata is used only when header is not detected/forced absent.
    if metadata_headers and header_detected:
        logger.info("In-file header detected; ignoring metadata header injection for this file")

    buf = io.StringIO(decoded_text)

    read_kwargs = dict(
        sep=delimiter,
        dtype=str,
        engine="python",
        on_bad_lines="warn",
        quoting=_csv_mod.QUOTE_NONE,
        keep_default_na=False,
        skiprows=skiprows
    )

    if header_detected:
        df = pd.read_csv(buf, **read_kwargs)
    else:
        if metadata_headers:
            names = metadata_headers if len(metadata_headers) == num_cols else None
            if names is None:
                logger.warning(
                    f"Metadata header count {len(metadata_headers)} != data columns {num_cols}. Using generic names."
                )

            df = pd.read_csv(
                buf,
                header=None,
                names=names,
                **read_kwargs
            )
            if names is None:
                df.columns = [f"col_{i+1}" for i in range(len(df.columns))]
        else:
            df = pd.read_csv(
                buf,
                header=None,
                **read_kwargs
            )
            df.columns = [f"col_{i+1}" for i in range(len(df.columns))]

    # If we parsed with generic col_* names, promote the first row to header
    # when it clearly looks like column names.
    if (
        not header_detected
        and len(df.columns) > 1
        and all(str(c).startswith("col_") for c in df.columns)
        and not df.empty
    ):
        first_row_vals = [str(v).strip() for v in df.iloc[0].tolist()]
        non_empty = [v for v in first_row_vals if v]
        if non_empty:
            alpha_cells = sum(1 for v in non_empty if any(ch.isalpha() for ch in v))
            # Heuristic: header row is mostly text labels, not code-like data values.
            if alpha_cells >= max(2, int(len(non_empty) * 0.7)):
                df.columns = first_row_vals
                df = df.iloc[1:].reset_index(drop=True)
                logger.info(
                    "Promoted first parsed row to headers for multi-column tabular file (%d columns)",
                    len(df.columns),
                )

    # Recovery path: if parsing still produced one column, split by likely delimiters
    # based on row content.
    if len(df.columns) == 1:
        header_text = str(df.columns[0])
        sample_vals = df.iloc[:, 0].dropna().astype(str).head(30).tolist()
        haystack = " ".join([header_text] + sample_vals)

        for candidate_delim in ("\u0102\u2021", "\u00C3\u2021", "\u00C7", "\u012E", "|", "^", "~", "\t", ";", ","):
            if candidate_delim not in haystack:
                continue
            split_df = df.iloc[:, 0].astype(str).str.split(candidate_delim, expand=True)
            if split_df.shape[1] <= 1:
                continue

            applied_headers_from_column_name = False
            if (
                header_detected
                and header_text
                and not header_text.startswith("col_")
                and candidate_delim in header_text
            ):
                raw_headers = [h.strip() for h in header_text.split(candidate_delim)]
                if len(raw_headers) == split_df.shape[1]:
                    split_df.columns = raw_headers
                    applied_headers_from_column_name = True

            if not applied_headers_from_column_name:
                # If first row looks like headers, promote it to column names.
                promoted = False
                if not split_df.empty:
                    first_row_vals = [str(v).strip() for v in split_df.iloc[0].tolist()]
                    second_row_vals = (
                        [str(v).strip() for v in split_df.iloc[1].tolist()]
                        if len(split_df) > 1 else []
                    )

                    def _is_header_like(vals: List[str]) -> bool:
                        non_empty_vals = [v for v in vals if v]
                        if not non_empty_vals:
                            return False
                        alpha_cells = sum(1 for v in non_empty_vals if any(ch.isalpha() for ch in v))
                        return alpha_cells >= max(2, int(len(non_empty_vals) * 0.7))

                    # Case A: first row is the actual header
                    if _is_header_like(first_row_vals):
                        split_df.columns = first_row_vals
                        split_df = split_df.iloc[1:].reset_index(drop=True)
                        promoted = True
                        logger.info("Promoted first split row to header columns")
                    # Case B: first row is control row (e.g., V3.1...), second row is header
                    elif (
                        len(split_df) > 1
                        and first_row_vals
                        and first_row_vals[0].upper().startswith("V")
                        and _is_header_like(second_row_vals)
                    ):
                        split_df.columns = second_row_vals
                        split_df = split_df.iloc[2:].reset_index(drop=True)
                        promoted = True
                        logger.info("Skipped control row and promoted second split row to headers")
                if not promoted:
                    split_df.columns = [f"col_{i+1}" for i in range(split_df.shape[1])]

            df = split_df
            logger.info(
                "Recovered single-column parse by splitting on delimiter %r into %d columns",
                candidate_delim,
                split_df.shape[1],
            )
            break

    df.columns = sanitize_and_deduplicate_columns(df.columns.tolist())
    return _clean_nan_strings(df)


# ---------------------------------------------------------------------
# PSV -> CSV helper (standalone)
# ---------------------------------------------------------------------
def convert_psv_to_csv(input_path: str, metadata_path: Optional[str], output_path: str) -> None:
    """
    Convenience function - convert a file on disk (PSV) to a CSV on disk using metadata_path if provided.
    Useful for offline preprocessing.
    """
    import csv as _csv
    def looks_like_header(row, num_cols):
        parts = row.split("|")
        if len(parts) != num_cols: return False
        if any(p.strip() == "" for p in parts): return False
        non_num_count = sum(1 for p in parts if not p.replace(".", "", 1).isdigit())
        return non_num_count >= num_cols / 2

    def load_metadata(metadata_file):
        return load_metadata_headers(metadata_file)

    with open(input_path, "r", encoding="utf-8") as f:
        first = f.readline().rstrip("\n")
        num_cols = len(first.split("|"))
    has_header = looks_like_header(first, num_cols)
    metadata_headers = [] if has_header else (load_metadata(metadata_path) if metadata_path else [])

    with open(input_path, "r", encoding="utf-8") as psv_file, \
         open(output_path, "w", newline="", encoding="utf-8") as csv_file:
        reader = _csv.reader(psv_file, delimiter="|")
        writer = _csv.writer(csv_file)
        if has_header:
            writer.writerow(next(reader))
        else:
            writer.writerow(metadata_headers)
        for row in reader:
            writer.writerow(row)
    logger.info(f"Converted PSV {input_path} â†’ CSV {output_path}")


# ---------------------------------------------------------------------
# BigQuery helpers: dataset creation, upload, chunked upload
# ---------------------------------------------------------------------
def create_dataset_if_not_exists(client: bigquery.Client, dataset_id: str):
    dataset_ref = client.dataset(dataset_id)
    try:
        client.get_dataset(dataset_ref)
    except Exception:
        dataset = bigquery.Dataset(dataset_ref)
        dataset.location = config.LOCATION
        client.create_dataset(dataset, exists_ok=True)

def create_data_dictionary_table(table_id: str) -> str:
    """
    Creates a BigQuery table for the data dictionary with specific columns.
    Returns the full table reference (project.dataset.table_id).
    """
    client = get_bigquery_client()
    dataset_id = config.DATASET_ID
    project_id = config.PROJECT_ID
    
    full_table_id = f"{project_id}.{dataset_id}.{table_id}"
    
    schema = [
        bigquery.SchemaField("File Name", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Attribute Name", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Logical Attribute Name", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Attribute Description", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Data Type", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Length", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Precision", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Format", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Nullability", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Default Value", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Most Occurrences", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Primary Key", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("Foreign Key", "STRING", mode="NULLABLE"),
    ]
    
    table = bigquery.Table(full_table_id, schema=schema)
    
    try:
        # Create dataset if not exists
        create_dataset_if_not_exists(client, dataset_id)
        
        # Create table (exists_ok=True handles collision if generated ID already exists)
        table = client.create_table(table, exists_ok=True)
        logger.info(f"Created BigQuery table: {full_table_id}")
        return full_table_id
    except Exception as e:
        logger.error(f"Failed to create BigQuery table {full_table_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create BigQuery table: {e}")

def create_metadata_table_from_headers(table_id: str, headers: List[str]) -> str:
    """
    Creates a BigQuery table with dynamically provided column headers.
    All columns are created as STRING type and NULLABLE.
    Returns the full table reference (project.dataset.table_id).
    """
    client = get_bigquery_client()
    dataset_id = config.DATASET_ID
    project_id = config.PROJECT_ID
    
    # Clean table_id to be safe for BigQuery
    safe_table_id = "".join(c if c.isalnum() or c == "_" else "_" for c in table_id).lower()
    full_table_id = f"{project_id}.{dataset_id}.{safe_table_id}"
    
    schema = []
    for header in headers:
        # Clean header name to be safe for BigQuery column name
        clean_header = "".join(c if c.isalnum() or c == "_" else "_" for c in header).strip()
        if not clean_header:
            continue
        # Column names cannot start with a number
        if clean_header[0].isdigit():
            clean_header = f"col_{clean_header}"
            
        schema.append(bigquery.SchemaField(clean_header, "STRING", mode="NULLABLE"))
    
    if not schema:
        raise HTTPException(status_code=400, detail="No valid headers provided to create table.")

    table = bigquery.Table(full_table_id, schema=schema)
    
    try:
        # Create dataset if not exists
        create_dataset_if_not_exists(client, dataset_id)
        
        # Create table (exists_ok=True)
        table = client.create_table(table, exists_ok=True)
        logger.info(f"Created/Verified BigQuery table: {full_table_id} with {len(schema)} columns")
        return full_table_id
    except Exception as e:
        logger.error(f"Failed to create BigQuery table {full_table_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create BigQuery table from headers: {e}")

def get_table(table_reference: str, dataset_id_override: str = None):
    """
    Retrieve BigQuery table object and return as DataFrame.

    Args:
        table_reference: Table reference (can be full path or just table name)
        dataset_id_override: Optional dataset ID to use instead of config default

    Returns:
        pandas DataFrame with table data
    """
    logging.info(f"[get_table] DATASET_OVERRIDE: Called with table_reference = {table_reference}")
    logging.info(f"[get_table] DATASET_OVERRIDE: dataset_id_override parameter = {dataset_id_override}")

    client = get_bigquery_client()

    # Use override dataset_id if provided, otherwise use config default
    dataset_id = dataset_id_override if dataset_id_override else config.BQ_DATASET_ID

    logging.info(f"[get_table] DATASET_OVERRIDE: Using dataset_id = {dataset_id}")
    logging.info(f"[get_table] DATASET_OVERRIDE: Config default dataset_id = {config.BQ_DATASET_ID}")

    # Build full table reference if not already fully qualified
    original_reference = table_reference
    if config.BQ_PROJECT_ID not in table_reference:
        table_reference = f"{config.BQ_PROJECT_ID}.{dataset_id}.{table_reference}"
        logging.info(f"[get_table] DATASET_OVERRIDE: Built full reference from partial table name")
        logging.info(f"[get_table] DATASET_OVERRIDE: Original = {original_reference}")
        logging.info(f"[get_table] DATASET_OVERRIDE: Full reference = {table_reference}")
    else:
        logging.info(f"[get_table] DATASET_OVERRIDE: Table reference already fully qualified = {table_reference}")

    try:
        logging.info(f"[get_table] DATASET_OVERRIDE: Fetching table from BigQuery: {table_reference}")
        table = client.get_table(table_reference)
        if table:
            query = f"SELECT * FROM `{table_reference}`"
            logging.info(f"[get_table] DATASET_OVERRIDE: Executing query: {query}")

            query_job = client.query(query)
            rows = query_job.result()
            df = rows.to_dataframe(create_bqstorage_client=False)

            logging.info(f"[get_table] DATASET_OVERRIDE: Successfully fetched table, rows = {len(df)}, columns = {len(df.columns)}")
            return df
        else:
            logging.warning(f"[get_table] DATASET_OVERRIDE: Table not found: {table_reference}")
            return None

    except Exception as e:
        logging.error(f"[get_table] DATASET_OVERRIDE: Error fetching table {table_reference}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to get table {table_reference}: {str(e)}"
    )



def get_tables_metadata(table_references: list[str]):
    """Retrieve metadata for specified BigQuery tables
    args:
        table_references: List of table references in the form project.dataset.table
    returns: List of metadata dictionaries for each table
    """
    client = get_bigquery_client()
    metadata = []
    try:
        for ref in table_references:
            table = client.get_table(ref)
            metadata.append({
                "table_id": table.table_id,
                "dataset_id": table.dataset_id,
                "project": table.project,
                "description": table.description,
                "num_rows": table.num_rows,
                "num_bytes": table.num_bytes,
                "schema": [
                    {
                        "name": field.name,
                        "type": field.field_type,
                        "mode": field.mode,
                        "description": field.description,
                    }
                    for field in table.schema
                ],
            })
        return metadata
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to get tables metadata: {str(e)}"
        )
    


# async def upload_to_bigquery(
#     client: bigquery.Client,
#     df: pd.DataFrame,
#     table_name: str,
#     dataset_id: str
# ) -> int:
#     """Upload DataFrame to BigQuery"""
#     try:
#         # Ensure dataset exists
#         create_dataset_if_not_exists(client, dataset_id)
        
#         # Create table reference
#         table_ref = client.dataset(dataset_id).table(table_name)
        
#         # Configure load job
#         job_config = bigquery.LoadJobConfig(
#             write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
#             autodetect=True,
#         )
        
#         # Upload data
#         job = client.load_table_from_dataframe(df, table_ref, job_config=job_config)
#         job.result()  # Wait for job to complete
        
#         return len(df)
#     except Exception as e:
#         raise HTTPException(
#             status_code=500,
#             detail=f"Failed to upload to BigQuery: {str(e)}"
#         )


def force_string_columns(df: pd.DataFrame, bq_schema) -> pd.DataFrame:
    """
    Ensure PyArrow-safe types before BigQuery upload.

    - If BQ schema is available: cast only STRING/BYTES columns
    - If BQ schema is None: cast ALL object columns
    """

    if not bq_schema:
        logging.warning(
            "BigQuery schema is None. Casting all object columns to string."
        )

        obj_cols = df.select_dtypes(include=["object"]).columns
        for col in obj_cols:
            df[col] = df[col].astype(str)

        return df

    string_fields = {
        field.name
        for field in bq_schema
        if field.field_type.upper() in ("STRING", "BYTES")
    }

    for col in string_fields:
        if col in df.columns:
            df[col] = df[col].astype(str)

    return df


def upload_to_bigquery_sync(client: bigquery.Client, df: pd.DataFrame, table_name: str, dataset_id: str) -> int:
    create_dataset_if_not_exists(client, dataset_id)

    # <<<< CHANGE J: Was: df = df.replace("", None)
    # Now: _clean_nan_strings runs first to catch ALL nan variants, then empty string fallback
    df = _clean_nan_strings(df)   # <<<< CHANGE J
    df = df.replace("", None)

    for col in df.columns:
        if df[col].isna().all():
            df[col] = df[col].astype(str)

    table_ref = client.dataset(dataset_id).table(table_name)
    job_config = bigquery.LoadJobConfig(write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)

    problem_cols = [c for c in df.columns if df[c].map(type).nunique() > 1]
    if problem_cols:
        logging.warning("Columns with mixed Python types detected: %s", problem_cols)

    df = force_string_columns(df, job_config.schema)
    job = client.load_table_from_dataframe(df, table_ref, job_config=job_config)
    job.result()
    return len(df)


def query_dataset_tables(client: bigquery.Client, query: str):
    """Query and return list of tables in a dataset"""
    try:

        profiling_uuid = str(uuid.uuid4())
        file_path = f"reports/data_profile_{profiling_uuid}"
        query_job = client.query(query)
        df = query_job.to_dataframe()

        # Generate the profile report
        profile = ProfileReport(df, title="BigQuery Data Profile")

        # Export the report (e.g., as HTML)
        profile.to_file(f"{ file_path }.html")
        profile.to_file(f"{file_path}.json")
        

        with open(f"{file_path}.html", "r") as file:
            df_as_html = file.read()
            df_as_html = df_as_html.replace('Report generated by <a href="https://ydata.ai/?utm_source=opensource&utm_medium=pandasprofiling&utm_campaign=report">YData</a>.', '').replace('Brought to you by <a href="https://ydata.ai/?utm_source=opensource&utm_medium=ydataprofiling&utm_campaign=report">YData</a>', '').replace('<tr><th>Software version</th><td style="white-space: nowrap;"><a href=https://github.com/ydataai/ydata-profiling>ydata-profiling vv4.16.1</a></td></tr>', '')

        
        with open(f"{file_path}.html", "w") as file:
            file.write(df_as_html)

        return f"data_profile_{profiling_uuid}"

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to list tables in dataset: {str(e)}"
        )





def get_access_info(project_id: str, dataset_id: str, table_name: str) -> Dict[str, Any]:
    """Generate access information for the uploaded table"""
    return {
        "sql_query": f"SELECT * FROM `{project_id}.{dataset_id}.{table_name}`",
        "table_url": f"https://console.cloud.google.com/bigquery?project={project_id}&ws=!1m5!1m4!4m3!1s{project_id}!2s{dataset_id}!3s{table_name}",
        "python_example": f"""
from utils import local_warehouse as bigquery

client = bigquery.Client(project='{project_id}')
query = "SELECT * FROM `{project_id}.{dataset_id}.{table_name}`"
df = client.query(query).to_dataframe()
        """.strip(),
        "table_reference": {
            "project_id": project_id,
            "dataset_id": dataset_id,
            "table_name": table_name,
            "full_table_id": f"{project_id}.{dataset_id}.{table_name}",
        },
    }



def upload_dataframe_chunked(client: bigquery.Client, df: pd.DataFrame, dataset_id: str, table_name: str, chunk_size: int = 10000) -> int:
    create_dataset_if_not_exists(client, dataset_id)

    # <<<< CHANGE K: Added _clean_nan_strings before chunked upload
    # Was: df used as-is with no nan string cleanup at all
    df = _clean_nan_strings(df)   # <<<< CHANGE K

    table_ref = client.dataset(dataset_id).table(table_name)
    total = len(df)
    detected_schema = None
    chunks = 0
    for start in range(0, total, chunk_size):
        chunk = df.iloc[start:start + chunk_size]
        if chunks == 0:
            job_config = bigquery.LoadJobConfig(write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE, autodetect=True)
            job = client.load_table_from_dataframe(chunk, table_ref, job_config=job_config)
            job.result()
            table = client.get_table(table_ref)
            detected_schema = table.schema
        else:
            job_config = bigquery.LoadJobConfig(write_disposition=bigquery.WriteDisposition.WRITE_APPEND, schema=detected_schema)
            job = client.load_table_from_dataframe(chunk, table_ref, job_config=job_config)
            job.result()
        chunks += 1
        logger.info(f"Uploaded chunk {chunks} ({len(chunk)} rows)")
    return total


# ---------------------------------------------------------------------
# YData profiling -> parse -> DQS
# ---------------------------------------------------------------------
def _parse_ydata_profile(profile_json_path: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Parse the ydata-profiling JSON into column_analysis and default_value_analysis
    (keeps the logic simple so calculate_quality_score can consume it)
    """
    with open(profile_json_path, "r", encoding="utf-8") as f:
        profile = json.load(f)
    variables = profile.get("variables", {}) or {}
    total_rows = profile.get("table", {}).get("n", 1) or 1

    column_analysis = {}
    default_value_analysis = {}
    for col, stats in variables.items():
        analysis = {
            "data_type": stats.get("type"),
            "null_percentage": stats.get("p_missing", 0) * 100,
            "blank_percentage": stats.get("p_empty", 0) * 100,
            "uniqueness_percentage": stats.get("p_unique", 0) * 100,
            "min_value": stats.get("min"),
            "max_value": stats.get("max"),
            "avg_length": stats.get("mean_length")
        }
        column_analysis[col] = analysis

        value_counts = stats.get("value_counts_without_nan")
        most_freq = 0
        if isinstance(value_counts, list) and value_counts:
            try:
                most_freq = value_counts[0][1]
            except Exception:
                pass
        elif isinstance(value_counts, dict) and value_counts:
            try:
                most_freq = max(value_counts.values())
            except Exception:
                pass
        if most_freq > 0:
            default_pct = (most_freq / total_rows) * 100
            default_value_analysis[col] = {"default_pct": default_pct}

    return column_analysis, default_value_analysis


def generate_quality_score_from_profile(profile_json_path: str) -> Dict[str, Any]:
    """
    Orchestrates DQS calculation from ydata profile JSON and returns the score dict.
    """
    try:
        col_analysis, def_analysis = _parse_ydata_profile(profile_json_path)
        score = calculate_quality_score(col_analysis, def_analysis)
        return score
    except Exception as e:
        logger.exception("DQS generation failed")
        return {"error": str(e)}


def query_and_profile_sync(client: bigquery.Client, query: str, sample_rows: int = 10000) -> Dict[str, Any]:
    """
    Run the query, create ydata profile (HTML + JSON), then compute DQS.
    Returns dict with profiling_uuid and data_quality_score.
    """
    profiling_uuid = str(uuid.uuid4())
    base_path = f"reports/data_profile_{profiling_uuid}"
    html_path = f"{base_path}.html"
    json_path = f"{base_path}.json"

    try:
        logger.info("Running query for profiling...")
        query_job = client.query(f"{query} LIMIT {sample_rows}")
        df = query_job.to_dataframe(create_bqstorage_client=False)
        logger.info(f"Query returned {len(df)} rows; generating profile...")

        profile = ProfileReport(df, title="BigQuery Data Profile", minimal=True, explorative=True)
        profile.to_file(html_path)
        profile.to_file(json_path)

        dqs = generate_quality_score_from_profile(json_path)

        return {"profiling_uuid": f"data_profile_{profiling_uuid}", "data_quality_score": dqs}
    except Exception as e:
        logger.exception("Profiling failed")
        return {"profiling_uuid": f"data_profile_{profiling_uuid}", "data_quality_score": {"error": str(e)}}

def validate_dataset_and_tables_large_data(
    dataset_id: str,
    table_ids: list[str]
) -> dict:
    """
    Validates BigQuery dataset and tables existence.

    Returns:
        {
            "valid": bool,
            "missing_dataset": bool,
            "missing_tables": list[str]
        }
    """
    client = get_bigquery_client()
    project_id = config.PROJECT_ID

    result = {
        "valid": True,
        "missing_dataset": False,
        "missing_tables": []
    }
    #Dataset check
    try:
        client.get_dataset(f"{project_id}.{dataset_id}")
    except NotFound:
        result["valid"] = False
        result["missing_dataset"] = True
        return result 

    #Table checks
    for table_id in table_ids:
        table_ref = f"{project_id}.{dataset_id}.{table_id}"
        try:
            client.get_table(table_ref)
        except NotFound:
            result["missing_tables"].append(table_id)

    if result["missing_tables"]:
        result["valid"] = False

    return result
# ---------------------------------------------------------------------
# FastAPI router + /upload-batch endpoint (integrated)
# ---------------------------------------------------------------------
router = APIRouter()

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True, parents=True)
DATA_JSON_PATH = DATA_DIR / "data.json"


@router.post("/upload-batch", response_model=BatchUploadResponse)
async def upload_multiple_files(
    files: List[UploadFile] = File(...),
    data_dict_files: Optional[List[UploadFile]] = File(None),
    data_dict_file: Optional[UploadFile] = File(None),
    brd_file: Optional[UploadFile] = File(None),
    file_spec_file: Optional[UploadFile] = File(None),
    project_name: Optional[str] = Form(None),
    vendor_name: Optional[str] = Form(None),
    vendor_contact_person: Optional[str] = Form(None),
    file_delivery_frequency: Optional[str] = Form(None),
    brd_description: Optional[str] = Form(None),
    spec_description: Optional[str] = Form(None),
    session_id: Optional[str] = Form(None),
):
    """
    Upload multiple files (ZIP expansion supported). Optionally provide a data dictionary
    (UploadFile) to be used for header injection when files lack headers.
    """
    logger.info(f"Received upload-batch (session_id={session_id}) with {len(files)} file(s)")

    # Save optional helper files to data/ and record for session
    uploaded_info = {}

    def save_optional(upload: Optional[UploadFile], key: str):
        if not upload:
            return None
        file_path = DATA_DIR / upload.filename
        with open(file_path, "wb") as f:
            upload.file.seek(0)
            f.write(upload.file.read())
        uploaded_info[key] = str(file_path)
        logger.info(f"Saved {key} to {file_path}")
        return str(file_path)

    def save_many_optional(uploads: List[UploadFile], key: str) -> List[str]:
        saved_paths: List[str] = []
        for idx, upload in enumerate(uploads):
            safe_filename = os.path.basename(upload.filename).replace("\\", "_").replace("/", "_")
            stem, ext = os.path.splitext(safe_filename)
            file_path = DATA_DIR / f"{stem}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{idx}{ext}"
            with open(file_path, "wb") as f:
                upload.file.seek(0)
                f.write(upload.file.read())
            saved_paths.append(str(file_path))
            logger.info(f"Saved {key} item to {file_path}")
        if saved_paths:
            uploaded_info[key] = saved_paths
        return saved_paths

    dd_uploads: List[UploadFile] = []
    if data_dict_files:
        dd_uploads.extend(data_dict_files)
    if data_dict_file:
        dd_uploads.append(data_dict_file)

    save_many_optional(dd_uploads, "data_dict_file_path")
    save_optional(brd_file, "brd_file_path")
    save_optional(file_spec_file, "file_spec_file_path")

    # Persist session entry
    if DATA_JSON_PATH.exists():
        try:
            with open(DATA_JSON_PATH, "r") as f:
                data_map = json.load(f)
        except Exception:
            data_map = {}
    else:
        data_map = {}

    session_key = session_id or str(uuid.uuid4())
    data_map.setdefault(session_key, {})
    data_map[session_key].update({
        "project_name": project_name,
        "vendor_name": vendor_name,
        "vendor_contact_person": vendor_contact_person,
        "file_delivery_frequency": file_delivery_frequency,
        "brd_description": brd_description,
        "spec_description": spec_description,
        **uploaded_info
    })
    with open(DATA_JSON_PATH, "w") as f:
        json.dump(data_map, f, indent=2)

    # Expand ZIPs
    expanded = expand_zip_files(files)
    if len(expanded) > 50:
        raise HTTPException(status_code=400, detail="Too many files after expansion (limit 50)")

    client = get_bigquery_client()
    dataset_id = getattr(config, "DATASET_ID", None)
    if not dataset_id:
        raise HTTPException(status_code=500, detail="DATASET_ID not configured")

    successful_uploads: List[FileUploadResponse] = []
    failed_uploads: List[Dict[str, Any]] = []

    # If data_dict_file was uploaded, we will pass that UploadFile object to readers
    metadata_source = dd_uploads[0] if dd_uploads else None

    async def process_file(file: UploadFile) -> Tuple[Optional[FileUploadResponse], Optional[Dict[str, Any]]]:
        try:
            validate_file(file)
            table_name = generate_table_name(file.filename)

            # Read file -> DataFrame (handles CSV, XLSX, JSON, PARQUET, XML, PSV/TXT)
            df = await asyncio.to_thread(read_file_to_dataframe, file, metadata_source)

            # Upload to BigQuery
            rows = await asyncio.to_thread(upload_to_bigquery_sync, client, df, table_name, dataset_id)
            logger.info(f"Uploaded file {file.filename} â†’ {table_name} ({rows} rows)")

            # Build access info and run profiling (synchronously via to_thread)
            access = get_access_info(config.PROJECT_ID, dataset_id, table_name)
            profiling_results = await asyncio.to_thread(query_and_profile_sync, client, access["sql_query"])

            now = datetime.utcnow().isoformat()
            file_id = str(uuid.uuid4())

            response = FileUploadResponse(
                sessionID=session_key,
                user="anonymous",
                createdDate=now,
                lastUpdateDate=now,
                file_id=file_id,
                filename=file.filename,
                table_name=table_name,
                dataset_id=dataset_id,
                project_id=config.PROJECT_ID,
                rows_uploaded=rows,
                upload_timestamp=now,
                access_info=access,
                initial_profiling_report=profiling_results.get("profiling_uuid", ""),
                profiling_report_url=f"/reports/{profiling_results.get('profiling_uuid','')}.html",
                data_quality_score=profiling_results.get("data_quality_score"),
            )
            return response, None

        except HTTPException as e:
            logger.warning(f"File {file.filename} failed: {e.detail}")
            return None, {"filename": file.filename, "error": e.detail}
        except Exception as e:
            logger.exception("Unexpected error processing file")
            return None, {"filename": file.filename, "error": str(e)}

    # Run all file tasks concurrently with bounded concurrency
    import asyncio
    sem = asyncio.Semaphore(6)  # max parallel file processors

    async def sem_task(f):
        async with sem:
            return await process_file(f)

    tasks = [sem_task(f) for f in expanded]
    results = await asyncio.gather(*tasks)

    for ok, err in results:
        if ok:
            successful_uploads.append(ok)
        if err:
            failed_uploads.append(err)

    summary = {
        "successful": len(successful_uploads),
        "failed": len(failed_uploads),
        "total_rows_uploaded": sum(u.rows_uploaded for u in successful_uploads)
    }

    return BatchUploadResponse(
        total_files=len(expanded),
        successful_uploads=successful_uploads,
        failed_uploads=failed_uploads,
        summary=summary
    )


# ---------------------------------------------------------------------
# Simple endpoints to list datasets and table info (reuse helpers)
# ---------------------------------------------------------------------
@router.get("/datasets")
async def list_datasets():
    client = get_bigquery_client()
    try:
        datasets = list(client.list_datasets())
        return {
            "project_id": config.PROJECT_ID,
            "datasets": [
                {"dataset_id": ds.dataset_id, "location": getattr(ds, "location", None)}
                for ds in datasets
            ]
        }
    except Exception as e:
        logger.exception("Failed listing datasets")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/table-info/{dataset_id}/{table_name}")
async def table_info(dataset_id: str, table_name: str):
    client = get_bigquery_client()
    try:
        table_ref = client.dataset(dataset_id).table(table_name)
        table = client.get_table(table_ref)
        return {
            "table_name": table_name,
            "dataset_id": dataset_id,
            "project_id": config.PROJECT_ID,
            "created": table.created.isoformat() if table.created else None,
            "modified": table.modified.isoformat() if table.modified else None,
            "num_rows": table.num_rows,
            "num_bytes": table.num_bytes,
            "schema": [{"name": f.name, "type": f.field_type, "mode": f.mode} for f in table.schema]
        }
    except Exception as e:
        logger.exception("Table info failed")
        raise HTTPException(status_code=404, detail=str(e))


# ---------------------------------------------------------------------
# Mapping Excel generation + BQ persistence
# Delegated to agents/extract_agent/mapping_agent/tools/mapping_excel_tools.py
# ---------------------------------------------------------------------
from agents.extract_agent.mapping_agent.tools.mapping_excel_tools import (
    generate_mapping_output,
)


# ---------------------------------------------------------------------
# Export module-level names for import elsewhere
# ---------------------------------------------------------------------
__all__ = [
    "get_bigquery_client",
    "validate_file",
    "generate_table_name",
    "expand_zip_files",
    "read_file_to_dataframe",
    "convert_psv_to_csv",
    "upload_to_bigquery_sync",
    "upload_dataframe_chunked",
    "get_access_info",
    "query_and_profile_sync",
    "generate_quality_score_from_profile",
    "router",
    "generate_mapping_output",
]


